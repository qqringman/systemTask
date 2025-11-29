#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Task Dashboard v23 - Windows Outlook Integration
新功能：
1. 成員超期圖表加寬，解決文字重疊
2. 任務列表、成員統計、貢獻度 新增下拉篩選（模組、負責人、Due Date、優先級、超期、狀態）
3. HTML 匯出完全重寫，與原頁面一致
4. Review 模式 - 無統計規則時顯示一般 mail 列表
"""

import re
import os
import io
import json
import tempfile
from datetime import datetime, timedelta
from collections import defaultdict
from dataclasses import dataclass, field
from typing import List, Dict, Optional, Set

from flask import Flask, render_template_string, request, jsonify, send_file, Response

# Windows Outlook
try:
    import win32com.client
    import pythoncom
    HAS_OUTLOOK = True
except ImportError:
    HAS_OUTLOOK = False
    print("⚠️ pywin32 未安裝，Outlook 功能停用")

# Excel
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False

# .msg 解析
try:
    import extract_msg
    HAS_EXTRACT_MSG = True
except ImportError:
    HAS_EXTRACT_MSG = False

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024

# 全域變數
FOLDER_TREE = []
FOLDERS = {}
OUTLOOK_OK = False
LAST_RESULT = None
LAST_DATA = None
LAST_MAILS_LIST = []  # 儲存郵件列表供匯出用
PRIORITY_WEIGHTS = {'high': 3, 'medium': 2, 'normal': 1}

# 儲存 mail 內容
MAIL_CONTENTS = {}
# 儲存 mail 的 entry_id（用於下載附件）
MAIL_ENTRIES = {}

@dataclass
class Task:
    title: str
    owners: List[str]
    priority: str = "normal"
    due_date: Optional[str] = None
    status: Optional[str] = None
    mail_date: str = ""
    mail_subject: str = ""
    module: str = ""
    mail_id: str = ""
    has_attachments: bool = False
    attachments: list = field(default_factory=list)

@dataclass
class TaskTracker:
    title: str
    owners: List[str]
    priority: str = "normal"
    due_date: Optional[str] = None
    status: Optional[str] = None
    first_seen: str = ""
    last_seen: str = ""
    appearances: List[str] = field(default_factory=list)
    in_last_mail: bool = False
    has_attachments: bool = False
    attachments: list = field(default_factory=list)
    
    def days_spent(self) -> int:
        if not self.first_seen or not self.last_seen:
            return 0
        try:
            d1 = datetime.strptime(self.first_seen, "%Y-%m-%d")
            d2 = datetime.strptime(self.last_seen, "%Y-%m-%d")
            return (d2 - d1).days + 1
        except:
            return 0

# ===== Outlook 功能 =====
def load_folders():
    global FOLDER_TREE, FOLDERS, OUTLOOK_OK
    if not HAS_OUTLOOK:
        print("❌ Outlook 功能不可用")
        return
    
    try:
        pythoncom.CoInitialize()
        outlook = win32com.client.Dispatch("Outlook.Application")
        namespace = outlook.GetNamespace("MAPI")
        
        tree = []
        folders = {}
        
        def scan(folder, parent_list, level=0):
            if level > 5:
                return
            try:
                entry_id = folder.EntryID
                store_id = folder.StoreID
                name = folder.Name
                
                folders[entry_id] = {"name": name, "store_id": store_id}
                
                node = {"name": name, "entry_id": entry_id, "store_id": store_id, "children": []}
                parent_list.append(node)
                
                for i in range(1, folder.Folders.Count + 1):
                    scan(folder.Folders.Item(i), node["children"], level + 1)
            except Exception as e:
                pass
        
        for i in range(1, namespace.Folders.Count + 1):
            scan(namespace.Folders.Item(i), tree)
        
        FOLDER_TREE = tree
        FOLDERS = folders
        OUTLOOK_OK = True
        print(f"    ✅ 共載入 {len(folders)} 個資料夾")
    except Exception as e:
        print(f"❌ Outlook 連接失敗: {e}")
        OUTLOOK_OK = False

def get_messages(entry_id, store_id, start_date, end_date, exclude_after_5pm: bool = True):
    global MAIL_ENTRIES
    pythoncom.CoInitialize()
    outlook = win32com.client.Dispatch("Outlook.Application")
    namespace = outlook.GetNamespace("MAPI")
    
    folder = namespace.GetFolderFromID(entry_id, store_id)
    items = folder.Items
    items.Sort("[ReceivedTime]", True)
    
    start_dt = datetime.strptime(start_date, "%Y-%m-%d")
    end_dt = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)
    
    try:
        items = items.Restrict(f"[ReceivedTime] >= '{start_dt.strftime('%m/%d/%Y')}' AND [ReceivedTime] < '{end_dt.strftime('%m/%d/%Y')}'")
    except:
        pass
    
    messages = []
    for item in items:
        try:
            rt = item.ReceivedTime
            if hasattr(rt, 'date') and not (start_dt.date() <= rt.date() < end_dt.date()):
                continue
            
            if exclude_after_5pm and hasattr(rt, 'hour'):
                if rt.hour >= 17:
                    continue
            
            html_body = ""
            try:
                html_body = item.HTMLBody or ""
            except:
                pass
            
            # 檢查是否有附件並取得附件資訊
            has_attachments = False
            attachments_info = []
            try:
                if hasattr(item, 'Attachments') and item.Attachments.Count > 0:
                    has_attachments = True
                    for j in range(1, item.Attachments.Count + 1):
                        try:
                            att = item.Attachments.Item(j)
                            attachments_info.append({
                                "index": j,
                                "name": att.FileName if hasattr(att, 'FileName') else f"attachment_{j}",
                                "size": att.Size if hasattr(att, 'Size') else 0
                            })
                        except:
                            pass
            except:
                pass
            
            # 生成 mail_id
            import hashlib
            mail_id = hashlib.md5(f"{rt.strftime('%Y-%m-%d') if hasattr(rt, 'strftime') else ''}_{rt.strftime('%H:%M') if hasattr(rt, 'strftime') else ''}_{item.Subject or ''}".encode()).hexdigest()[:12]
            
            # 儲存 entry_id 供附件下載用
            try:
                item_entry_id = item.EntryID
                MAIL_ENTRIES[mail_id] = {
                    'entry_id': item_entry_id,
                    'store_id': store_id
                }
            except:
                pass
            
            messages.append({
                "subject": item.Subject or "", 
                "body": item.Body or "",
                "html_body": html_body,
                "date": rt.strftime("%Y-%m-%d") if hasattr(rt, 'strftime') else "",
                "time": rt.strftime("%H:%M") if hasattr(rt, 'strftime') else "",
                "sender": str(item.SenderName) if hasattr(item, 'SenderName') else "",
                "has_attachments": has_attachments,
                "attachments": attachments_info,
                "mail_id": mail_id
            })
        except:
            continue
    
    return messages

# ===== 任務解析 =====
class TaskParser:
    def __init__(self, exclude_middle_priority: bool = True):
        self.tasks: List[Task] = []
        self.current_module = ""
        self.exclude_middle_priority = exclude_middle_priority
        self.stop_parsing = False
    
    def _is_valid_module(self, bracket_content: str) -> bool:
        """檢查是否是有效的模組標題"""
        inner = bracket_content.strip('[]').strip()
        inner_lower = inner.lower()
        
        # 排除狀態標記
        invalid_patterns = [
            r'^\s*status\s*:', r'^\s*due\s*:', r'^\s*duedate\s*:',
            r'^\s*pending\s*$', r'^\s*resolved\s*$', r'^\s*done\s*$',
            r'^\s*completed\s*$', r'^\s*in\s*progress\s*$'
        ]
        for pattern in invalid_patterns:
            if re.match(pattern, inner_lower, re.IGNORECASE):
                return False
        
        # 排除日期格式 [20250821], [2025/08/21], [08/21], [8/21] 等
        if re.match(r'^\d{8}$', inner):  # YYYYMMDD
            return False
        if re.match(r'^\d{4}[-/]\d{1,2}[-/]\d{1,2}$', inner):  # YYYY-MM-DD or YYYY/MM/DD
            return False
        if re.match(r'^\d{1,2}[-/]\d{1,2}$', inner):  # MM/DD or M/D
            return False
        if re.match(r'^\d{1,2}[-/]\d{1,2}[-/]\d{2,4}$', inner):  # MM/DD/YY or MM/DD/YYYY
            return False
        
        return True
    
    def _is_middle_priority_marker(self, line: str) -> bool:
        line_lower = line.lower().strip()
        return 'middle priority' in line_lower or 'low priority' in line_lower
    
    def parse(self, subject: str, body: str, mail_date: str = "", mail_time: str = "", html_body: str = "", has_attachments: bool = False, attachments: list = None, mail_id: str = None):
        import hashlib
        if not mail_id:
            mail_id = hashlib.md5(f"{mail_date}_{mail_time}_{subject}".encode()).hexdigest()[:12]
        
        MAIL_CONTENTS[mail_id] = {
            "subject": subject,
            "body": body,
            "html_body": html_body,
            "date": mail_date,
            "time": mail_time,
            "attachments": attachments or []
        }
        
        # 記錄此郵件是否有附件
        self._current_has_attachments = has_attachments
        self._current_attachments = attachments or []
        self._current_mail_id = mail_id
        
        if '<html' in body.lower() or '<' in body:
            body = re.sub(r'<style[^>]*>.*?</style>', '', body, flags=re.DOTALL | re.IGNORECASE)
            body = re.sub(r'<[^>]+>', '\n', body)
            body = re.sub(r'&nbsp;', ' ', body)
            body = re.sub(r'&[a-z]+;', ' ', body)
        
        self.current_module = ""
        self.stop_parsing = False
        
        for line in body.split('\n'):
            line = line.strip()
            
            if self.exclude_middle_priority and self._is_middle_priority_marker(line):
                self.stop_parsing = True
                break
            
            module_match = re.match(r'^(\[[^\]]+\](?:\[[^\]]+\])*)\s*$', line)
            if module_match:
                potential_module = module_match.group(1)
                first_bracket = re.match(r'^(\[[^\]]+\])', potential_module)
                if first_bracket and self._is_valid_module(first_bracket.group(1)):
                    self.current_module = potential_module
                continue
            
            match = re.match(r'^(\d+)[.\)、]\s*(.+)$', line)
            if match:
                content = match.group(2).strip()
                task = self._parse_task(content, mail_date, subject)
                if task:
                    task.module = self.current_module
                    task.mail_id = mail_id
                    task.has_attachments = has_attachments
                    task.attachments = self._current_attachments
                    self.tasks.append(task)
    
    def _parse_task(self, content: str, mail_date: str = "", mail_subject: str = "") -> Optional[Task]:
        priority = "normal"
        star_match = re.match(r'^(\*{1,3})\s*(.+)$', content)
        if star_match:
            stars = len(star_match.group(1))
            content = star_match.group(2).strip()
            priority = {3: "high", 2: "medium", 1: "normal"}.get(stars, "normal")
        
        due_match = re.search(r'\[Due\s*(?:date)?[:\s]*([^\]]+)\]', content, re.IGNORECASE)
        if not due_match:
            due_match = re.search(r'\[(\d{1,2}/\d{1,2})\]', content)
        if not due_match:
            return None
        
        due_date = due_match.group(1).strip()
        # 清理 due_date，移除多餘的前綴
        due_date = re.sub(r'^(?:date)?[:\s]*', '', due_date, flags=re.IGNORECASE).strip()
        content = content[:due_match.start()] + content[due_match.end():]
        
        status = None
        status_match = re.search(r'\[(pending|resolved|done|completed|status[:\s]*[^\]]+)\]', content, re.IGNORECASE)
        if status_match:
            status = status_match.group(1).strip().lower()
            if ':' in status:
                status = status.split(':')[-1].strip()
            content = content[:status_match.start()] + content[status_match.end():]
        
        parts = re.split(r'\s*[-–—]\s*', content, maxsplit=1)
        if len(parts) < 2:
            parts = re.split(r'\s+', content, maxsplit=1)
        
        if len(parts) < 2:
            return None
        
        task_name = parts[0].strip()
        members_str = parts[1].strip() if len(parts) > 1 else ""
        members_str = re.sub(r'\[.*?\]', '', members_str).strip()
        
        owners = self._parse_members(members_str)
        if not owners:
            return None
        
        task_name = re.sub(r'\s+', ' ', task_name).strip()
        task_name = re.sub(r'\[.*?\]', '', task_name).strip()
        
        if not task_name or len(task_name) < 2:
            return None
        
        return Task(title=task_name, owners=owners, priority=priority, due_date=due_date, status=status, mail_date=mail_date, mail_subject=mail_subject)
    
    def _parse_members(self, text: str) -> List[str]:
        if not text:
            return []
        parts = re.split(r'[/,、]', text)
        members = []
        for p in parts:
            p = p.strip()
            if not p:
                continue
            if re.match(r'^[\u4e00-\u9fff]{1,10}$', p):
                members.append(p)
            elif re.match(r'^[A-Za-z][A-Za-z0-9_]{0,19}$', p):
                members.append(p)
        return members


# ===== 統計類別 =====
class Stats:
    def __init__(self):
        self.raw_tasks: List[Dict] = []
        self.unique_members: Set[str] = set()
        self.last_mail_date: str = ""
    
    def _task_key(self, title: str, due: str, owners: List[str]) -> str:
        return f"{title.strip().lower()}|{due}|{','.join(sorted(owners))}"
    
    def add(self, task: Task):
        self.raw_tasks.append({
            "title": task.title,
            "owners": task.owners,
            "owners_str": "/".join(task.owners),
            "priority": task.priority,
            "due": task.due_date,
            "status": task.status or "-",
            "mail_date": task.mail_date,
            "mail_subject": task.mail_subject,
            "mail_id": task.mail_id,
            "module": task.module or "",
            "has_attachments": task.has_attachments,
            "attachments": task.attachments if hasattr(task, 'attachments') else [],
            "_key": self._task_key(task.title, task.due_date, task.owners)
        })
        
        for owner in task.owners:
            self.unique_members.add(owner)
        
        if task.mail_date > self.last_mail_date:
            self.last_mail_date = task.mail_date
    
    def _process_tasks(self) -> List[Dict]:
        if not self.raw_tasks:
            return []
        
        tasks_by_date = defaultdict(list)
        for t in self.raw_tasks:
            tasks_by_date[t["mail_date"]].append(t)
        
        sorted_dates = sorted(tasks_by_date.keys())
        task_tracker = {}
        final_tasks = []
        prev_date_keys = set()
        
        for date_idx, mail_date in enumerate(sorted_dates):
            day_tasks = tasks_by_date[mail_date]
            day_task_map = {}
            for t in day_tasks:
                key = t["_key"]
                if key not in day_task_map:
                    day_task_map[key] = t
                else:
                    existing = day_task_map[key]
                    priority_order = {"high": 3, "medium": 2, "normal": 1}
                    if priority_order.get(t["priority"], 0) > priority_order.get(existing["priority"], 0):
                        day_task_map[key] = t
            
            current_date_keys = set(day_task_map.keys())
            
            for key in prev_date_keys:
                if key not in current_date_keys and key in task_tracker and task_tracker[key]["active"]:
                    tracker = task_tracker[key]
                    task_data = tracker["task_data"].copy()
                    prev_date = sorted_dates[date_idx - 1] if date_idx > 0 else mail_date
                    task_data["first_seen"] = tracker["first_seen"]
                    task_data["last_seen"] = prev_date
                    task_data["completed_date"] = prev_date
                    task_data["task_status"] = "completed"
                    task_data["overdue_days"] = self._calc_overdue_days_v2(task_data["due"], tracker["first_seen"], prev_date)
                    task_data["days_spent"] = self._calc_days_between(tracker["first_seen"], prev_date)
                    final_tasks.append(task_data)
                    task_tracker[key]["active"] = False
            
            for key, task_data in day_task_map.items():
                if key not in task_tracker or not task_tracker[key]["active"]:
                    task_tracker[key] = {"first_seen": mail_date, "task_data": task_data, "active": True}
                else:
                    task_tracker[key]["task_data"] = task_data
            
            prev_date_keys = current_date_keys
        
        last_date = sorted_dates[-1] if sorted_dates else ""
        today = datetime.now().strftime("%Y-%m-%d")
        
        for key, tracker in task_tracker.items():
            if tracker["active"]:
                task_data = tracker["task_data"].copy()
                task_data["first_seen"] = tracker["first_seen"]
                task_data["last_seen"] = last_date
                
                status_val = task_data.get("status", "-").lower()
                if status_val in ["pending", "hold", "blocked"]:
                    task_data["task_status"] = "pending"
                else:
                    task_data["task_status"] = "in_progress"
                
                task_data["overdue_days"] = self._calc_overdue_days_v2(task_data["due"], tracker["first_seen"], today)
                task_data["days_spent"] = self._calc_days_between(tracker["first_seen"], last_date)
                final_tasks.append(task_data)
        
        return final_tasks
    
    def _calc_overdue_days_v2(self, due_str: str, first_seen: str, end_date: str) -> int:
        if not due_str or not end_date:
            return 0
        try:
            due_str = due_str.replace('/', '-').strip()
            parts = due_str.split('-')
            if len(parts) == 2:
                month, day = int(parts[0]), int(parts[1])
                first_year = int(first_seen[:4]) if first_seen else datetime.now().year
                due_date = datetime(first_year, month, day)
                first_dt = datetime.strptime(first_seen, "%Y-%m-%d") if first_seen else datetime.now()
                if due_date < first_dt - timedelta(days=180):
                    due_date = datetime(first_year + 1, month, day)
            elif len(parts) == 3:
                year, month, day = int(parts[0]), int(parts[1]), int(parts[2])
                if year < 100:
                    year += 2000
                due_date = datetime(year, month, day)
            else:
                return 0
            
            end_dt = datetime.strptime(end_date, "%Y-%m-%d")
            diff = (end_dt - due_date).days
            return max(0, diff)
        except:
            return 0
    
    def _calc_days_between(self, start: str, end: str) -> int:
        try:
            d1 = datetime.strptime(start, "%Y-%m-%d")
            d2 = datetime.strptime(end, "%Y-%m-%d")
            return (d2 - d1).days + 1
        except:
            return 0
    
    def summary(self):
        all_tasks = self._process_tasks()
        total_tasks = len(all_tasks)
        
        completed_count = sum(1 for t in all_tasks if t.get("task_status") == "completed")
        pending_count = sum(1 for t in all_tasks if t.get("task_status") == "pending")
        in_progress_count = sum(1 for t in all_tasks if t.get("task_status") == "in_progress")
        
        active_tasks = [t for t in all_tasks if t.get("task_status") != "completed"]
        overdue_count = sum(1 for t in active_tasks if t.get("overdue_days", 0) > 0)
        not_overdue_count = len(active_tasks) - overdue_count
        
        for t in all_tasks:
            t["is_overdue"] = t.get("overdue_days", 0) > 0 and t.get("task_status") != "completed"
        
        sorted_tasks = sorted(all_tasks, key=lambda x: (x.get("last_seen", "") or "", x.get("due", "") or ""), reverse=True)
        
        members = []
        overdue_by_member = {}
        contribution = []
        
        for n in sorted(self.unique_members):
            m_tasks = [t for t in all_tasks if n in t.get("owners", [])]
            high_count = sum(1 for t in m_tasks if t["priority"] == "high")
            med_count = sum(1 for t in m_tasks if t["priority"] == "medium")
            nor_count = sum(1 for t in m_tasks if t["priority"] == "normal")
            
            members.append({
                "name": n,
                "total": len(m_tasks),
                "completed": sum(1 for t in m_tasks if t.get("task_status") == "completed"),
                "pending": sum(1 for t in m_tasks if t.get("task_status") == "pending"),
                "in_progress": sum(1 for t in m_tasks if t.get("task_status") == "in_progress"),
                "high": high_count, "medium": med_count, "normal": nor_count
            })
            
            task_count = len(m_tasks)
            weighted_score = high_count * 3 + med_count * 2 + nor_count * 1
            
            overdue_tasks = [t for t in m_tasks if t.get("overdue_days", 0) > 0]
            overdue_task_count = len(overdue_tasks)
            total_overdue_days = sum(t.get("overdue_days", 0) for t in overdue_tasks)
            avg_overdue_days = total_overdue_days / overdue_task_count if overdue_task_count > 0 else 0
            
            completed_overdue_tasks = [t for t in overdue_tasks if t.get("task_status") == "completed"]
            active_overdue_tasks = [t for t in overdue_tasks if t.get("task_status") != "completed"]
            completed_overdue_days = sum(t.get("overdue_days", 0) for t in completed_overdue_tasks)
            active_overdue_days = sum(t.get("overdue_days", 0) for t in active_overdue_tasks)
            
            overdue_penalty = 0
            if task_count > 0:
                overdue_rate = overdue_task_count / task_count
                overdue_penalty += overdue_task_count * 0.5
                if avg_overdue_days > 7:
                    overdue_penalty += avg_overdue_days / 7
                if overdue_rate > 0.3:
                    overdue_penalty += 2
            
            final_score = max(0, weighted_score - overdue_penalty)
            
            overdue_by_member[n] = {
                "overdue_count": overdue_task_count,
                "total_overdue_days": total_overdue_days,
                "avg_overdue_days": round(avg_overdue_days, 1),
                "overdue_rate": round(overdue_task_count / task_count * 100, 1) if task_count > 0 else 0
            }
            
            contribution.append({
                "name": n,
                "task_count": task_count,
                "high": high_count, "medium": med_count, "normal": nor_count,
                "base_score": weighted_score,
                "overdue_count": overdue_task_count,
                "overdue_days": total_overdue_days,
                "completed_overdue_days": completed_overdue_days,
                "active_overdue_days": active_overdue_days,
                "overdue_penalty": round(overdue_penalty, 1),
                "score": round(final_score, 1)
            })
        
        contribution.sort(key=lambda x: -x["score"])
        for i, c in enumerate(contribution):
            c["rank"] = i + 1
        
        priority_counts = {"high": 0, "medium": 0, "normal": 0}
        for task in all_tasks:
            priority_counts[task["priority"]] += 1
        
        module_stats = defaultdict(int)
        for task in all_tasks:
            module = task.get("module", "") or "未分類"
            module_stats[module] += 1
        
        # 取得所有唯一值用於篩選下拉
        all_modules = sorted(set(t.get("module", "") or "未分類" for t in all_tasks))
        all_owners = sorted(self.unique_members)
        all_dues = sorted(set(t.get("due", "") for t in all_tasks if t.get("due")))
        
        return {
            "total_tasks": total_tasks, 
            "total_members": len(self.unique_members),
            "completed_count": completed_count,
            "pending_count": pending_count,
            "in_progress_count": in_progress_count,
            "overdue_count": overdue_count,
            "not_overdue_count": not_overdue_count,
            "last_mail_date": self.last_mail_date,
            "priority_counts": priority_counts,
            "module_stats": dict(module_stats),
            "overdue_by_member": overdue_by_member,
            "members": members, 
            "all_tasks": sorted_tasks,
            "member_list": all_owners,
            "module_list": all_modules,
            "due_list": all_dues,
            "contribution": contribution
        }
    
    def excel(self):
        wb = Workbook()
        hfill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        hfont = Font(bold=True, color="FFFFFF")
        redfont = Font(color="FF0000", bold=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        summary = self.summary()
        
        ws1 = wb.active
        ws1.title = "總覽"
        ws1['A1'] = "Task Dashboard Report"
        ws1['A1'].font = Font(bold=True, size=14)
        ws1['A3'] = "報表日期"; ws1['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M")
        ws1['A4'] = "最後郵件"; ws1['B4'] = summary["last_mail_date"]
        ws1['A6'] = "總任務數"; ws1['B6'] = summary["total_tasks"]
        ws1['A7'] = "已完成"; ws1['B7'] = summary["completed_count"]
        ws1['A8'] = "進行中"; ws1['B8'] = summary["in_progress_count"]
        ws1['A9'] = "Pending"; ws1['B9'] = summary["pending_count"]
        ws1['A10'] = "超期"; ws1['B10'] = summary["overdue_count"]
        ws1['A11'] = "成員數"; ws1['B11'] = summary["total_members"]
        
        ws2 = wb.create_sheet("任務明細")
        headers2 = ["模組", "任務", "負責人", "優先級", "Due Date", "超期天數", "狀態", "任務狀態", "首次出現", "最後出現", "花費天數"]
        for i, h in enumerate(headers2, 1):
            c = ws2.cell(1, i, h); c.fill, c.font, c.border = hfill, hfont, border
        for r, t in enumerate(summary["all_tasks"], 2):
            overdue_days = t.get("overdue_days", 0)
            status_map = {"completed": "已完成", "pending": "Pending", "in_progress": "進行中"}
            for i, v in enumerate([
                t.get("module", ""), t["title"], t["owners_str"], t["priority"],
                t.get("due", ""), overdue_days, t.get("status", "-"),
                status_map.get(t.get("task_status", ""), t.get("task_status", "")),
                t.get("first_seen", ""), t.get("last_seen", ""), t.get("days_spent", 0)
            ], 1):
                cell = ws2.cell(r, i, v)
                cell.border = border
                if i == 5 and overdue_days > 0:
                    cell.font = redfont
                if i == 6 and overdue_days > 0:
                    cell.font = redfont
        
        ws3 = wb.create_sheet("貢獻度排名")
        headers3 = ["排名", "成員", "任務數", "基礎分", "超期任務數", "總超期天數", "扣分", "總分"]
        for i, h in enumerate(headers3, 1):
            c = ws3.cell(1, i, h); c.fill, c.font, c.border = hfill, hfont, border
        for r, c in enumerate(summary["contribution"], 2):
            for i, v in enumerate([c["rank"], c["name"], c["task_count"], c["base_score"], c["overdue_count"], c["overdue_days"], c["overdue_penalty"], c["score"]], 1):
                cell = ws3.cell(r, i, v)
                cell.border = border
                if i in [5, 6, 7] and v > 0:
                    cell.font = redfont
        
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf


HTML = '''
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>System Task Dashboard</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
    <link href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.10.0/font/bootstrap-icons.css" rel="stylesheet">
    <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
    <style>
        :root { --primary: #2E75B6; --primary-dark: #1a4f7a; }
        body { background: #f5f7fa; font-size: 14px; }
        .navbar { background: #2E75B6; }
        .card { border: none; border-radius: 10px; box-shadow: 0 2px 12px rgba(0,0,0,0.08); margin-bottom: 12px; }
        .card-header { background: #2E75B6; color: white; border-radius: 10px 10px 0 0 !important; padding: 8px 12px; display: flex; justify-content: space-between; align-items: center; }
        .card-header-title { font-weight: 500; }
        .stat-card { text-align: center; padding: 10px; cursor: pointer; transition: all 0.2s; height: 85px; display: flex; flex-direction: column; justify-content: center; }
        .stat-card:hover { transform: translateY(-2px); box-shadow: 0 4px 15px rgba(0,0,0,0.15); }
        .stat-number { font-size: 1.5rem; font-weight: bold; color: var(--primary); }
        .stat-number.danger { color: #dc3545; }
        .stat-number.warning { color: #FFA500; }
        .stat-number.success { color: #28a745; }
        .stat-number.info { color: #17a2b8; }
        .stat-label { color: #666; font-size: 0.7rem; }
        .badge-high { background: #FF6B6B !important; }
        .badge-medium { background: #FFE066 !important; color: #333 !important; }
        .badge-normal { background: #74C0FC !important; }
        .badge-completed { background: #28a745 !important; }
        .badge-pending { background: #FFA500 !important; }
        .badge-in_progress { background: #17a2b8 !important; }
        .config-ok { background: #d4edda; color: #155724; padding: 6px 12px; border-radius: 6px; margin-bottom: 8px; font-size: 0.8rem; }
        .drop-zone { border: 2px dashed #dee2e6; border-radius: 6px; padding: 15px; text-align: center; cursor: pointer; }
        .drop-zone.dragover { border-color: var(--primary); background: #f0f7ff; }
        .loading { position: fixed; top: 0; left: 0; width: 100%; height: 100%; background: rgba(255,255,255,0.9); display: flex; justify-content: center; align-items: center; z-index: 9999; }
        .tree-box { max-height: 150px; overflow-y: auto; border: 1px solid #dee2e6; border-radius: 6px; padding: 5px; background: #fff; font-size: 0.75rem; }
        .tree ul { list-style: none; padding-left: 12px; margin: 0; }
        .tree > ul { padding-left: 0; }
        .tree-toggle { cursor: pointer; color: #666; }
        .tree-toggle::before { content: "▶ "; font-size: 7px; }
        .tree-toggle.open::before { content: "▼ "; }
        .tree-item { cursor: pointer; padding: 1px 4px; border-radius: 3px; }
        .tree-item:hover { background: #e8f4fc; }
        .tree-item.selected { background: var(--primary); color: white; }
        .tree-item::before { content: "📁 "; }
        
        /* 結果頁籤樣式 */
        #resultTabs .nav-link { color: #666; background: #f8f9fa; border: 1px solid #dee2e6; border-bottom: none; margin-right: 2px; }
        #resultTabs .nav-link:hover { color: #2E75B6; background: #e8f4fc; }
        #resultTabs .nav-link.active { color: #fff; background: #2E75B6; border-color: #2E75B6; }
        
        .data-table { width: 100%; font-size: 0.8rem; border-collapse: collapse; table-layout: auto; }
        .data-table thead th { background: #4a4a4a !important; color: white !important; font-weight: 600; cursor: pointer; padding: 8px 5px; white-space: nowrap; border: 1px solid #666; }
        .data-table thead th:hover { background: #333 !important; }
        .data-table tbody td { padding: 6px 5px; vertical-align: middle; border: 1px solid #ddd; }
        .data-table tbody tr { cursor: pointer; }
        .data-table tbody tr:nth-child(even) { background: #f9f9f9; }
        .data-table tbody tr:hover { background: #e8f4fc !important; }
        .data-table tbody tr.row-pending { background: #fff8e1; }
        .data-table tbody tr.row-in_progress { background: #e3f2fd; }
        .data-table tbody tr.row-overdue { background: #ffebee; }
        .table-toolbar { display: flex; gap: 8px; padding: 8px 10px; background: #f8f9fa; border-bottom: 1px solid #dee2e6; flex-wrap: wrap; align-items: center; }
        .table-toolbar input, .table-toolbar select { font-size: 0.75rem; }
        .table-toolbar select { min-width: 90px; }
        .table-container { overflow-x: auto; max-height: 400px; }
        .text-overdue { color: #dc3545 !important; font-weight: bold; }
        
        .pagination-controls { display: flex; justify-content: space-between; align-items: center; padding: 8px 10px; background: #f8f9fa; border-top: 1px solid #dee2e6; font-size: 0.75rem; }
        .pagination-controls button { padding: 3px 10px; font-size: 0.75rem; }
        .pagination-controls select { font-size: 0.75rem; padding: 2px 5px; width: 70px; }
        
        .footer { text-align: center; padding: 12px; color: #999; font-size: 0.7rem; border-top: 1px solid #eee; margin-top: 10px; }
        .rank-badge { display: inline-block; width: 22px; height: 22px; line-height: 22px; border-radius: 50%; text-align: center; font-weight: bold; color: white; font-size: 0.7rem; }
        .rank-1 { background: linear-gradient(135deg, #FFD700, #FFA500); }
        .rank-2 { background: linear-gradient(135deg, #C0C0C0, #A0A0A0); }
        .rank-3 { background: linear-gradient(135deg, #CD7F32, #8B4513); }
        .rank-other { background: #6c757d; }
        .progress { height: 18px; }
        .chart-container { height: 280px; }
        .chart-select { font-size: 0.75rem; padding: 3px 8px; width: 80px; }
        
        /* Review 模式樣式 */
        .mail-item { padding: 10px; border-bottom: 1px solid #eee; cursor: pointer; }
        .mail-item:hover { background: #f8f9fa; }
        .mail-item.selected { background: #e3f2fd; }
        .mail-subject { font-weight: 500; }
        .mail-meta { font-size: 0.75rem; color: #666; }
        .mail-preview { max-height: 60vh; overflow-y: auto; padding: 0; background: #fff; }
        
        /* 頁籤樣式 */
        .nav-tabs { border-bottom: none; }
        .nav-tabs .nav-link { color: rgba(255,255,255,0.7); border: none; padding: 8px 16px; margin-right: 4px; border-radius: 6px 6px 0 0; background: rgba(255,255,255,0.1); }
        .nav-tabs .nav-link:hover { color: white; background: rgba(255,255,255,0.2); }
        .nav-tabs .nav-link.active { color: #333; background: white; font-weight: 500; }
        
        /* 最大化功能 */
        .card-maximize-btn { cursor: pointer; opacity: 0.7; font-size: 0.8rem; }
        .card-maximize-btn:hover { opacity: 1; }
        .card-fullscreen { position: fixed !important; top: 0 !important; left: 0 !important; width: 100vw !important; height: 100vh !important; max-height: 100vh !important; z-index: 9999; border-radius: 0 !important; margin: 0 !important; display: flex !important; flex-direction: column !important; }
        .card-fullscreen > .card-header { flex-shrink: 0 !important; }
        .card-fullscreen > .card-body, .card-fullscreen .card-body { flex: 1 !important; height: 0 !important; min-height: 0 !important; max-height: none !important; overflow: hidden !important; display: flex !important; flex-direction: column !important; }
        .card-fullscreen .table-container { flex: 1 !important; height: 0 !important; min-height: 0 !important; max-height: none !important; }
        .card-fullscreen .chart-container { flex: 1 !important; height: 0 !important; min-height: 0 !important; }
        .card-fullscreen #mailList { flex: 1 !important; height: 0 !important; min-height: 0 !important; max-height: none !important; }
        /* 郵件內容卡片最大化 - 關鍵：移除 max-height 限制 */
        .card-fullscreen #mailHeader { flex-shrink: 0 !important; }
        .card-fullscreen #mailContentHtml { flex: 1 !important; min-height: 0 !important; max-height: none !important; overflow: hidden !important; position: relative !important; }
        .card-fullscreen #mailContentText { flex: 1 !important; min-height: 0 !important; max-height: none !important; overflow-y: auto !important; }
        .card-fullscreen #mailIframe { position: absolute !important; top: 0 !important; left: 0 !important; width: 100% !important; height: 100% !important; border: none !important; }
        .card-fullscreen .mail-preview { max-height: none !important; }
        .fullscreen-overlay { position: fixed; top: 0; left: 0; width: 100vw; height: 100vh; background: rgba(0,0,0,0.5); z-index: 9998; display: none; }
    </style>
</head>
<body>
    <!-- 全螢幕遮罩 -->
    <div id="fullscreenOverlay" class="fullscreen-overlay" onclick="exitFullscreen()"></div>
    
    <nav class="navbar navbar-dark mb-2 py-1">
        <div class="container-fluid">
            <span class="navbar-brand mb-0 h6"><i class="bi bi-clipboard-data me-2"></i>System Task Dashboard</span>
            <div class="d-flex gap-2">
                <button class="btn btn-outline-light btn-sm" onclick="exportExcel()"><i class="bi bi-file-excel me-1"></i>Excel</button>
                <button class="btn btn-outline-light btn-sm" onclick="exportHTML()"><i class="bi bi-filetype-html me-1"></i>HTML</button>
            </div>
        </div>
    </nav>

    <div class="container-fluid">
        <div id="loading" class="loading" style="display:none;">
            <div class="text-center"><div class="spinner-border text-primary"></div><div class="mt-2">分析中...</div></div>
        </div>

        <!-- 設定區 - 頁籤 -->
        <div class="row g-2 mb-2">
            <div class="col-md-12">
                <div class="card">
                    <div class="card-header py-2">
                        <ul class="nav nav-tabs card-header-tabs" role="tablist">
                            <li class="nav-item">
                                <button class="nav-link active" data-bs-toggle="tab" data-bs-target="#tabOutlook" type="button">
                                    <i class="bi bi-envelope me-1"></i>Outlook 資料夾
                                </button>
                            </li>
                            <li class="nav-item">
                                <button class="nav-link" data-bs-toggle="tab" data-bs-target="#tabUpload" type="button">
                                    <i class="bi bi-cloud-upload me-1"></i>上傳 .msg 檔案
                                </button>
                            </li>
                        </ul>
                    </div>
                    <div class="card-body py-2">
                        <div class="tab-content">
                            <!-- Outlook 資料夾頁籤 -->
                            <div class="tab-pane fade show active" id="tabOutlook" role="tabpanel">
                                {% if fc > 0 %}<div class="config-ok"><i class="bi bi-check-circle me-1"></i>Outlook 已連接 ({{ fc }} 資料夾)</div>{% endif %}
                                <div class="row g-2">
                                    <div class="col-md-3">
                                        <label class="form-label small mb-1">選擇 Outlook 資料夾</label>
                                        <div class="tree-box" id="folderTree">
                                            <div class="tree" id="tree"></div>
                                        </div>
                                        <div class="small text-muted mt-1">已選: <span id="selectedFolder" style="color:#2E75B6;font-weight:600;">-</span></div>
                                    </div>
                                    <div class="col-md-9">
                                        <!-- 日期和篩選 -->
                                        <div class="row g-2 mb-2">
                                            <div class="col-md-2">
                                                <label class="form-label small mb-1">開始日期</label>
                                                <input type="date" class="form-control form-control-sm" id="startDate">
                                            </div>
                                            <div class="col-md-2">
                                                <label class="form-label small mb-1">結束日期</label>
                                                <input type="date" class="form-control form-control-sm" id="endDate">
                                            </div>
                                            <div class="col-md-4">
                                                <label class="form-label small mb-1">進階篩選</label>
                                                <div class="input-group input-group-sm">
                                                    <select class="form-select form-select-sm" id="filterField" style="max-width:100px;">
                                                        <option value="">全部</option>
                                                        <option value="subject">主旨</option>
                                                        <option value="sender">寄件者</option>
                                                        <option value="recipient">收件者</option>
                                                        <option value="body">內容</option>
                                                    </select>
                                                    <input type="text" class="form-control form-control-sm" id="filterKeyword" placeholder="關鍵字...">
                                                    <select class="form-select form-select-sm" id="filterAttType" style="max-width:90px;" title="附件類型篩選">
                                                        <option value="">附件</option>
                                                        <option value="any">有附件</option>
                                                        <option value="xlsx">Excel</option>
                                                        <option value="docx">Word</option>
                                                        <option value="pptx">PPT</option>
                                                        <option value="pdf">PDF</option>
                                                        <option value="image">圖片</option>
                                                        <option value="zip">壓縮檔</option>
                                                        <option value="txt">文字檔</option>
                                                    </select>
                                                </div>
                                            </div>
                                            <div class="col-md-4">
                                                <label class="form-label small mb-1">&nbsp;</label>
                                                <div class="d-flex gap-1">
                                                    <button class="btn btn-outline-secondary btn-sm" onclick="toggleFilterSettings()"><i class="bi bi-gear me-1"></i>分析選項</button>
                                                    <button class="btn btn-primary btn-sm" id="btnAnalyze" onclick="analyze()"><i class="bi bi-bar-chart me-1"></i>統計分析</button>
                                                    <button class="btn btn-outline-primary btn-sm" id="btnReview" onclick="loadReviewMode()"><i class="bi bi-envelope-open me-1"></i>Review</button>
                                                </div>
                                            </div>
                                        </div>
                                        <!-- 分析選項 -->
                                        <div id="filterSettings" class="p-2 bg-light rounded" style="display:none;">
                                            <div class="row g-2">
                                                <div class="col-auto">
                                                    <div class="form-check form-check-inline">
                                                        <input class="form-check-input" type="checkbox" id="excludeMiddlePriority" checked>
                                                        <label class="form-check-label small">排除 Middle priority 以下</label>
                                                    </div>
                                                </div>
                                                <div class="col-auto">
                                                    <div class="form-check form-check-inline">
                                                        <input class="form-check-input" type="checkbox" id="excludeAfter5pm" checked>
                                                        <label class="form-check-label small">排除下午 5:00 後</label>
                                                    </div>
                                                </div>
                                            </div>
                                        </div>
                                    </div>
                                </div>
                            </div>
                            <!-- 上傳 .msg 檔案頁籤 -->
                            <div class="tab-pane fade" id="tabUpload" role="tabpanel">
                                <div class="row g-2">
                                    <div class="col-md-6">
                                        <div class="drop-zone py-4" id="dropZone">
                                            <i class="bi bi-cloud-upload fs-1 text-muted"></i>
                                            <div class="mt-2">拖放 .msg 檔案到此處，或點擊選擇檔案</div>
                                            <input type="file" id="fileInput" multiple accept=".msg" style="display:none;">
                                        </div>
                                        <div id="uploadFileList" class="mt-2 small"></div>
                                    </div>
                                    <div class="col-md-6">
                                        <div class="mb-2">
                                            <label class="form-label small mb-1">篩選設定</label>
                                            <div class="form-check">
                                                <input class="form-check-input" type="checkbox" id="uploadExcludeMiddlePriority" checked>
                                                <label class="form-check-label small">排除 Middle priority 以下</label>
                                            </div>
                                            <div class="form-check">
                                                <input class="form-check-input" type="checkbox" id="uploadExcludeAfter5pm" checked>
                                                <label class="form-check-label small">排除下午 5:00 後</label>
                                            </div>
                                        </div>
                                        <button class="btn btn-primary btn-sm" id="btnUploadAnalyze" onclick="analyzeUploadedFiles()"><i class="bi bi-search me-1"></i>分析上傳檔案</button>
                                    </div>
                                </div>
                            </div>
                        </div>
                    </div>
                </div>
            </div>
        </div>

        <!-- 結果區域 - 頁籤結構 -->
        <div id="resultArea" style="display:none;">
            <ul class="nav nav-tabs mb-2" id="resultTabs" role="tablist" style="border-bottom: 2px solid #2E75B6;">
                <li class="nav-item" id="tabItem-stats">
                    <button class="nav-link active" id="tab-stats" data-bs-toggle="tab" data-bs-target="#pane-stats" type="button" style="font-weight:600;">
                        <i class="bi bi-bar-chart me-1"></i>統計分析
                    </button>
                </li>
                <li class="nav-item" id="tabItem-review">
                    <button class="nav-link" id="tab-review" data-bs-toggle="tab" data-bs-target="#pane-review" type="button" style="font-weight:600;">
                        <i class="bi bi-eye me-1"></i>Review <span id="reviewMailCount" class="badge bg-warning text-dark ms-1">0</span>
                    </button>
                </li>
            </ul>
            <div class="tab-content">
                <!-- Review 頁籤 -->
                <div class="tab-pane fade" id="pane-review" role="tabpanel">
                    <div class="row g-2">
                        <div class="col-md-4">
                            <div class="card" id="cardMailList" style="height:600px;">
                                <div class="card-header">
                                    <span class="card-header-title"><i class="bi bi-envelope me-1"></i>郵件列表</span>
                                    <div class="d-flex align-items-center">
                                        <span id="reviewMailCountDetail" class="badge bg-info text-white me-2" style="font-size:0.7rem;"></span>
                                        <i class="bi bi-arrows-fullscreen card-maximize-btn text-white" onclick="toggleFullscreen('cardMailList')" title="最大化/還原"></i>
                                    </div>
                                </div>
                                <div class="table-toolbar">
                                    <input type="text" class="form-control form-control-sm" placeholder="🔍 搜尋主旨/寄件者..." id="mailSearch" onkeyup="filterMailList()">
                                </div>
                                <div id="mailList" style="flex:1;overflow-y:auto;" onscroll="onMailListScroll(event)"></div>
                            </div>
                        </div>
                        <div class="col-md-8">
                                            <div class="card" id="cardMailContent" style="height:600px;display:flex;flex-direction:column;">
                                                <div class="card-header" style="flex-shrink:0;">
                                                    <span class="card-header-title"><i class="bi bi-file-text me-1"></i>郵件內容</span>
                                                    <div class="d-flex align-items-center">
                                                        <div class="btn-group btn-group-sm me-2">
                                                            <button class="btn btn-outline-light btn-sm active" id="btnMailHtml" onclick="setMailViewMode('html')">HTML</button>
                                                            <button class="btn btn-outline-light btn-sm" id="btnMailText" onclick="setMailViewMode('text')">純文字</button>
                                                        </div>
                                                        <i class="bi bi-arrows-fullscreen card-maximize-btn text-white" onclick="toggleFullscreen('cardMailContent')" title="最大化/還原"></i>
                                                    </div>
                                                </div>
                                                <div class="card-body p-0" style="display:flex;flex-direction:column;flex:1 1 auto;overflow:hidden;min-height:0;">
                                                    <div id="mailHeader" class="p-2 bg-light border-bottom" style="display:none;flex-shrink:0;">
                                                        <div><strong>主旨:</strong> <span id="mailSubjectView">-</span></div>
                                                        <div><strong>日期:</strong> <span id="mailDateView">-</span></div>
                                                        <div id="mailAttachmentsRow" style="display:none;"><strong>附件:</strong> <span id="mailAttachmentsList"></span></div>
                                                    </div>
                                                    <div id="mailContentHtml" class="mail-preview" style="flex:1 1 auto;overflow:hidden;min-height:0;position:relative;"><iframe id="mailIframe" style="position:absolute;top:0;left:0;width:100%;height:100%;border:none;"></iframe></div>
                                                    <div id="mailContentText" class="mail-preview" style="display:none;flex:1 1 auto;overflow-y:auto;font-family:monospace;white-space:pre-wrap;padding:15px;min-height:0;"></div>
                                                </div>
                                            </div>
                                        </div>
                    </div>
                </div>
                
                <!-- 統計分析頁籤 -->
                <div class="tab-pane fade show active" id="pane-stats" role="tabpanel">
            <!-- 統計卡片 -->
            <div class="row g-2 mb-2">
                <div class="col"><div class="card stat-card" onclick="showAllTasks()"><div class="stat-number" id="totalTasks">0</div><div class="stat-label">總任務</div></div></div>
                <div class="col"><div class="card stat-card" onclick="showByStatus('pending')"><div class="stat-number warning" id="pendingCount">0</div><div class="stat-label">Pending</div></div></div>
                <div class="col"><div class="card stat-card" onclick="showByStatus('in_progress')"><div class="stat-number info" id="inProgressCount">0</div><div class="stat-label">進行中</div></div></div>
                <div class="col"><div class="card stat-card" onclick="showByStatus('completed')"><div class="stat-number success" id="completedCount">0</div><div class="stat-label">已完成</div></div></div>
                <div class="col"><div class="card stat-card" onclick="showOverdue()"><div class="stat-number danger" id="overdueCount">0</div><div class="stat-label">超期</div></div></div>
            </div>

            <!-- 進度條 -->
            <div class="card mb-2">
                <div class="card-body py-2">
                    <div class="d-flex justify-content-between small mb-1">
                        <strong>任務進度</strong>
                        <span>最後郵件: <span id="lastMailDate">-</span></span>
                    </div>
                    <div class="progress">
                        <div class="progress-bar bg-success" id="completedBar" title="已完成"></div>
                        <div class="progress-bar bg-info" id="inProgressBar" title="進行中"></div>
                        <div class="progress-bar bg-warning" id="pendingBar" title="Pending"></div>
                    </div>
                </div>
            </div>

            <!-- 圖表區 - 2x2 佈局讓圖表更寬 -->
            <div class="row g-2 mb-2">
                <div class="col-md-6">
                    <div class="card" id="cardChart1">
                        <div class="card-header">
                            <span class="card-header-title"><i class="bi bi-pie-chart me-1"></i>狀態分佈</span>
                            <div class="d-flex align-items-center">
                                <select class="form-select chart-select me-2" style="width:90px" id="chart1Type" onchange="updateChart1()">
                                    <option value="doughnut">環形</option><option value="pie">圓餅</option><option value="bar">長條</option><option value="polarArea">極區</option>
                                </select>
                                <i class="bi bi-arrows-fullscreen card-maximize-btn text-white" onclick="toggleFullscreen('cardChart1')" title="最大化"></i>
                            </div>
                        </div>
                        <div class="card-body py-2"><div class="chart-container"><canvas id="chart1"></canvas></div></div>
                    </div>
                </div>
                <div class="col-md-6">
                    <div class="card" id="cardChart4">
                        <div class="card-header">
                            <span class="card-header-title"><i class="bi bi-person-exclamation me-1"></i>成員超期天數</span>
                            <div class="d-flex align-items-center">
                                <select class="form-select chart-select me-2" style="width:120px" id="chart4Type" onchange="updateChart4()">
                                    <option value="stacked" selected>水平堆疊</option><option value="vstacked">垂直堆疊</option><option value="line">折線圖</option>
                                </select>
                                <i class="bi bi-arrows-fullscreen card-maximize-btn text-white" onclick="toggleFullscreen('cardChart4')" title="最大化"></i>
                            </div>
                        </div>
                        <div class="card-body py-2"><div class="chart-container"><canvas id="chart4"></canvas></div></div>
                    </div>
                </div>
            </div>
            <div class="row g-2 mb-2">
                <div class="col-md-6">
                    <div class="card" id="cardChart2">
                        <div class="card-header">
                            <span class="card-header-title"><i class="bi bi-bar-chart me-1"></i>優先級分佈</span>
                            <div class="d-flex align-items-center">
                                <select class="form-select chart-select me-2" style="width:90px" id="chart2Type" onchange="updateChart2()">
                                    <option value="doughnut">環形</option><option value="pie">圓餅</option><option value="bar">長條</option><option value="polarArea">極區</option>
                                </select>
                                <i class="bi bi-arrows-fullscreen card-maximize-btn text-white" onclick="toggleFullscreen('cardChart2')" title="最大化"></i>
                            </div>
                        </div>
                        <div class="card-body py-2"><div class="chart-container"><canvas id="chart2"></canvas></div></div>
                    </div>
                </div>
                <div class="col-md-6">
                    <div class="card" id="cardChart3">
                        <div class="card-header">
                            <span class="card-header-title"><i class="bi bi-exclamation-triangle me-1"></i>超期狀況</span>
                            <div class="d-flex align-items-center">
                                <select class="form-select chart-select me-2" style="width:90px" id="chart3Type" onchange="updateChart3()">
                                    <option value="doughnut">環形</option><option value="pie">圓餅</option><option value="bar">長條</option><option value="polarArea">極區</option>
                                </select>
                                <i class="bi bi-arrows-fullscreen card-maximize-btn text-white" onclick="toggleFullscreen('cardChart3')" title="最大化"></i>
                            </div>
                        </div>
                        <div class="card-body py-2"><div class="chart-container"><canvas id="chart3"></canvas></div></div>
                    </div>
                </div>
            </div>

            <!-- 任務列表 -->
            <div class="card mb-2" id="cardTaskList">
                <div class="card-header">
                    <span class="card-header-title"><i class="bi bi-list-task me-1"></i>任務列表</span>
                    <div class="d-flex align-items-center">
                        <button class="btn btn-outline-light btn-sm me-1" onclick="toggleTaskFilter()"><i class="bi bi-funnel me-1"></i>篩選</button>
                        <button class="btn btn-outline-light btn-sm me-2" onclick="exportTableCSV('task')"><i class="bi bi-download me-1"></i>CSV</button>
                        <i class="bi bi-arrows-fullscreen card-maximize-btn text-white" onclick="toggleFullscreen('cardTaskList')" title="最大化"></i>
                    </div>
                </div>
                <div class="table-toolbar" id="taskFilterBar" style="display:none;">
                    <input type="text" class="form-control form-control-sm" style="width:150px" placeholder="🔍 搜尋..." id="taskSearch" onkeyup="filterAndRenderTaskTable()">
                    <select class="form-select form-select-sm" style="width:130px" id="filterModule" onchange="filterAndRenderTaskTable()"><option value="">全部模組</option></select>
                    <select class="form-select form-select-sm" style="width:130px" id="filterOwner" onchange="filterAndRenderTaskTable()"><option value="">全部負責人</option></select>
                    <select class="form-select form-select-sm" style="width:110px" id="filterPriority" onchange="filterAndRenderTaskTable()">
                        <option value="">全部優先</option><option value="high">High</option><option value="medium">Medium</option><option value="normal">Normal</option>
                    </select>
                    <select class="form-select form-select-sm" style="width:110px" id="filterStatus" onchange="filterAndRenderTaskTable()">
                        <option value="">全部狀態</option><option value="in_progress">進行中</option><option value="pending">Pending</option><option value="completed">已完成</option>
                    </select>
                    <select class="form-select form-select-sm" style="width:110px" id="filterOverdue" onchange="filterAndRenderTaskTable()">
                        <option value="">全部超期</option><option value="yes">超期</option><option value="no">未超期</option>
                    </select>
                    <button class="btn btn-outline-secondary btn-sm" onclick="clearTaskFilters()"><i class="bi bi-x-circle"></i> 清除</button>
                </div>
                <div class="table-container">
                    <table class="table table-sm data-table mb-0">
                        <thead>
                            <tr>
                                <th onclick="sortTable('task','last_seen')">Mail日期 ↕</th>
                                <th onclick="sortTable('task','module')">模組 ↕</th>
                                <th onclick="sortTable('task','title')">任務 ↕</th>
                                <th onclick="sortTable('task','owners_str')">負責人 ↕</th>
                                <th onclick="sortTable('task','priority')">優先級 ↕</th>
                                <th onclick="sortTable('task','due')">Due ↕</th>
                                <th onclick="sortTable('task','overdue_days')">超期 ↕</th>
                                <th onclick="sortTable('task','task_status')">狀態 ↕</th>
                            </tr>
                        </thead>
                        <tbody id="taskTableBody"></tbody>
                    </table>
                </div>
                <div class="pagination-controls">
                    <div>
                        <button class="btn btn-outline-secondary btn-sm" onclick="prevPage('task')">上一頁</button>
                        <select class="form-select form-select-sm d-inline-block ms-1" id="taskPageSize" onchange="renderTaskTable()">
                            <option value="30">30</option><option value="50" selected>50</option><option value="100">100</option><option value="200">200</option>
                        </select>
                    </div>
                    <span id="taskPageInfo">-</span>
                    <button class="btn btn-outline-secondary btn-sm" onclick="nextPage('task')">下一頁</button>
                </div>
            </div>

            <!-- 成員統計 & 貢獻度 -->
            <div class="row g-2">
                <div class="col-md-7">
                    <div class="card" id="cardMemberStats">
                        <div class="card-header">
                            <span class="card-header-title"><i class="bi bi-people me-1"></i>成員統計</span>
                            <div class="d-flex align-items-center">
                                <button class="btn btn-outline-light btn-sm me-1" onclick="toggleMemberFilter()"><i class="bi bi-funnel me-1"></i>篩選</button>
                                <button class="btn btn-outline-light btn-sm me-2" onclick="exportTableCSV('member')"><i class="bi bi-download me-1"></i>CSV</button>
                                <i class="bi bi-arrows-fullscreen card-maximize-btn text-white" onclick="toggleFullscreen('cardMemberStats')" title="最大化"></i>
                            </div>
                        </div>
                        <div class="table-toolbar" id="memberFilterBar" style="display:none;">
                            <input type="text" class="form-control form-control-sm" style="width:150px" placeholder="🔍 搜尋..." id="memberSearch" onkeyup="filterAndRenderMemberTable()">
                            <select class="form-select form-select-sm" style="width:130px" id="filterMemberModule" onchange="filterAndRenderMemberTable()"><option value="">全部模組</option></select>
                            <select class="form-select form-select-sm" style="width:110px" id="filterMemberPriority" onchange="filterAndRenderMemberTable()">
                                <option value="">全部優先</option><option value="high">High</option><option value="medium">Medium</option><option value="normal">Normal</option>
                            </select>
                            <select class="form-select form-select-sm" style="width:110px" id="filterMemberTaskStatus" onchange="filterAndRenderMemberTable()">
                                <option value="">全部狀態</option><option value="in_progress">進行中</option><option value="pending">Pending</option><option value="completed">已完成</option>
                            </select>
                            <select class="form-select form-select-sm" style="width:110px" id="filterMemberOverdue" onchange="filterAndRenderMemberTable()">
                                <option value="">全部超期</option><option value="hasOverdue">有超期</option><option value="noOverdue">無超期</option>
                            </select>
                            <button class="btn btn-outline-secondary btn-sm" onclick="clearMemberFilters()"><i class="bi bi-x-circle"></i></button>
                        </div>
                        <div class="table-container" style="height:400px;overflow-y:auto;">
                            <table class="table table-sm data-table mb-0">
                                <thead>
                                    <tr>
                                        <th onclick="sortTable('member','name')">成員 ↕</th>
                                        <th onclick="sortTable('member','total')">總數 ↕</th>
                                        <th onclick="sortTable('member','completed')">完成 ↕</th>
                                        <th onclick="sortTable('member','in_progress')">進行 ↕</th>
                                        <th onclick="sortTable('member','pending')">Pend ↕</th>
                                        <th onclick="sortTable('member','high')">H ↕</th>
                                        <th onclick="sortTable('member','medium')">M ↕</th>
                                        <th onclick="sortTable('member','normal')">N ↕</th>
                                    </tr>
                                </thead>
                                <tbody id="memberTableBody"></tbody>
                            </table>
                        </div>
                    </div>
                </div>
                <div class="col-md-5">
                    <div class="card" id="cardContrib">
                        <div class="card-header">
                            <span class="card-header-title"><i class="bi bi-trophy me-1"></i>貢獻度 <small class="text-warning">(含超期減分)</small></span>
                            <div class="d-flex align-items-center">
                                <button class="btn btn-outline-light btn-sm me-1" onclick="toggleContribFilter()"><i class="bi bi-funnel me-1"></i>篩選</button>
                                <button class="btn btn-outline-light btn-sm me-2" onclick="exportTableCSV('contrib')"><i class="bi bi-download me-1"></i>CSV</button>
                                <i class="bi bi-arrows-fullscreen card-maximize-btn text-white" onclick="toggleFullscreen('cardContrib')" title="最大化"></i>
                            </div>
                        </div>
                        <div class="table-toolbar" id="contribFilterBar" style="display:none;">
                            <input type="text" class="form-control form-control-sm" style="width:150px" placeholder="🔍 搜尋..." id="contribSearch" onkeyup="filterAndRenderContribTable()">
                            <select class="form-select form-select-sm" style="width:130px" id="filterContribModule" onchange="filterAndRenderContribTable()"><option value="">全部模組</option></select>
                            <select class="form-select form-select-sm" style="width:110px" id="filterContribPriority" onchange="filterAndRenderContribTable()">
                                <option value="">全部優先</option><option value="high">High</option><option value="medium">Medium</option><option value="normal">Normal</option>
                            </select>
                            <select class="form-select form-select-sm" style="width:110px" id="filterContribTaskStatus" onchange="filterAndRenderContribTable()">
                                <option value="">全部狀態</option><option value="in_progress">進行中</option><option value="pending">Pending</option><option value="completed">已完成</option>
                            </select>
                            <select class="form-select form-select-sm" style="width:110px" id="filterContribOverdue" onchange="filterAndRenderContribTable()">
                                <option value="">全部超期</option><option value="hasOverdue">有超期</option><option value="noOverdue">無超期</option>
                            </select>
                            <button class="btn btn-outline-secondary btn-sm" onclick="clearContribFilters()"><i class="bi bi-x-circle"></i></button>
                        </div>
                        <div class="table-container" style="height:400px;overflow-y:auto;">
                            <table class="table table-sm data-table mb-0">
                                <thead>
                                    <tr>
                                        <th onclick="sortTable('contrib','rank')"># ↕</th>
                                        <th onclick="sortTable('contrib','name')">成員 ↕</th>
                                        <th onclick="sortTable('contrib','task_count')">任務 ↕</th>
                                        <th onclick="sortTable('contrib','base_score')">基礎分 ↕</th>
                                        <th onclick="sortTable('contrib','overdue_count')">超期數 ↕</th>
                                        <th onclick="sortTable('contrib','overdue_penalty')">扣分 ↕</th>
                                        <th onclick="sortTable('contrib','score')">總分 ↕</th>
                                    </tr>
                                </thead>
                                <tbody id="contribTableBody"></tbody>
                            </table>
                        </div>
                    </div>
                </div>
            </div>
        </div><!-- end pane-stats -->
        </div><!-- end tab-content -->
        </div><!-- end resultArea -->

        <div class="footer">© 2025 Task Dashboard v23 | Powered by Vince</div>
    </div>

    <!-- Modal -->
    <div class="modal fade" id="detailModal" tabindex="-1">
        <div class="modal-dialog modal-xl">
            <div class="modal-content">
                <div class="modal-header py-2">
                    <h6 class="modal-title" id="modalTitle">明細</h6>
                    <button type="button" class="btn-close" data-bs-dismiss="modal"></button>
                </div>
                <div class="modal-body" style="max-height:70vh;overflow-y:auto;"><div id="modalContent"></div></div>
                <div class="modal-footer py-1">
                    <button class="btn btn-outline-secondary btn-sm" onclick="exportModalCSV()"><i class="bi bi-download me-1"></i>CSV</button>
                    <button type="button" class="btn btn-secondary btn-sm" data-bs-dismiss="modal">關閉</button>
                </div>
            </div>
        </div>
    </div>
    
    <!-- Mail Preview Modal -->
    <div class="modal fade" id="mailModal" tabindex="-1">
        <div class="modal-dialog modal-xl">
            <div class="modal-content">
                <div class="modal-header py-2 bg-primary text-white">
                    <h6 class="modal-title"><i class="bi bi-envelope me-1"></i>Mail 預覽</h6>
                    <button type="button" class="btn-close btn-close-white" data-bs-dismiss="modal"></button>
                </div>
                <div class="modal-body p-0">
                    <div class="p-2 bg-light border-bottom d-flex justify-content-between align-items-center">
                        <div>
                            <div><strong>主旨：</strong><span id="mailSubject">-</span></div>
                            <div><strong>日期：</strong><span id="mailDate">-</span> <span id="mailTime" class="text-muted"></span></div>
                            <div id="mailPreviewAttachments" style="display:none;" class="mt-1"></div>
                        </div>
                        <div class="btn-group btn-group-sm">
                            <button class="btn btn-outline-secondary active" onclick="setMailView('html')" id="btnHtml">HTML</button>
                            <button class="btn btn-outline-secondary" onclick="setMailView('text')" id="btnText">純文字</button>
                        </div>
                    </div>
                    <div id="mailBodyHtml" style="height:60vh;overflow:hidden;">
                        <iframe id="mailPreviewIframe" style="width:100%;height:100%;border:none;"></iframe>
                    </div>
                    <div id="mailBodyText" style="max-height:60vh;overflow-y:auto;padding:15px;font-family:monospace;font-size:13px;white-space:pre-wrap;background:#fafafa;display:none;"></div>
                </div>
                <div class="modal-footer py-1">
                    <button type="button" class="btn btn-secondary btn-sm" data-bs-dismiss="modal">關閉</button>
                </div>
            </div>
        </div>
    </div>

    <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/js/bootstrap.bundle.min.js"></script>
    <script>
        const treeData = {{ tree | tojson | safe }};
        let selectedEntry = null, selectedStore = null, resultData = null;
        let chart1 = null, chart2 = null, chart3 = null, chart4 = null, currentModal = null;
        let reviewModeActive = false;
        let allMails = [];
        
        const statusLabels = { completed: '已完成', pending: 'Pending', in_progress: '進行中' };
        
        // 表格狀態
        let tableState = {
            task: { data: [], filtered: [], page: 0, pageSize: 50, sortKey: 'last_seen', sortDir: -1 },
            member: { data: [], filtered: [], page: 0, pageSize: 50, sortKey: 'total', sortDir: -1 },
            contrib: { data: [], filtered: [], page: 0, pageSize: 50, sortKey: 'rank', sortDir: 1 }
        };

        // 初始化日期
        const today = new Date();
        document.getElementById('endDate').value = today.toISOString().split('T')[0];
        const startDate = new Date(today); startDate.setDate(today.getDate() - 30);
        document.getElementById('startDate').value = startDate.toISOString().split('T')[0];

        // 資料夾樹 - 預設全部展開
        let firstLeafNode = null;  // 記錄第一個葉節點
        let preferredNode = null;  // 優先選擇的節點 (收件匣下的 Dias-System team 協助事項)
        let inInbox = false;  // 追蹤是否在收件匣下
        let inArchive = false;  // 追蹤是否在封存下
        
        function buildTree(data, parent, parentName = '') {
            const ul = document.createElement('ul');
            data.forEach(node => {
                const li = document.createElement('li');
                const nodeLower = node.name.toLowerCase();
                const isInbox = node.name === '收件匣' || nodeLower === 'inbox' || nodeLower.includes('inbox');
                const isArchive = node.name === '封存' || nodeLower === 'archive' || nodeLower.includes('archive') || nodeLower.includes('封存');
                
                if (node.children && node.children.length > 0) {
                    const toggle = document.createElement('span');
                    toggle.className = 'tree-toggle open';  // 預設展開
                    toggle.textContent = node.name;
                    toggle.onclick = function(e) { e.stopPropagation(); this.classList.toggle('open'); li.querySelector(':scope > ul').style.display = this.classList.contains('open') ? 'block' : 'none'; };
                    li.appendChild(toggle);
                    
                    // 進入收件匣或封存時設定標記
                    const prevInInbox = inInbox;
                    const prevInArchive = inArchive;
                    if (isInbox) inInbox = true;
                    if (isArchive) inArchive = true;
                    const childUl = buildTree(node.children, li, node.name);
                    inInbox = prevInInbox;  // 離開時恢復
                    inArchive = prevInArchive;
                    
                    childUl.style.display = 'block';  // 預設顯示
                    li.appendChild(childUl);
                } else {
                    const item = document.createElement('span');
                    item.className = 'tree-item';
                    item.textContent = node.name;
                    item.dataset.entryId = node.entry_id;
                    item.dataset.storeId = node.store_id;
                    item.onclick = async function() {
                        document.querySelectorAll('.tree-item').forEach(i => i.classList.remove('selected'));
                        this.classList.add('selected');
                        selectedEntry = node.entry_id;
                        selectedStore = node.store_id;
                        document.getElementById('selectedFolder').textContent = node.name;
                        
                        // 點擊資料夾時，直接載入郵件（不套用日期篩選）並切換到 Review
                        useUploadedMails = false;
                        await loadFolderMailsDirect(true);
                        
                        // 顯示結果區域和 Review 頁籤
                        showResultArea();
                        document.getElementById('tabItem-stats').style.display = 'none';  // 隱藏統計
                        document.getElementById('tabItem-review').style.display = 'block';
                        
                        // 切換到 Review 頁籤
                        const reviewTab = document.getElementById('tab-review');
                        if (reviewTab) {
                            const bsTab = new bootstrap.Tab(reviewTab);
                            bsTab.show();
                        }
                        reviewModeActive = true;
                    };
                    li.appendChild(item);
                    
                    // 優先選擇收件匣下的 "Dias-System team 協助事項"（排除封存）
                    const isDiasFolder = node.name.includes('Dias-System') || node.name.includes('協助事項');
                    const isInInboxNotArchive = (inInbox || parentName === '收件匣' || parentName === 'Inbox') && !inArchive;
                    
                    if (isDiasFolder && isInInboxNotArchive && !preferredNode) {
                        // 只選第一個找到的，不覆蓋
                        preferredNode = { item: item, node: node };
                    } else if (!firstLeafNode) {
                        firstLeafNode = { item: item, node: node };
                    }
                }
                ul.appendChild(li);
            });
            return ul;
        }
        document.getElementById('tree').appendChild(buildTree(treeData, null, ''));
        
        // 預設選擇：優先收件匣下的 Dias-System team 協助事項，否則第一個資料夾
        const defaultNode = preferredNode || firstLeafNode;
        if (defaultNode) {
            defaultNode.item.classList.add('selected');
            selectedEntry = defaultNode.node.entry_id;
            selectedStore = defaultNode.node.store_id;
            document.getElementById('selectedFolder').textContent = defaultNode.node.name;
        }

        // 切換篩選設定
        function toggleFilterSettings() {
            const el = document.getElementById('filterSettings');
            el.style.display = el.style.display === 'none' ? 'block' : 'none';
        }
        
        // 切換任務列表篩選
        function toggleTaskFilter() {
            const el = document.getElementById('taskFilterBar');
            el.style.display = el.style.display === 'none' ? 'flex' : 'none';
        }
        
        // 清除任務列表篩選
        function clearTaskFilters() {
            document.getElementById('taskSearch').value = '';
            document.getElementById('filterModule').value = '';
            document.getElementById('filterOwner').value = '';
            document.getElementById('filterPriority').value = '';
            document.getElementById('filterStatus').value = '';
            document.getElementById('filterOverdue').value = '';
            filterAndRenderTaskTable();
        }
        
        // 切換成員統計篩選
        function toggleMemberFilter() {
            const el = document.getElementById('memberFilterBar');
            el.style.display = el.style.display === 'none' ? 'flex' : 'none';
        }
        
        // 清除成員統計篩選
        function clearMemberFilters() {
            document.getElementById('memberSearch').value = '';
            document.getElementById('filterMemberModule').value = '';
            document.getElementById('filterMemberPriority').value = '';
            document.getElementById('filterMemberTaskStatus').value = '';
            document.getElementById('filterMemberOverdue').value = '';
            filterAndRenderMemberTable();
        }
        
        // 切換貢獻度篩選
        function toggleContribFilter() {
            const el = document.getElementById('contribFilterBar');
            el.style.display = el.style.display === 'none' ? 'flex' : 'none';
        }
        
        // 清除貢獻度篩選
        function clearContribFilters() {
            document.getElementById('contribSearch').value = '';
            document.getElementById('filterContribModule').value = '';
            document.getElementById('filterContribPriority').value = '';
            document.getElementById('filterContribTaskStatus').value = '';
            document.getElementById('filterContribOverdue').value = '';
            filterAndRenderContribTable();
        }
        
        // 標記是否使用上傳的郵件
        let useUploadedMails = false;
        // reviewModeActive 已在上方宣告
        
        // 顯示結果區域
        function showResultArea() {
            document.getElementById('resultArea').style.display = 'block';
        }
        
        // 載入 Review 模式
        async function loadReviewMode() {
            if (!selectedEntry && !useUploadedMails) {
                alert('請先選擇資料夾');
                return;
            }
            
            document.getElementById('loading').style.display = 'flex';
            
            // 重設直接模式（使用日期篩選）
            directFolderMode = false;
            
            try {
                // Review 模式：使用與統計分析相同的 API，但不套用分析選項（不排除 Middle priority、不排除下午 5 點後）
                const r = await fetch('/api/outlook', { 
                    method: 'POST', 
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({
                        entry_id: selectedEntry, 
                        store_id: selectedStore, 
                        start: document.getElementById('startDate').value, 
                        end: document.getElementById('endDate').value,
                        exclude_middle_priority: false,  // Review 不排除
                        exclude_after_5pm: false,        // Review 不排除
                        include_mails: true
                    }) 
                });
                const data = await r.json();
                if (data.error) throw new Error(data.error);
                
                showResultArea();
                
                // 顯示兩個頁籤
                document.getElementById('tabItem-stats').style.display = 'block';
                document.getElementById('tabItem-review').style.display = 'block';
                
                // 切換到 Review 頁籤
                const reviewTab = document.getElementById('tab-review');
                const bsTab = new bootstrap.Tab(reviewTab);
                bsTab.show();
                
                reviewModeActive = true;
                
                resultData = data;
                updateUI();
                
                // 儲存郵件列表供 Review 使用
                if (data.mails) {
                    allMailsOriginal = data.mails;
                    allMails = data.mails;
                    reviewMailsTotal = data.mails.length;
                    reviewMailsLoaded = data.mails.length;
                    renderMailList();
                    updateReviewCount();
                }
            } catch (e) {
                alert('錯誤: ' + e.message);
            }
            
            document.getElementById('loading').style.display = 'none';
        }
        
        // 進階篩選郵件
        function applyMailFilters() {
            const field = document.getElementById('filterField').value;
            const keyword = document.getElementById('filterKeyword').value.toLowerCase();
            const attType = document.getElementById('filterAttType')?.value || '';
            
            let filtered = [...allMailsOriginal];
            
            // 附件類型篩選
            if (attType) {
                filtered = filtered.filter(m => {
                    const hasAtt = (m.attachments && m.attachments.length > 0) || m.has_attachments || (m.attachment_count > 0);
                    if (!hasAtt) return false;
                    if (attType === 'any') return true;
                    
                    // 如果有詳細附件資訊
                    if (m.attachments && m.attachments.length > 0) {
                        return m.attachments.some(att => {
                            const name = (att.name || '').toLowerCase();
                            switch (attType) {
                                case 'xlsx': return name.endsWith('.xlsx') || name.endsWith('.xls') || name.endsWith('.csv');
                                case 'docx': return name.endsWith('.docx') || name.endsWith('.doc');
                                case 'pptx': return name.endsWith('.pptx') || name.endsWith('.ppt');
                                case 'pdf': return name.endsWith('.pdf');
                                case 'image': return name.endsWith('.png') || name.endsWith('.jpg') || name.endsWith('.jpeg') || name.endsWith('.gif') || name.endsWith('.bmp');
                                case 'zip': return name.endsWith('.zip') || name.endsWith('.rar') || name.endsWith('.7z');
                                case 'txt': return name.endsWith('.txt') || name.endsWith('.log');
                                default: return true;
                            }
                        });
                    }
                    // 沒有詳細資訊，只要有附件就通過
                    return attType === 'any';
                });
            }
            
            // 關鍵字篩選
            if (keyword) {
                filtered = filtered.filter(m => {
                    if (field === 'subject') return (m.subject || '').toLowerCase().includes(keyword);
                    if (field === 'sender') return (m.sender || '').toLowerCase().includes(keyword);
                    if (field === 'recipient') return (m.recipient || m.to || '').toLowerCase().includes(keyword);
                    if (field === 'body') return (m.body || '').toLowerCase().includes(keyword);
                    // 全部欄位
                    return (m.subject || '').toLowerCase().includes(keyword) ||
                           (m.sender || '').toLowerCase().includes(keyword) ||
                           (m.recipient || m.to || '').toLowerCase().includes(keyword) ||
                           (m.body || '').toLowerCase().includes(keyword);
                });
            }
            
            allMails = filtered;
            renderMailList();
            updateReviewCount();
        }
        
        // 保存原始郵件列表
        let allMailsOriginal = [];
        
        // 篩選變更時重新篩選
        function onFilterChange() {
            if (reviewModeActive && allMailsOriginal.length > 0) {
                applyMailFilters();
            }
        }
        
        // 日期變更事件 - 如果在 Review 模式，重新載入郵件
        async function onDateChange() {
            if (reviewModeActive && selectedEntry && !useUploadedMails) {
                console.log('[Date] Date changed, reloading mails...');
                await loadMailsForReview(true);
            }
        }
        
        // Review 模式狀態
        let reviewMailsTotal = 0;
        let reviewMailsLoaded = 0;
        let reviewMailsLoading = false;
        const REVIEW_PAGE_SIZE = 100;
        
        // 載入郵件列表 (Review 模式) - 支援分頁動態載入
        async function loadMailsForReview(reset = false) {
            if (!selectedEntry || reviewMailsLoading) return;
            
            // 如果使用上傳的郵件，不從 Outlook 載入
            if (useUploadedMails) return;
            
            if (reset) {
                allMails = [];
                reviewMailsLoaded = 0;
                reviewMailsTotal = 0;
            }
            
            // 如果已載入全部，不再載入
            if (!reset && reviewMailsLoaded >= reviewMailsTotal && reviewMailsTotal > 0) return;
            
            reviewMailsLoading = true;
            
            // 顯示載入中
            const mailList = document.getElementById('mailList');
            if (reset) {
                mailList.innerHTML = '<div class="text-center p-3"><div class="spinner-border spinner-border-sm text-primary"></div> 載入中...</div>';
            } else {
                // 使用者捲動載入時，在底部顯示 loading
                showMailListLoading();
            }
            
            try {
                const r = await fetch('/api/review-mails', { 
                    method: 'POST', 
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({
                        entry_id: selectedEntry, 
                        store_id: selectedStore, 
                        start: document.getElementById('startDate').value, 
                        end: document.getElementById('endDate').value,
                        offset: reviewMailsLoaded,
                        limit: REVIEW_PAGE_SIZE
                    }) 
                });
                
                // 檢查回應類型
                const contentType = r.headers.get('content-type');
                if (!contentType || !contentType.includes('application/json')) {
                    throw new Error('伺服器返回非 JSON 格式');
                }
                
                const data = await r.json();
                if (data.error) throw new Error(data.error);
                
                reviewMailsTotal = data.total || 0;
                reviewMailsLoaded += (data.mails || []).length;
                
                // 合併郵件到原始列表
                if (reset) {
                    allMailsOriginal = data.mails || [];
                } else {
                    allMailsOriginal = allMailsOriginal.concat(data.mails || []);
                }
                
                // 應用篩選
                applyMailFilters();
                
                // 隱藏底部 loading
                hideMailListLoading();
                
            } catch (e) {
                console.error('載入郵件失敗:', e);
                if (reset) {
                    mailList.innerHTML = '<div class="text-center text-danger p-3">載入失敗: ' + e.message + '</div>';
                }
                hideMailListLoading();
            }
            
            reviewMailsLoading = false;
        }
        
        // 直接載入資料夾郵件（不套用日期篩選）- 用於點擊資料夾
        let folderMailsTotal = 0;
        let folderMailsLoaded = 0;
        let folderMailsLoading = false;
        let directFolderMode = false;  // 標記是否為直接載入模式
        let backgroundLoadTimer = null;  // 背景載入定時器
        
        async function loadFolderMailsDirect(reset = false, isBackgroundLoad = false) {
            if (!selectedEntry || folderMailsLoading) return;
            if (useUploadedMails) return;
            
            if (reset) {
                allMails = [];
                allMailsOriginal = [];
                folderMailsLoaded = 0;
                folderMailsTotal = 0;
                directFolderMode = true;
                // 清除背景載入定時器
                if (backgroundLoadTimer) {
                    clearTimeout(backgroundLoadTimer);
                    backgroundLoadTimer = null;
                }
            }
            
            if (!reset && folderMailsLoaded >= folderMailsTotal && folderMailsTotal > 0) return;
            
            folderMailsLoading = true;
            
            const mailList = document.getElementById('mailList');
            if (reset) {
                mailList.innerHTML = '<div class="text-center p-3"><div class="spinner-border spinner-border-sm text-primary"></div> 載入中...</div>';
            } else if (!isBackgroundLoad) {
                // 使用者捲動載入時，在底部顯示 loading
                showMailListLoading();
            }
            
            try {
                const r = await fetch('/api/folder-mails', { 
                    method: 'POST', 
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({
                        entry_id: selectedEntry, 
                        store_id: selectedStore,
                        offset: folderMailsLoaded,
                        limit: REVIEW_PAGE_SIZE
                    }) 
                });
                
                const contentType = r.headers.get('content-type');
                if (!contentType || !contentType.includes('application/json')) {
                    throw new Error('伺服器返回非 JSON 格式');
                }
                
                const data = await r.json();
                if (data.error) throw new Error(data.error);
                
                folderMailsTotal = data.total || 0;
                folderMailsLoaded += (data.mails || []).length;
                
                if (reset) {
                    allMailsOriginal = data.mails || [];
                } else {
                    allMailsOriginal = allMailsOriginal.concat(data.mails || []);
                }
                
                // 直接模式不套用篩選
                allMails = [...allMailsOriginal];
                renderMailList();
                updateFolderMailCount();
                
                // 隱藏底部 loading
                hideMailListLoading();
                
                // 啟動背景載入（如果還有更多）
                if (folderMailsLoaded < folderMailsTotal) {
                    scheduleBackgroundLoad();
                }
                
            } catch (e) {
                console.error('載入資料夾郵件失敗:', e);
                if (reset) {
                    mailList.innerHTML = '<div class="text-center text-danger p-3">載入失敗: ' + e.message + '</div>';
                }
                hideMailListLoading();
            }
            
            folderMailsLoading = false;
        }
        
        // 背景載入排程
        function scheduleBackgroundLoad() {
            if (backgroundLoadTimer) return;  // 已有排程
            if (!directFolderMode) return;
            if (folderMailsLoaded >= folderMailsTotal) return;
            
            backgroundLoadTimer = setTimeout(() => {
                backgroundLoadTimer = null;
                if (directFolderMode && !folderMailsLoading && folderMailsLoaded < folderMailsTotal) {
                    console.log('[Background] Loading more...', folderMailsLoaded, '/', folderMailsTotal);
                    loadFolderMailsDirect(false, true);  // 背景載入
                }
            }, 500);  // 500ms 後載入下一批
        }
        
        // 顯示郵件列表底部 loading
        function showMailListLoading() {
            const mailList = document.getElementById('mailList');
            let loadingEl = document.getElementById('mailListLoading');
            if (!loadingEl) {
                loadingEl = document.createElement('div');
                loadingEl.id = 'mailListLoading';
                loadingEl.className = 'text-center p-2';
                loadingEl.style.cssText = 'background: linear-gradient(135deg, #e3f2fd 0%, #bbdefb 100%); border-top: 1px solid #90caf9;';
                loadingEl.innerHTML = '<div class="spinner-border spinner-border-sm text-primary me-2"></div><span class="text-primary">載入更多郵件中...</span>';
                mailList.appendChild(loadingEl);
            }
            loadingEl.style.display = 'block';
        }
        
        // 隱藏郵件列表底部 loading
        function hideMailListLoading() {
            const loadingEl = document.getElementById('mailListLoading');
            if (loadingEl) {
                loadingEl.style.display = 'none';
            }
        }
        
        function updateFolderMailCount() {
            const countEl = document.getElementById('reviewMailCount');
            const countDetailEl = document.getElementById('reviewMailCountDetail');
            // 黃色 badge 顯示總數
            if (countEl) countEl.textContent = folderMailsTotal;
            // 藍色 badge 顯示已載入/總數
            if (countDetailEl) {
                countDetailEl.textContent = `已載入 ${folderMailsLoaded} / ${folderMailsTotal} 封`;
                // 如果還沒載入完，顯示更醒目的顏色
                if (folderMailsLoaded < folderMailsTotal) {
                    countDetailEl.className = 'badge bg-info text-white me-2';
                    countDetailEl.style.fontSize = '0.7rem';
                } else {
                    countDetailEl.className = 'badge bg-success text-white me-2';
                    countDetailEl.style.fontSize = '0.7rem';
                }
            }
        }
        
        // 更新 Review 模式計數（篩選模式）
        function updateReviewCount() {
            const countEl = document.getElementById('reviewMailCount');
            const countDetailEl = document.getElementById('reviewMailCountDetail');
            // 黃色 badge 顯示總數
            if (countEl) countEl.textContent = reviewMailsTotal;
            // 藍色 badge 顯示已載入/總數
            if (countDetailEl) {
                countDetailEl.textContent = `已載入 ${allMailsOriginal.length} / ${reviewMailsTotal} 封`;
                if (allMailsOriginal.length < reviewMailsTotal) {
                    countDetailEl.className = 'badge bg-info text-white me-2';
                    countDetailEl.style.fontSize = '0.7rem';
                } else {
                    countDetailEl.className = 'badge bg-success text-white me-2';
                    countDetailEl.style.fontSize = '0.7rem';
                }
            }
        }
        
        // 滾動載入節流
        let scrollThrottleTimer = null;
        let lastScrollTime = 0;
        
        // 滾動載入更多
        function onMailListScroll(e) {
            if (!reviewModeActive || useUploadedMails) return;
            
            const now = Date.now();
            if (now - lastScrollTime < 300) return;  // 300ms 節流
            
            const el = e.target;
            // 當滾動到底部 150px 內時載入更多
            if (el.scrollHeight - el.scrollTop - el.clientHeight < 150) {
                lastScrollTime = now;
                
                // 根據模式選擇載入方式
                if (directFolderMode) {
                    if (folderMailsLoading) return;
                    if (folderMailsLoaded >= folderMailsTotal) return;
                    console.log('[Scroll] Loading more (direct)...', folderMailsLoaded, '/', folderMailsTotal);
                    loadFolderMailsDirect(false);
                } else {
                    if (reviewMailsLoading) return;
                    if (reviewMailsLoaded >= reviewMailsTotal) return;
                    console.log('[Scroll] Loading more (filtered)...', reviewMailsLoaded, '/', reviewMailsTotal);
                    loadMailsForReview(false);
                }
            }
        }

        // 分析
        async function analyze() {
            if (!selectedEntry) { alert('請選擇資料夾'); return; }
            document.getElementById('loading').style.display = 'flex';
            
            // 使用 Outlook 分析，重置上傳標誌和直接模式
            useUploadedMails = false;
            directFolderMode = false;  // 使用篩選模式
            
            const excludeMiddlePriority = document.getElementById('excludeMiddlePriority').checked;
            const excludeAfter5pm = document.getElementById('excludeAfter5pm').checked;
            
            try {
                const r = await fetch('/api/outlook', { 
                    method: 'POST', 
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({
                        entry_id: selectedEntry, 
                        store_id: selectedStore, 
                        start: document.getElementById('startDate').value, 
                        end: document.getElementById('endDate').value,
                        exclude_middle_priority: excludeMiddlePriority,
                        exclude_after_5pm: excludeAfter5pm,
                        include_mails: true
                    }) 
                });
                const data = await r.json();
                if (data.error) throw new Error(data.error);
                
                // 顯示結果區域
                showResultArea();
                
                // 顯示兩個頁籤
                document.getElementById('tabItem-stats').style.display = 'block';
                document.getElementById('tabItem-review').style.display = 'block';
                
                // 切換到統計頁籤
                const statsTab = document.getElementById('tab-stats');
                const bsTab = new bootstrap.Tab(statsTab);
                bsTab.show();
                
                reviewModeActive = false;
                
                // 檢查是否有任務
                if (data.total_tasks === 0) {
                    alert('未找到符合條件的任務，請切換到 Review 模式查看郵件');
                }
                
                resultData = data;
                updateUI();
                
                // 儲存郵件列表供 Review 使用
                if (data.mails) {
                    allMailsOriginal = data.mails;
                    allMails = data.mails;
                    reviewMailsTotal = data.mails.length;
                    reviewMailsLoaded = data.mails.length;
                    updateReviewCount();
                }
            } catch (e) {
                alert('錯誤: ' + e.message);
            }
            document.getElementById('loading').style.display = 'none';
        }

        // 更新 UI
        function updateUI() {
            document.getElementById('totalTasks').textContent = resultData.total_tasks;
            document.getElementById('pendingCount').textContent = resultData.pending_count;
            document.getElementById('inProgressCount').textContent = resultData.in_progress_count;
            document.getElementById('completedCount').textContent = resultData.completed_count;
            document.getElementById('overdueCount').textContent = resultData.overdue_count;
            document.getElementById('lastMailDate').textContent = resultData.last_mail_date || '-';
            
            const total = resultData.total_tasks || 1;
            document.getElementById('completedBar').style.width = (resultData.completed_count / total * 100) + '%';
            document.getElementById('inProgressBar').style.width = (resultData.in_progress_count / total * 100) + '%';
            document.getElementById('pendingBar').style.width = (resultData.pending_count / total * 100) + '%';
            
            // 填充篩選下拉
            fillFilterOptions();
            
            // 初始化表格資料
            tableState.task.data = resultData.all_tasks || [];
            tableState.member.data = resultData.members || [];
            tableState.contrib.data = resultData.contribution || [];
            
            filterAndRenderTaskTable();
            filterAndRenderMemberTable();
            filterAndRenderContribTable();
            
            updateChart1(); updateChart2(); updateChart3(); updateChart4();
        }
        
        // 填充篩選下拉選項
        function fillFilterOptions() {
            // 模組 - 任務列表
            const moduleSelect = document.getElementById('filterModule');
            moduleSelect.innerHTML = '<option value="">全部模組</option>';
            (resultData.module_list || []).forEach(m => {
                moduleSelect.innerHTML += `<option value="${esc(m)}">${m || '未分類'}</option>`;
            });
            
            // 負責人 - 任務列表
            const ownerSelect = document.getElementById('filterOwner');
            ownerSelect.innerHTML = '<option value="">全部負責人</option>';
            (resultData.member_list || []).forEach(o => {
                ownerSelect.innerHTML += `<option value="${esc(o)}">${o}</option>`;
            });
            
            // 模組 - 成員統計
            const memberModuleSelect = document.getElementById('filterMemberModule');
            if (memberModuleSelect) {
                memberModuleSelect.innerHTML = '<option value="">全部模組</option>';
                (resultData.module_list || []).forEach(m => {
                    memberModuleSelect.innerHTML += `<option value="${esc(m)}">${m || '未分類'}</option>`;
                });
            }
            
            // 模組 - 貢獻度
            const contribModuleSelect = document.getElementById('filterContribModule');
            if (contribModuleSelect) {
                contribModuleSelect.innerHTML = '<option value="">全部模組</option>';
                (resultData.module_list || []).forEach(m => {
                    contribModuleSelect.innerHTML += `<option value="${esc(m)}">${m || '未分類'}</option>`;
                });
            }
        }

        // 表格篩選與渲染
        function filterAndRenderTaskTable() {
            const search = (document.getElementById('taskSearch')?.value || '').toLowerCase();
            const module = document.getElementById('filterModule')?.value || '';
            const owner = document.getElementById('filterOwner')?.value || '';
            const priority = document.getElementById('filterPriority')?.value || '';
            const status = document.getElementById('filterStatus')?.value || '';
            const overdue = document.getElementById('filterOverdue')?.value || '';
            
            tableState.task.filtered = tableState.task.data.filter(t => {
                if (search && !JSON.stringify(t).toLowerCase().includes(search)) return false;
                if (module && (t.module || '') !== module) return false;
                if (owner && !t.owners_str.includes(owner)) return false;
                if (priority && t.priority !== priority) return false;
                if (status && t.task_status !== status) return false;
                if (overdue === 'yes' && t.overdue_days <= 0) return false;
                if (overdue === 'no' && t.overdue_days > 0) return false;
                return true;
            });
            tableState.task.page = 0;
            renderTaskTable();
        }
        
        function filterAndRenderMemberTable() {
            const search = (document.getElementById('memberSearch')?.value || '').toLowerCase();
            const module = document.getElementById('filterMemberModule')?.value || '';
            const priority = document.getElementById('filterMemberPriority')?.value || '';
            const taskStatus = document.getElementById('filterMemberTaskStatus')?.value || '';
            const overdueFilter = document.getElementById('filterMemberOverdue')?.value || '';
            
            // 根據篩選條件重新計算成員統計
            let filteredTasks = resultData.all_tasks;
            if (module) filteredTasks = filteredTasks.filter(t => (t.module || '') === module);
            if (priority) filteredTasks = filteredTasks.filter(t => t.priority === priority);
            if (taskStatus) filteredTasks = filteredTasks.filter(t => t.task_status === taskStatus);
            
            // 重新計算成員統計
            const memberStats = {};
            filteredTasks.forEach(t => {
                (t.owners || t.owners_str?.split('/') || []).forEach(owner => {
                    if (!memberStats[owner]) memberStats[owner] = { name: owner, total: 0, completed: 0, in_progress: 0, pending: 0, high: 0, medium: 0, normal: 0, overdue_count: 0 };
                    memberStats[owner].total++;
                    memberStats[owner][t.task_status] = (memberStats[owner][t.task_status] || 0) + 1;
                    memberStats[owner][t.priority] = (memberStats[owner][t.priority] || 0) + 1;
                    if (t.overdue_days > 0 && t.task_status !== 'completed') memberStats[owner].overdue_count++;
                });
            });
            
            let memberList = Object.values(memberStats);
            
            // 搜尋和超期篩選
            tableState.member.filtered = memberList.filter(m => {
                if (search && !m.name.toLowerCase().includes(search)) return false;
                if (overdueFilter === 'hasOverdue' && m.overdue_count === 0) return false;
                if (overdueFilter === 'noOverdue' && m.overdue_count > 0) return false;
                return true;
            });
            
            tableState.member.page = 0;
            renderMemberTable();
        }
        
        function filterAndRenderContribTable() {
            const search = (document.getElementById('contribSearch')?.value || '').toLowerCase();
            const module = document.getElementById('filterContribModule')?.value || '';
            const priorityFilter = document.getElementById('filterContribPriority')?.value || '';
            const statusFilter = document.getElementById('filterContribTaskStatus')?.value || '';
            const overdueFilter = document.getElementById('filterContribOverdue')?.value || '';
            
            // 先根據模組和優先級篩選所有任務（用於計算任務數，包含 pending）
            let allFilteredTasks = resultData.all_tasks;
            if (module) allFilteredTasks = allFilteredTasks.filter(t => (t.module || '') === module);
            if (priorityFilter) allFilteredTasks = allFilteredTasks.filter(t => t.priority === priorityFilter);
            if (statusFilter) allFilteredTasks = allFilteredTasks.filter(t => t.task_status === statusFilter);
            
            // 用於計算分數的任務（排除 pending）
            let scoringTasks = allFilteredTasks.filter(t => t.task_status !== 'pending');
            
            // 重新計算貢獻度
            const contribStats = {};
            
            // 先計算任務數（包含 pending）
            allFilteredTasks.forEach(t => {
                (t.owners || t.owners_str?.split('/') || []).forEach(owner => {
                    if (!contribStats[owner]) {
                        contribStats[owner] = { 
                            name: owner, task_count: 0, high: 0, medium: 0, normal: 0,
                            base_score: 0, overdue_count: 0, overdue_days: 0, overdue_penalty: 0, score: 0
                        };
                    }
                    contribStats[owner].task_count++;
                });
            });
            
            // 再計算分數（排除 pending）
            scoringTasks.forEach(t => {
                (t.owners || t.owners_str?.split('/') || []).forEach(owner => {
                    if (!contribStats[owner]) {
                        contribStats[owner] = { 
                            name: owner, task_count: 0, high: 0, medium: 0, normal: 0,
                            base_score: 0, overdue_count: 0, overdue_days: 0, overdue_penalty: 0, score: 0
                        };
                    }
                    contribStats[owner][t.priority] = (contribStats[owner][t.priority] || 0) + 1;
                    if (t.overdue_days > 0 && t.task_status !== 'completed') {
                        contribStats[owner].overdue_count++;
                        contribStats[owner].overdue_days += t.overdue_days;
                    }
                });
            });
            
            // 計算分數
            Object.values(contribStats).forEach(c => {
                c.base_score = c.high * 3 + c.medium * 2 + c.normal * 1;
                c.overdue_penalty = Math.round(c.overdue_days * 0.1 * 10) / 10;
                c.score = Math.round((c.base_score - c.overdue_penalty) * 10) / 10;
            });
            
            // 排序並加入排名
            let contribList = Object.values(contribStats).sort((a, b) => b.score - a.score);
            contribList.forEach((c, i) => c.rank = i + 1);
            
            // 搜尋和超期篩選
            tableState.contrib.filtered = contribList.filter(c => {
                if (search && !c.name.toLowerCase().includes(search)) return false;
                if (overdueFilter === 'hasOverdue' && c.overdue_count === 0) return false;
                if (overdueFilter === 'noOverdue' && c.overdue_count > 0) return false;
                return true;
            });
            
            tableState.contrib.page = 0;
            renderContribTable();
        }
        
        function sortTable(table, key) {
            const state = tableState[table];
            if (state.sortKey === key) state.sortDir *= -1;
            else { state.sortKey = key; state.sortDir = 1; }
            
            state.filtered.sort((a, b) => {
                let va = a[key], vb = b[key];
                if (va == null) va = '';
                if (vb == null) vb = '';
                if (typeof va === 'number') return (va - vb) * state.sortDir;
                return String(va).localeCompare(String(vb)) * state.sortDir;
            });
            
            if (table === 'task') renderTaskTable();
            else if (table === 'member') renderMemberTable();
            else renderContribTable();
        }
        
        function esc(s) { return String(s || '').replace(/'/g, "\\'").replace(/"/g, '&quot;'); }

        // 根據附件類型返回對應圖示（可點擊開啟 Mail 預覽）
        function getAttachmentIcons(attachments, hasAttachments, mailId = null) {
            const clickAttr = mailId ? `style="cursor:pointer;font-size:0.75rem" onclick="showMailPreview('${mailId}', event)"` : `style="font-size:0.75rem"`;
            
            // 如果有詳細附件資訊
            if (attachments && attachments.length > 0) {
                const icons = [];
                let hasExcel = false, hasWord = false, hasPpt = false, hasPdf = false, hasImage = false, hasOther = false;
                
                attachments.forEach(att => {
                    const name = (att.name || '').toLowerCase();
                    if (name.endsWith('.xlsx') || name.endsWith('.xls') || name.endsWith('.csv')) hasExcel = true;
                    else if (name.endsWith('.docx') || name.endsWith('.doc')) hasWord = true;
                    else if (name.endsWith('.pptx') || name.endsWith('.ppt')) hasPpt = true;
                    else if (name.endsWith('.pdf')) hasPdf = true;
                    else if (name.endsWith('.png') || name.endsWith('.jpg') || name.endsWith('.jpeg') || name.endsWith('.gif')) hasImage = true;
                    else hasOther = true;
                });
                
                if (hasExcel) icons.push(`<i class="bi bi-file-earmark-excel text-success" ${clickAttr} title="Excel 附件"></i>`);
                if (hasWord) icons.push(`<i class="bi bi-file-earmark-word text-primary" ${clickAttr} title="Word 附件"></i>`);
                if (hasPpt) icons.push(`<i class="bi bi-file-earmark-ppt text-danger" ${clickAttr} title="PPT 附件"></i>`);
                if (hasPdf) icons.push(`<i class="bi bi-file-earmark-pdf text-danger" ${clickAttr} title="PDF 附件"></i>`);
                if (hasImage) icons.push(`<i class="bi bi-file-earmark-image text-info" ${clickAttr} title="圖片附件"></i>`);
                if (hasOther) icons.push(`<i class="bi bi-paperclip text-secondary" ${clickAttr} title="其他附件"></i>`);
                
                return icons.length > 0 ? `<span class="ms-1">${icons.join('')}</span>` : '';
            }
            // 如果只有 has_attachments flag
            if (hasAttachments) {
                return `<i class="bi bi-paperclip ms-1 text-secondary" ${clickAttr} title="有附件"></i>`;
            }
            return '';
        }
        
        function renderTaskTable() {
            const state = tableState.task;
            state.pageSize = parseInt(document.getElementById('taskPageSize').value);
            const start = state.page * state.pageSize;
            const pageData = state.filtered.slice(start, start + state.pageSize);
            
            document.getElementById('taskTableBody').innerHTML = pageData.map(t => `
                <tr class="row-${t.task_status} ${t.overdue_days > 0 ? 'row-overdue' : ''}">
                    <td>${t.last_seen || '-'}</td>
                    <td><span class="badge bg-secondary" style="font-size:0.65rem">${t.module || '-'}</span></td>
                    <td>
                        <span style="cursor:pointer" onclick="showTaskDetail('${esc(t.title)}')">${t.title}</span>
                        ${t.mail_id ? `<i class="bi bi-envelope ms-1 text-primary" style="cursor:pointer;font-size:0.8rem" onclick="showMailPreview('${t.mail_id}', event)" title="預覽 Mail"></i>` : ''}
                        ${getAttachmentIcons(t.attachments, t.has_attachments, t.mail_id)}
                    </td>
                    <td>${t.owners_str}</td>
                    <td><span class="badge badge-${t.priority}">${t.priority}</span></td>
                    <td class="${t.overdue_days > 0 ? 'text-overdue' : ''}">${t.due || '-'}</td>
                    <td class="${t.overdue_days > 0 ? 'text-overdue' : ''}">${t.overdue_days > 0 ? '+' + t.overdue_days + '天' : '-'}</td>
                    <td><span class="badge badge-${t.task_status}">${statusLabels[t.task_status]}</span></td>
                </tr>
            `).join('');
            
            const totalPages = Math.ceil(state.filtered.length / state.pageSize) || 1;
            document.getElementById('taskPageInfo').textContent = `第 ${state.page + 1}/${totalPages} 頁 (共 ${state.filtered.length} 筆)`;
        }
        
        function renderMemberTable() {
            const state = tableState.member;
            document.getElementById('memberTableBody').innerHTML = state.filtered.map(m => `
                <tr>
                    <td><strong style="cursor:pointer" onclick="showMemberTasks('${esc(m.name)}')">${m.name}</strong></td>
                    <td style="cursor:pointer" onclick="showMemberTasks('${esc(m.name)}')">${m.total}</td>
                    <td style="cursor:pointer" onclick="showMemberTasksByStatus('${esc(m.name)}', 'completed')"><span class="badge badge-completed">${m.completed}</span></td>
                    <td style="cursor:pointer" onclick="showMemberTasksByStatus('${esc(m.name)}', 'in_progress')"><span class="badge badge-in_progress">${m.in_progress}</span></td>
                    <td style="cursor:pointer" onclick="showMemberTasksByStatus('${esc(m.name)}', 'pending')"><span class="badge badge-pending">${m.pending}</span></td>
                    <td style="cursor:pointer" onclick="showMemberTasksByPriority('${esc(m.name)}', 'high')"><span class="badge badge-high">${m.high}</span></td>
                    <td style="cursor:pointer" onclick="showMemberTasksByPriority('${esc(m.name)}', 'medium')"><span class="badge badge-medium">${m.medium}</span></td>
                    <td style="cursor:pointer" onclick="showMemberTasksByPriority('${esc(m.name)}', 'normal')"><span class="badge badge-normal">${m.normal}</span></td>
                </tr>
            `).join('');
        }
        
        function renderContribTable() {
            const state = tableState.contrib;
            document.getElementById('contribTableBody').innerHTML = state.filtered.map(c => `
                <tr>
                    <td><span class="rank-badge ${c.rank <= 3 ? 'rank-' + c.rank : 'rank-other'}">${c.rank}</span></td>
                    <td style="cursor:pointer" onclick="showMemberTasks('${esc(c.name)}')">${c.name}</td>
                    <td style="cursor:pointer" onclick="showMemberTasks('${esc(c.name)}')">${c.task_count}</td>
                    <td style="cursor:pointer" onclick="showContribDetail('${esc(c.name)}')">${c.base_score}</td>
                    <td class="${c.overdue_count > 0 ? 'text-overdue' : ''}" style="cursor:pointer" onclick="showMemberOverdueTasks('${esc(c.name)}')">${c.overdue_count}</td>
                    <td class="${c.overdue_penalty > 0 ? 'text-overdue' : ''}" style="cursor:pointer" onclick="showContribDetail('${esc(c.name)}')">-${c.overdue_penalty}</td>
                    <td style="cursor:pointer" onclick="showContribDetail('${esc(c.name)}')"><strong>${c.score}</strong></td>
                </tr>
            `).join('');
        }
        
        // 成員任務查看函數
        function showMemberTasks(name) {
            if (!resultData) return;
            const tasks = resultData.all_tasks.filter(t => t.owners_str.includes(name));
            showModal(`${name} 的任務 (${tasks.length})`, modalTableWithFilters(tasks, 'memberTasks'));
        }
        
        function showMemberTasksByStatus(name, status) {
            if (!resultData) return;
            const tasks = resultData.all_tasks.filter(t => t.owners_str.includes(name) && t.task_status === status);
            showModal(`${name} - ${statusLabels[status]} (${tasks.length})`, modalTableWithFilters(tasks, 'memberStatusTasks'));
        }
        
        function showMemberTasksByPriority(name, priority) {
            if (!resultData) return;
            const tasks = resultData.all_tasks.filter(t => t.owners_str.includes(name) && t.priority === priority);
            showModal(`${name} - ${priority.toUpperCase()} 優先級 (${tasks.length})`, modalTableWithFilters(tasks, 'memberPriorityTasks'));
        }
        
        function showContribDetail(name) {
            if (!resultData) return;
            // 優先從動態篩選後的數據取，確保與表格顯示一致
            let c = tableState.contrib.filtered.find(x => x.name === name);
            if (!c) c = resultData.contribution.find(x => x.name === name);
            if (!c) return;
            const detail = `
                <div class="p-3">
                    <div class="d-flex align-items-center mb-3">
                        <i class="bi bi-person-circle fs-2 text-primary me-2"></i>
                        <h5 class="mb-0">${name} 貢獻度計算明細</h5>
                    </div>
                    <table class="table table-sm data-table mb-0">
                        <tbody>
                            <tr><td class="fw-bold" style="width:50%">任務數</td><td>${c.task_count}</td></tr>
                            <tr><td>High 任務 × 3</td><td><span class="badge badge-high me-1">${c.high}</span>× 3 = ${c.high * 3}</td></tr>
                            <tr><td>Medium 任務 × 2</td><td><span class="badge badge-medium me-1">${c.medium}</span>× 2 = ${c.medium * 2}</td></tr>
                            <tr><td>Normal 任務 × 1</td><td><span class="badge badge-normal me-1">${c.normal}</span>× 1 = ${c.normal}</td></tr>
                            <tr class="table-active"><td class="fw-bold">基礎分</td><td class="fw-bold">${c.base_score}</td></tr>
                        </tbody>
                    </table>
                    <table class="table table-sm data-table mt-2 mb-0">
                        <tbody>
                            <tr class="row-overdue"><td style="width:50%">超期任務數</td><td>${c.overdue_count}</td></tr>
                            <tr class="row-overdue"><td>總超期天數</td><td>${c.overdue_days}</td></tr>
                            <tr class="row-overdue"><td class="fw-bold">扣分 (天數 × 0.1 × -1)</td><td class="fw-bold text-danger">-${c.overdue_penalty}</td></tr>
                        </tbody>
                    </table>
                    <table class="table table-sm data-table mt-2 mb-0">
                        <tbody>
                            <tr style="background:#d4edda"><td style="width:50%" class="fw-bold fs-5">總分</td><td class="fw-bold fs-5 text-success">${c.score}</td></tr>
                        </tbody>
                    </table>
                    <div class="text-muted small mt-3">
                        <i class="bi bi-info-circle me-1"></i>計算公式: 總分 = 基礎分 - 扣分 = ${c.base_score} - ${c.overdue_penalty} = ${c.score}
                    </div>
                </div>
            `;
            showModal(`${name} 貢獻度明細`, detail);
        }
        
        function filterTaskByOwner(name) {
            document.getElementById('filterOwner').value = name;
            filterAndRenderTaskTable();
        }
        
        function prevPage(table) {
            if (tableState[table].page > 0) {
                tableState[table].page--;
                if (table === 'task') renderTaskTable();
            }
        }
        
        function nextPage(table) {
            const state = tableState[table];
            if ((state.page + 1) * state.pageSize < state.filtered.length) {
                state.page++;
                if (table === 'task') renderTaskTable();
            }
        }

        // 圖表
        function updateChart1() {
            const type = document.getElementById('chart1Type').value;
            if (chart1) chart1.destroy();
            chart1 = new Chart(document.getElementById('chart1'), {
                type: type,
                data: { labels: ['進行中', 'Pending', '已完成'], datasets: [{ data: [resultData.in_progress_count, resultData.pending_count, resultData.completed_count], backgroundColor: ['#17a2b8', '#FFA500', '#28a745'] }] },
                options: { maintainAspectRatio: false, plugins: { legend: { display: type !== 'bar', position: 'right' } }, onClick: (e, el) => { if (el.length) showByStatus(['in_progress', 'pending', 'completed'][el[0].index]); } }
            });
        }

        function updateChart2() {
            const type = document.getElementById('chart2Type').value;
            if (chart2) chart2.destroy();
            chart2 = new Chart(document.getElementById('chart2'), {
                type: type,
                data: { labels: ['High', 'Medium', 'Normal'], datasets: [{ data: [resultData.priority_counts.high, resultData.priority_counts.medium, resultData.priority_counts.normal], backgroundColor: ['#FF6B6B', '#FFE066', '#74C0FC'] }] },
                options: { maintainAspectRatio: false, plugins: { legend: { display: type !== 'bar', position: 'right' } }, onClick: (e, el) => { if (el.length) showByPriority(['high', 'medium', 'normal'][el[0].index]); } }
            });
        }

        function updateChart3() {
            const type = document.getElementById('chart3Type').value;
            if (chart3) chart3.destroy();
            chart3 = new Chart(document.getElementById('chart3'), {
                type: type,
                data: { labels: ['超期', '未超期'], datasets: [{ data: [resultData.overdue_count, resultData.not_overdue_count], backgroundColor: ['#dc3545', '#28a745'] }] },
                options: { maintainAspectRatio: false, plugins: { legend: { display: type !== 'bar', position: 'right' } }, onClick: (e, el) => { if (el.length && el[0].index === 0) showOverdue(); else if (el.length && el[0].index === 1) showNotOverdue(); } }
            });
        }

        function updateChart4() {
            const type = document.getElementById('chart4Type').value;
            if (chart4) chart4.destroy();
            const ctx = document.getElementById('chart4').getContext('2d');
            
            const overdueData = resultData.contribution.filter(c => c.overdue_days > 0).sort((a, b) => b.overdue_days - a.overdue_days).slice(0, 10);
            
            if (overdueData.length === 0) {
                chart4 = new Chart(ctx, { type: 'bar', data: { labels: ['無超期'], datasets: [{ data: [0], backgroundColor: '#28a745' }] }, options: { maintainAspectRatio: false, plugins: { legend: { display: false } } } });
                return;
            }
            
            const labels = overdueData.map(c => c.name);
            
            if (type === 'vstacked') {
                // 垂直堆疊
                chart4 = new Chart(ctx, {
                    type: 'bar',
                    data: { labels, datasets: [
                        { label: '已完成超期', data: overdueData.map(c => c.completed_overdue_days || 0), backgroundColor: '#6c757d', stack: 's' },
                        { label: '未完成超期', data: overdueData.map(c => c.active_overdue_days || 0), backgroundColor: '#dc3545', stack: 's' }
                    ]},
                    options: { maintainAspectRatio: false, plugins: { legend: { display: true, position: 'top' } }, scales: { x: { stacked: true }, y: { stacked: true, beginAtZero: true } }, onClick: (e, el) => { if (el.length) showMemberOverdueTasks(labels[el[0].index]); } }
                });
            } else if (type === 'line') {
                // 折線圖
                chart4 = new Chart(ctx, {
                    type: 'line',
                    data: { labels, datasets: [
                        { label: '已完成超期', data: overdueData.map(c => c.completed_overdue_days || 0), borderColor: '#6c757d', backgroundColor: 'rgba(108,117,125,0.2)', fill: true, tension: 0.3 },
                        { label: '未完成超期', data: overdueData.map(c => c.active_overdue_days || 0), borderColor: '#dc3545', backgroundColor: 'rgba(220,53,69,0.2)', fill: true, tension: 0.3 }
                    ]},
                    options: { maintainAspectRatio: false, plugins: { legend: { display: true, position: 'top' } }, scales: { y: { beginAtZero: true } }, onClick: (e, el) => { if (el.length) showMemberOverdueTasks(labels[el[0].index]); } }
                });
            } else {
                // 水平堆疊 (預設)
                chart4 = new Chart(ctx, {
                    type: 'bar',
                    data: { labels, datasets: [
                        { label: '已完成超期', data: overdueData.map(c => c.completed_overdue_days || 0), backgroundColor: '#6c757d', stack: 's' },
                        { label: '未完成超期', data: overdueData.map(c => c.active_overdue_days || 0), backgroundColor: '#dc3545', stack: 's' }
                    ]},
                    options: { maintainAspectRatio: false, indexAxis: 'y', plugins: { legend: { display: true, position: 'top' } }, scales: { x: { stacked: true, beginAtZero: true }, y: { stacked: true } }, onClick: (e, el) => { if (el.length) showMemberOverdueTasks(labels[el[0].index]); } }
                });
            }
        }

        // Modal 顯示
        function showModal(title, content) {
            document.getElementById('modalTitle').textContent = title;
            document.getElementById('modalContent').innerHTML = content;
            currentModal = new bootstrap.Modal(document.getElementById('detailModal'));
            currentModal.show();
        }
        
        // Modal 任務表格 - 基本版
        function modalTable(tasks, id = 'modalTableBody') {
            return `
                <div class="mb-2"><input type="text" class="form-control form-control-sm" style="max-width:250px" placeholder="🔍 搜尋..." onkeyup="filterModalTable('${id}', this.value)"></div>
                <div style="max-height: 50vh; overflow-y: auto;">
                    <table class="table table-sm data-table">
                        <thead><tr><th>Mail日期</th><th>模組</th><th>任務</th><th>負責人</th><th>優先級</th><th>Due</th><th>超期</th><th>狀態</th></tr></thead>
                        <tbody id="${id}">${tasks.map(t => `
                            <tr class="row-${t.task_status} ${t.overdue_days > 0 ? 'row-overdue' : ''}">
                                <td>${t.last_seen || t.mail_date || '-'}</td>
                                <td><span class="badge bg-secondary" style="font-size:0.6rem">${t.module || '-'}</span></td>
                                <td>${t.title} ${t.mail_id ? `<i class="bi bi-envelope ms-1 text-primary" style="cursor:pointer;font-size:0.8rem" onclick="showMailPreview('${t.mail_id}', event)" title="預覽"></i>` : ''}</td>
                                <td>${t.owners_str || (t.owners ? t.owners.join('/') : '-')}</td>
                                <td><span class="badge badge-${t.priority}">${t.priority}</span></td>
                                <td class="${t.overdue_days > 0 ? 'text-overdue' : ''}">${t.due || '-'}</td>
                                <td class="${t.overdue_days > 0 ? 'text-overdue' : ''}">${t.overdue_days > 0 ? '+' + t.overdue_days + '天' : '-'}</td>
                                <td><span class="badge badge-${t.task_status}">${statusLabels[t.task_status] || t.task_status}</span></td>
                            </tr>
                        `).join('')}</tbody>
                    </table>
                </div>`;
        }
        
        // Modal 任務表格 - 含快速下拉篩選（橫向排列）
        let modalTasks = [];  // 儲存當前 modal 的任務
        function modalTableWithFilters(tasks, id = 'modalTableBody') {
            modalTasks = tasks;
            // 取得唯一值
            const modules = [...new Set(tasks.map(t => t.module || '未分類'))].sort();
            const owners = [...new Set(tasks.flatMap(t => t.owners || []))].sort();
            const priorities = ['high', 'medium', 'normal'];
            const statuses = ['in_progress', 'pending', 'completed'];
            
            // 初始化時就渲染所有任務
            const initialRows = tasks.map(t => `
                <tr class="row-${t.task_status} ${t.overdue_days > 0 ? 'row-overdue' : ''}">
                    <td>${t.last_seen || t.mail_date || '-'}</td>
                    <td><span class="badge bg-secondary" style="font-size:0.6rem">${t.module || '-'}</span></td>
                    <td>${t.title} ${t.mail_id ? `<i class="bi bi-envelope ms-1 text-primary" style="cursor:pointer;font-size:0.8rem" onclick="showMailPreview('${t.mail_id}', event)" title="預覽"></i>` : ''}${getAttachmentIcons(t.attachments, t.has_attachments, t.mail_id)}</td>
                    <td>${t.owners_str || (t.owners ? t.owners.join('/') : '-')}</td>
                    <td><span class="badge badge-${t.priority}">${t.priority}</span></td>
                    <td class="${t.overdue_days > 0 ? 'text-overdue' : ''}">${t.due || '-'}</td>
                    <td class="${t.overdue_days > 0 ? 'text-overdue' : ''}">${t.overdue_days > 0 ? '+' + t.overdue_days + '天' : '-'}</td>
                    <td><span class="badge badge-${t.task_status}">${statusLabels[t.task_status] || t.task_status}</span></td>
                </tr>
            `).join('');
            
            return `
                <div class="d-flex flex-wrap gap-2 mb-2 align-items-center">
                    <input type="text" class="form-control form-control-sm" style="width:150px" placeholder="🔍 搜尋..." id="modal_search" onkeyup="filterModalTasks()">
                    <select class="form-select form-select-sm" style="width:130px" id="modal_module" onchange="filterModalTasks()">
                        <option value="">全部模組</option>
                        ${modules.map(m => `<option value="${m}">${m}</option>`).join('')}
                    </select>
                    <select class="form-select form-select-sm" style="width:130px" id="modal_owner" onchange="filterModalTasks()">
                        <option value="">全部負責人</option>
                        ${owners.map(o => `<option value="${o}">${o}</option>`).join('')}
                    </select>
                    <select class="form-select form-select-sm" style="width:110px" id="modal_priority" onchange="filterModalTasks()">
                        <option value="">全部優先</option>
                        ${priorities.map(p => `<option value="${p}">${p}</option>`).join('')}
                    </select>
                    <select class="form-select form-select-sm" style="width:110px" id="modal_status" onchange="filterModalTasks()">
                        <option value="">全部狀態</option>
                        ${statuses.map(s => `<option value="${s}">${statusLabels[s]}</option>`).join('')}
                    </select>
                    <select class="form-select form-select-sm" style="width:110px" id="modal_overdue" onchange="filterModalTasks()">
                        <option value="">全部超期</option>
                        <option value="yes">超期</option>
                        <option value="no">未超期</option>
                    </select>
                    <span id="modal_count" class="small text-muted">共 ${tasks.length} 筆</span>
                </div>
                <div style="max-height: 50vh; overflow-y: auto;">
                    <table class="table table-sm data-table">
                        <thead><tr><th>Mail日期</th><th>模組</th><th>任務</th><th>負責人</th><th>優先級</th><th>Due</th><th>超期</th><th>狀態</th></tr></thead>
                        <tbody id="${id}">${initialRows}</tbody>
                    </table>
                </div>`;
        }
        
        function filterModalTasks() {
            const search = (document.getElementById('modal_search')?.value || '').toLowerCase();
            const module = document.getElementById('modal_module')?.value || '';
            const owner = document.getElementById('modal_owner')?.value || '';
            const priority = document.getElementById('modal_priority')?.value || '';
            const status = document.getElementById('modal_status')?.value || '';
            const overdue = document.getElementById('modal_overdue')?.value || '';
            
            const filtered = modalTasks.filter(t => {
                if (search && !JSON.stringify(t).toLowerCase().includes(search)) return false;
                if (module && (t.module || '未分類') !== module) return false;
                if (owner && !(t.owners || []).includes(owner) && !t.owners_str?.includes(owner)) return false;
                if (priority && t.priority !== priority) return false;
                if (status && t.task_status !== status) return false;
                if (overdue === 'yes' && t.overdue_days <= 0) return false;
                if (overdue === 'no' && t.overdue_days > 0) return false;
                return true;
            });
            
            const tbody = document.querySelector('#modalContent tbody');
            if (tbody) {
                tbody.innerHTML = filtered.map(t => `
                    <tr class="row-${t.task_status} ${t.overdue_days > 0 ? 'row-overdue' : ''}">
                        <td>${t.last_seen || t.mail_date || '-'}</td>
                        <td><span class="badge bg-secondary" style="font-size:0.6rem">${t.module || '-'}</span></td>
                        <td>${t.title} ${t.mail_id ? `<i class="bi bi-envelope ms-1 text-primary" style="cursor:pointer;font-size:0.8rem" onclick="showMailPreview('${t.mail_id}', event)" title="預覽"></i>` : ''}${getAttachmentIcons(t.attachments, t.has_attachments, t.mail_id)}</td>
                        <td>${t.owners_str || (t.owners ? t.owners.join('/') : '-')}</td>
                        <td><span class="badge badge-${t.priority}">${t.priority}</span></td>
                        <td class="${t.overdue_days > 0 ? 'text-overdue' : ''}">${t.due || '-'}</td>
                        <td class="${t.overdue_days > 0 ? 'text-overdue' : ''}">${t.overdue_days > 0 ? '+' + t.overdue_days + '天' : '-'}</td>
                        <td><span class="badge badge-${t.task_status}">${statusLabels[t.task_status] || t.task_status}</span></td>
                    </tr>
                `).join('');
            }
            
            const countEl = document.getElementById('modal_count');
            if (countEl) countEl.textContent = `共 ${filtered.length} 筆`;
        }
        
        function filterModalTable(id, q) { q = q.toLowerCase(); for (let row of document.getElementById(id).rows) row.style.display = row.textContent.toLowerCase().includes(q) ? '' : 'none'; }

        function showAllTasks() { if (!resultData) return; showModal(`全部任務 (${resultData.total_tasks})`, modalTableWithFilters(resultData.all_tasks)); setTimeout(filterModalTasks, 100); }
        function showByStatus(status) { if (!resultData) return; const tasks = resultData.all_tasks.filter(t => t.task_status === status); showModal(`${statusLabels[status]} (${tasks.length})`, modalTableWithFilters(tasks, status + 'Table')); setTimeout(filterModalTasks, 100); }
        function showByPriority(priority) { if (!resultData) return; const tasks = resultData.all_tasks.filter(t => t.priority === priority); showModal(`${priority.toUpperCase()} 優先級 (${tasks.length})`, modalTableWithFilters(tasks, priority + 'Table')); setTimeout(filterModalTasks, 100); }
        function showOverdue() { if (!resultData) return; const tasks = resultData.all_tasks.filter(t => t.overdue_days > 0 && t.task_status !== 'completed'); showModal(`超期任務 (${tasks.length})`, modalTableWithFilters(tasks, 'overdueTable')); setTimeout(filterModalTasks, 100); }
        function showNotOverdue() { if (!resultData) return; const tasks = resultData.all_tasks.filter(t => t.overdue_days <= 0 && t.task_status !== 'completed'); showModal(`未超期任務 (${tasks.length})`, modalTableWithFilters(tasks, 'notOverdueTable')); setTimeout(filterModalTasks, 100); }
        function showMemberOverdueTasks(name) { if (!resultData) return; const tasks = resultData.all_tasks.filter(t => t.owners_str.includes(name) && t.overdue_days > 0); showModal(`${name} 超期任務 (${tasks.length})`, modalTableWithFilters(tasks, 'memberOverdueTable')); setTimeout(filterModalTasks, 100); }
        function showMembers() { if (!resultData) return; showModal('成員列表', resultData.member_list.map(m => `<span class="member-badge" onclick="filterTaskByOwner('${m}')">${m}</span>`).join('')); }
        function showTaskDetail(title) { if (!resultData) return; const tasks = resultData.all_tasks.filter(t => t.title === title); showModal(`任務: ${title}`, modalTableWithFilters(tasks, 'taskDetailTable')); setTimeout(filterModalTasks, 100); }

        // Mail Preview
        async function showMailPreview(mailId, event) {
            if (event) event.stopPropagation();
            if (!mailId) { alert('此任務沒有關聯的 Mail'); return; }
            
            try {
                const r = await fetch(`/api/mail/${mailId}`);
                if (!r.ok) { alert('無法取得 Mail 內容'); return; }
                const mail = await r.json();
                
                document.getElementById('mailSubject').textContent = mail.subject || '-';
                document.getElementById('mailDate').textContent = mail.date || '-';
                document.getElementById('mailTime').textContent = mail.time ? `(${mail.time})` : '';
                
                // 顯示附件（可下載）
                const attachContainer = document.getElementById('mailPreviewAttachments');
                if (mail.attachments && mail.attachments.length > 0) {
                    attachContainer.innerHTML = '<strong class="me-2">附件:</strong>' + mail.attachments.map(att => {
                        const name = att.name || 'attachment';
                        const icon = getFileIcon(name);
                        // 如果有 Base64 資料，使用 downloadAttachment；否則使用 API
                        if (att.data) {
                            const escapedName = name.replace(/'/g, "\\'");
                            return `<span class="badge bg-light text-dark me-1" style="cursor:pointer" onclick="downloadAttachment('${escapedName}', '${att.data}', '${att.mime || 'application/octet-stream'}')" title="點擊下載">${icon} ${name}</span>`;
                        } else {
                            return `<a href="/api/mail/${mailId}/attachment/${att.index}" class="badge bg-light text-dark me-1 text-decoration-none" style="cursor:pointer" title="點擊下載">${icon} ${name}</a>`;
                        }
                    }).join('');
                    attachContainer.style.display = 'block';
                } else {
                    attachContainer.style.display = 'none';
                }
                
                const hasHtml = mail.html_body && mail.html_body.trim().length > 0;
                if (hasHtml) {
                    setMailView('html');
                    document.getElementById('mailPreviewIframe').srcdoc = mail.html_body;
                } else {
                    setMailView('html');
                    const textAsHtml = `<!DOCTYPE html><html><head><meta charset="UTF-8"><style>body{font-family:Segoe UI,Arial,sans-serif;font-size:14px;padding:20px;line-height:1.6;}</style></head><body><pre style="white-space:pre-wrap;font-family:inherit;">${escapeHtml(mail.body || '(無內容)')}</pre></body></html>`;
                    document.getElementById('mailPreviewIframe').srcdoc = textAsHtml;
                }
                document.getElementById('mailBodyText').textContent = mail.body || '(無內容)';
                
                new bootstrap.Modal(document.getElementById('mailModal')).show();
            } catch (e) { alert('錯誤: ' + e); }
        }
        
        // 根據檔名取得圖示
        function getFileIcon(filename) {
            const name = (filename || '').toLowerCase();
            if (name.endsWith('.xlsx') || name.endsWith('.xls') || name.endsWith('.csv')) return '<i class="bi bi-file-earmark-excel text-success"></i>';
            if (name.endsWith('.docx') || name.endsWith('.doc')) return '<i class="bi bi-file-earmark-word text-primary"></i>';
            if (name.endsWith('.pptx') || name.endsWith('.ppt')) return '<i class="bi bi-file-earmark-ppt text-danger"></i>';
            if (name.endsWith('.pdf')) return '<i class="bi bi-file-earmark-pdf text-danger"></i>';
            if (name.endsWith('.png') || name.endsWith('.jpg') || name.endsWith('.jpeg') || name.endsWith('.gif')) return '<i class="bi bi-file-earmark-image text-info"></i>';
            if (name.endsWith('.txt')) return '<i class="bi bi-file-earmark-text text-secondary"></i>';
            if (name.endsWith('.zip') || name.endsWith('.rar') || name.endsWith('.7z')) return '<i class="bi bi-file-earmark-zip text-warning"></i>';
            return '<i class="bi bi-paperclip text-secondary"></i>';
        }
        
        function escapeHtml(text) { const div = document.createElement('div'); div.textContent = text; return div.innerHTML; }
        
        // 下載附件（Base64 資料）
        function downloadAttachment(name, data, mime) {
            const byteCharacters = atob(data);
            const byteNumbers = new Array(byteCharacters.length);
            for (let i = 0; i < byteCharacters.length; i++) {
                byteNumbers[i] = byteCharacters.charCodeAt(i);
            }
            const byteArray = new Uint8Array(byteNumbers);
            const blob = new Blob([byteArray], { type: mime || 'application/octet-stream' });
            const url = URL.createObjectURL(blob);
            const a = document.createElement('a');
            a.href = url;
            a.download = name;
            document.body.appendChild(a);
            a.click();
            document.body.removeChild(a);
            URL.revokeObjectURL(url);
        }
        
        function setMailView(mode) {
            document.getElementById('mailBodyHtml').style.display = mode === 'html' ? 'block' : 'none';
            document.getElementById('mailBodyText').style.display = mode === 'text' ? 'block' : 'none';
            document.getElementById('btnHtml').classList.toggle('active', mode === 'html');
            document.getElementById('btnText').classList.toggle('active', mode === 'text');
        }

        // Review 模式 - 郵件列表
        function renderMailList() {
            const search = document.getElementById('mailSearch').value.toLowerCase();
            const filtered = allMails.filter(m => !search || (m.subject || '').toLowerCase().includes(search) || (m.body || '').toLowerCase().includes(search));
            
            let html = filtered.map((m, i) => {
                // 判斷附件：優先用 attachments 陣列，其次用 has_attachments 或 attachment_count
                const hasAtt = (m.attachments && m.attachments.length > 0) || m.has_attachments || (m.attachment_count > 0);
                const mailId = m.mail_id || '';
                
                // 使用統一的附件圖示函數
                const attIcons = hasAtt ? getAttachmentIcons(m.attachments, hasAtt, mailId) : '';
                
                return `
                <div class="mail-item" onclick="selectMail(${i})" data-mail-id="${mailId}">
                    <div class="mail-subject d-flex align-items-center justify-content-between">
                        <span>${m.subject || '(無主旨)'} ${attIcons}</span>
                        ${mailId ? `<i class="bi bi-box-arrow-up-right text-primary" style="cursor:pointer;font-size:0.8rem" onclick="showMailPreview('${mailId}', event)" title="開啟 Mail 預覽"></i>` : ''}
                    </div>
                    <div class="mail-meta">${m.date} ${m.time || ''} | ${m.sender || ''}</div>
                </div>
            `}).join('');
            
            // 根據模式判斷是否還有更多未載入
            const hasMore = directFolderMode 
                ? (folderMailsLoaded < folderMailsTotal) 
                : (reviewMailsLoaded < reviewMailsTotal);
            const isLoading = directFolderMode ? folderMailsLoading : reviewMailsLoading;
            const loaded = directFolderMode ? folderMailsLoaded : reviewMailsLoaded;
            const total = directFolderMode ? folderMailsTotal : reviewMailsTotal;
            
            if (hasMore && !search) {
                html += `<div class="text-center p-2 text-muted small" id="loadMoreHint">
                    <span class="spinner-border spinner-border-sm me-1" style="display:${isLoading ? 'inline-block' : 'none'}"></span>
                    向下滾動載入更多... (${loaded}/${total})
                </div>`;
            }
            
            document.getElementById('mailList').innerHTML = html || '<div class="p-3 text-muted">無郵件</div>';
        }
        
        function filterMailList() { renderMailList(); }
        
        async function selectMail(index) {
            document.querySelectorAll('.mail-item').forEach(el => el.classList.remove('selected'));
            document.querySelectorAll('.mail-item')[index]?.classList.add('selected');
            
            const mail = allMails[index];
            if (!mail) return;
            
            console.log('[selectMail] mail:', mail.mail_id, 'html_body exists:', !!mail.html_body, 'cid_processed:', !!mail.cid_processed, 'attachments:', mail.attachments?.length || 0);
            
            document.getElementById('mailHeader').style.display = 'block';
            document.getElementById('mailSubjectView').textContent = mail.subject || '-';
            document.getElementById('mailDateView').textContent = `${mail.date} ${mail.time || ''}`;
            
            // 如果已有完整內容且已處理過 CID 圖片，則直接顯示
            if (mail.html_body && mail.html_body.length > 0 && mail.cid_processed && mail.attachments !== undefined) {
                console.log('[selectMail] Using existing data (CID processed)');
                displayMailContent(mail);
                return;
            }
            
            // 否則從 API 取得完整內容（包含 CID 處理後的 html_body 和附件資訊）
            if (mail.mail_id) {
                try {
                    console.log('[selectMail] Fetching from API:', mail.mail_id);
                    const r = await fetch(`/api/mail/${mail.mail_id}`);
                    if (r.ok) {
                        const fullMail = await r.json();
                        console.log('[selectMail] API response html_body len:', (fullMail.html_body || '').length, 'cid_processed:', fullMail.cid_processed, 'attachments:', fullMail.attachments?.length || 0);
                        // 更新本地資料
                        allMails[index] = { ...mail, ...fullMail };
                        displayMailContent(allMails[index]);
                    } else {
                        displayMailContent(mail);
                    }
                } catch (e) {
                    console.error('[selectMail] Error:', e);
                    displayMailContent(mail);
                }
            } else {
                displayMailContent(mail);
            }
        }
        
        function displayMailContent(mail) {
            // 儲存當前郵件 ID
            currentMailId = mail.mail_id;
            
            console.log('[displayMailContent] html_body:', !!mail.html_body, 'len:', (mail.html_body || '').length);
            
            if (mail.html_body && mail.html_body.length > 0) {
                console.log('[displayMailContent] Using HTML mode');
                document.getElementById('mailIframe').srcdoc = mail.html_body;
            } else {
                console.log('[displayMailContent] Using text mode');
                const textHtml = `<!DOCTYPE html><html><head><meta charset="UTF-8"><style>body{font-family:sans-serif;font-size:14px;padding:15px;}</style></head><body><pre style="white-space:pre-wrap;">${escapeHtml(mail.body || '')}</pre></body></html>`;
                document.getElementById('mailIframe').srcdoc = textHtml;
            }
            document.getElementById('mailContentText').textContent = mail.body || '';
            
            // 顯示附件
            const attachmentsRow = document.getElementById('mailAttachmentsRow');
            const attachmentsList = document.getElementById('mailAttachmentsList');
            console.log('[displayMailContent] attachments:', mail.attachments, 'attachment_count:', mail.attachment_count);
            if (mail.attachments && mail.attachments.length > 0) {
                attachmentsRow.style.display = 'block';
                attachmentsList.innerHTML = mail.attachments.map(att => {
                    // 如果有 Base64 資料，使用 downloadAttachment；否則使用 API
                    if (att.data) {
                        const escapedName = (att.name || 'attachment').replace(/'/g, "\\'");
                        return `<span class="badge bg-primary me-1" style="cursor:pointer" onclick="downloadAttachment('${escapedName}', '${att.data}', '${att.mime || 'application/octet-stream'}')" title="${formatFileSize(att.size)} - 點擊下載"><i class="bi bi-download"></i> ${att.name}</span>`;
                    } else {
                        return `<a href="/api/mail/${mail.mail_id}/attachment/${att.index}" class="badge bg-primary me-1 text-decoration-none" style="cursor:pointer" title="${formatFileSize(att.size)} - 點擊下載"><i class="bi bi-download"></i> ${att.name}</a>`;
                    }
                }).join('');
            } else {
                attachmentsRow.style.display = 'none';
            }
            
            // 如果卡片是最大化狀態，重新調整 iframe 高度
            setTimeout(adjustMailContentHeight, 50);
        }
        
        function adjustMailContentHeight() {
            const card = document.getElementById('cardMailContent');
            if (!card || !card.classList.contains('card-fullscreen')) return;
            
            const cardBody = card.querySelector('.card-body');
            const mailHeader = document.getElementById('mailHeader');
            const container = document.getElementById('mailContentHtml');
            const iframe = document.getElementById('mailIframe');
            
            if (!cardBody || !container || !iframe) return;
            
            const bodyHeight = cardBody.offsetHeight;
            const headerHeight = mailHeader ? mailHeader.offsetHeight : 0;
            const availableHeight = bodyHeight - headerHeight;
            
            container.style.cssText = `height: ${availableHeight}px; position: relative; overflow: hidden;`;
            iframe.style.cssText = 'position: absolute; top: 0; left: 0; width: 100%; height: 100%; border: none;';
        }
        
        // 儲存當前選中郵件的 mail_id
        let currentMailId = null;
        
        function formatFileSize(bytes) {
            if (!bytes || bytes === 0) return '0 B';
            const k = 1024;
            const sizes = ['B', 'KB', 'MB', 'GB'];
            const i = Math.floor(Math.log(bytes) / Math.log(k));
            return parseFloat((bytes / Math.pow(k, i)).toFixed(1)) + ' ' + sizes[i];
        }
        
        function setMailViewMode(mode) {
            document.getElementById('mailContentHtml').style.display = mode === 'html' ? 'block' : 'none';
            document.getElementById('mailContentText').style.display = mode === 'text' ? 'block' : 'none';
            document.getElementById('btnMailHtml').classList.toggle('active', mode === 'html');
            document.getElementById('btnMailText').classList.toggle('active', mode === 'text');
        }

        // CSV Export
        function exportTableCSV(table) {
            const state = tableState[table];
            let csv = [], headers, getData;
            
            if (table === 'task') {
                headers = ['Mail日期', '模組', '任務', '負責人', '優先級', 'Due', '超期天數', '狀態'];
                getData = t => [t.last_seen || '', t.module || '', t.title, t.owners_str, t.priority, t.due || '', t.overdue_days || 0, statusLabels[t.task_status]];
            } else if (table === 'member') {
                headers = ['成員', '總數', '完成', '進行', 'Pending', 'High', 'Medium', 'Normal'];
                getData = m => [m.name, m.total, m.completed, m.in_progress, m.pending, m.high, m.medium, m.normal];
            } else {
                headers = ['排名', '成員', '任務數', '基礎分', '超期數', '扣分', '總分'];
                getData = c => [c.rank, c.name, c.task_count, c.base_score, c.overdue_count, c.overdue_penalty, c.score];
            }
            
            csv.push(headers.join(','));
            state.filtered.forEach(item => csv.push(getData(item).map(v => '"' + String(v).replace(/"/g, '""') + '"').join(',')));
            downloadCSV(csv.join('\\n'), table + '.csv');
        }
        
        function exportModalCSV() {
            const table = document.querySelector('#modalContent table');
            if (!table) return;
            let csv = [];
            csv.push(Array.from(table.querySelectorAll('thead th')).map(th => th.textContent.trim()).join(','));
            table.querySelectorAll('tbody tr').forEach(row => csv.push(Array.from(row.cells).map(td => '"' + td.textContent.trim().replace(/"/g, '""') + '"').join(',')));
            downloadCSV(csv.join('\\n'), 'export.csv');
        }
        
        function downloadCSV(content, filename) {
            const blob = new Blob(['\\ufeff' + content], { type: 'text/csv;charset=utf-8' });
            const a = document.createElement('a'); a.href = URL.createObjectURL(blob); a.download = filename; a.click();
        }

        function exportExcel() { window.location.href = '/api/excel'; }
        function exportHTML() { if (!resultData) { alert('請先分析'); return; } window.open('/api/export-html', '_blank'); }

        // 檔案上傳
        const dropZone = document.getElementById('dropZone');
        const fileInput = document.getElementById('fileInput');
        let uploadedFiles = [];  // 儲存上傳的檔案
        
        dropZone.onclick = () => fileInput.click();
        dropZone.ondragover = e => { e.preventDefault(); dropZone.classList.add('dragover'); };
        dropZone.ondragleave = () => dropZone.classList.remove('dragover');
        dropZone.ondrop = e => { e.preventDefault(); dropZone.classList.remove('dragover'); addFiles(e.dataTransfer.files); };
        fileInput.onchange = () => { addFiles(fileInput.files); fileInput.value = ''; };
        
        function addFiles(files) {
            Array.from(files).forEach(f => {
                if (f.name.endsWith('.msg') && !uploadedFiles.find(uf => uf.name === f.name)) {
                    uploadedFiles.push(f);
                }
            });
            renderUploadFileList();
        }
        
        function removeUploadFile(index) {
            uploadedFiles.splice(index, 1);
            renderUploadFileList();
        }
        
        function renderUploadFileList() {
            const el = document.getElementById('uploadFileList');
            if (uploadedFiles.length === 0) {
                el.innerHTML = '<span class="text-muted">尚未選擇檔案</span>';
                return;
            }
            el.innerHTML = uploadedFiles.map((f, i) => `
                <div class="d-flex justify-content-between align-items-center p-1 border-bottom">
                    <span><i class="bi bi-file-earmark-text me-1"></i>${f.name}</span>
                    <button class="btn btn-sm btn-link text-danger p-0" onclick="removeUploadFile(${i})"><i class="bi bi-x-circle"></i></button>
                </div>
            `).join('');
        }
        
        async function analyzeUploadedFiles() {
            if (uploadedFiles.length === 0) { alert('請先選擇 .msg 檔案'); return; }
            
            const formData = new FormData();
            uploadedFiles.forEach(f => formData.append('f', f));
            
            const excludeMiddlePriority = document.getElementById('uploadExcludeMiddlePriority').checked;
            const excludeAfter5pm = document.getElementById('uploadExcludeAfter5pm').checked;
            formData.append('exclude_middle_priority', excludeMiddlePriority);
            formData.append('exclude_after_5pm', excludeAfter5pm);
            
            document.getElementById('loading').style.display = 'flex';
            try {
                const r = await fetch('/api/upload', { method: 'POST', body: formData });
                const data = await r.json();
                if (data.error) throw new Error(data.error);
                resultData = data;
                
                // 標記使用上傳的郵件
                useUploadedMails = true;
                
                // 顯示結果區域
                showResultArea();
                
                // 顯示兩個頁籤
                document.getElementById('tabItem-stats').style.display = 'block';
                document.getElementById('tabItem-review').style.display = 'block';
                
                // 切換到統計頁籤
                const statsTab = document.getElementById('tab-stats');
                const bsTab = new bootstrap.Tab(statsTab);
                bsTab.show();
                
                reviewModeActive = false;
                
                // 更新郵件列表（供後續 Review 使用）
                if (data.mails) {
                    allMailsOriginal = data.mails;
                    allMails = data.mails;
                    reviewMailsTotal = allMails.length;
                    reviewMailsLoaded = allMails.length;
                }
                
                // 先更新 UI 再填充篩選選項
                updateUI();
                
                // 強制渲染郵件列表
                renderMailList();
                updateReviewCount();
                
            } catch (e) { alert('錯誤: ' + e.message); }
            document.getElementById('loading').style.display = 'none';
        }

        // 初始化上傳檔案列表顯示
        renderUploadFileList();

        document.getElementById('detailModal').addEventListener('hidden.bs.modal', () => { currentModal = null; });
        
        // 防抖函數
        function debounce(func, wait) {
            let timeout;
            return function(...args) {
                clearTimeout(timeout);
                timeout = setTimeout(() => func.apply(this, args), wait);
            };
        }
        
        // 進階篩選事件監聯 - 安全檢查
        const filterFieldEl = document.getElementById('filterField');
        const filterKeywordEl = document.getElementById('filterKeyword');
        const filterAttTypeEl = document.getElementById('filterAttType');
        
        if (filterFieldEl) filterFieldEl.addEventListener('change', onFilterChange);
        if (filterKeywordEl) filterKeywordEl.addEventListener('input', debounce(onFilterChange, 300));
        if (filterAttTypeEl) filterAttTypeEl.addEventListener('change', onFilterChange);
        
        // 結果頁籤切換事件 - 安全檢查
        const tabReview = document.getElementById('tab-review');
        const tabStats = document.getElementById('tab-stats');
        
        if (tabReview) {
            tabReview.addEventListener('shown.bs.tab', () => {
                reviewModeActive = true;
                if (allMailsOriginal.length > 0) {
                    applyMailFilters();
                }
            });
        }
        
        if (tabStats) {
            tabStats.addEventListener('shown.bs.tab', () => {
                reviewModeActive = false;
            });
        }
        
        // 頁籤切換事件 - 重置狀態
        document.querySelectorAll('[data-bs-toggle="tab"]').forEach(tab => {
            tab.addEventListener('shown.bs.tab', (e) => {
                const targetId = e.target.getAttribute('data-bs-target');
                console.log('[Tab] Switched to:', targetId);
                
                // 切換到上傳頁籤時，重置 Outlook 模式狀態
                if (targetId === '#tabUpload') {
                    // 保持 selectedEntry 但標記為上傳模式
                    console.log('[Tab] Upload mode');
                }
                // 切換到 Outlook 頁籤時，重置上傳模式狀態
                else if (targetId === '#tabOutlook') {
                    useUploadedMails = false;
                    console.log('[Tab] Outlook mode, useUploadedMails reset');
                }
            });
        });
        
        // 最大化/還原功能
        let currentFullscreenCard = null;
        
        function toggleFullscreen(cardId) {
            const card = document.getElementById(cardId);
            const overlay = document.getElementById('fullscreenOverlay');
            
            if (card.classList.contains('card-fullscreen')) {
                // 還原
                card.classList.remove('card-fullscreen');
                overlay.style.display = 'none';
                currentFullscreenCard = null;
                
                // 還原原始高度
                if (card._originalHeight) {
                    card.style.height = card._originalHeight;
                }
                
                // 還原子元素樣式
                const cardBody = card.querySelector('.card-body');
                if (cardBody && cardBody._originalStyle) {
                    cardBody.style.cssText = cardBody._originalStyle;
                }
                
                // 還原郵件內容區域樣式
                const mailContentHtml = card.querySelector('#mailContentHtml');
                if (mailContentHtml) {
                    mailContentHtml.style.cssText = 'flex:1 1 auto;overflow:hidden;min-height:0;';
                }
                const mailIframe = card.querySelector('#mailIframe');
                if (mailIframe) {
                    mailIframe.style.cssText = 'width:100%;height:100%;border:none;';
                }
                
                // 重繪圖表
                if (chart1) chart1.resize();
                if (chart2) chart2.resize();
                if (chart3) chart3.resize();
                if (chart4) chart4.resize();
            } else {
                // 最大化
                if (currentFullscreenCard) {
                    currentFullscreenCard.classList.remove('card-fullscreen');
                    if (currentFullscreenCard._originalHeight) {
                        currentFullscreenCard.style.height = currentFullscreenCard._originalHeight;
                    }
                }
                
                // 儲存原始高度並清除 inline style
                card._originalHeight = card.style.height;
                card.style.height = '100vh';
                
                // 強制設定 card-body 樣式
                const cardBody = card.querySelector('.card-body');
                if (cardBody) {
                    cardBody._originalStyle = cardBody.style.cssText;
                    cardBody.style.cssText = 'flex: 1 !important; height: 0 !important; min-height: 0 !important; display: flex !important; flex-direction: column !important; overflow: hidden !important;';
                }
                
                // 強制設定郵件內容區域 - 使用相對定位+絕對定位讓 iframe 填滿
                const mailContentHtml = card.querySelector('#mailContentHtml');
                if (mailContentHtml) {
                    mailContentHtml.style.cssText = 'flex: 1 !important; min-height: 0 !important; overflow: hidden !important; position: relative !important;';
                }
                const mailIframe = card.querySelector('#mailIframe');
                if (mailIframe) {
                    mailIframe.style.cssText = 'position: absolute !important; top: 0 !important; left: 0 !important; width: 100% !important; height: 100% !important; border: none !important;';
                }
                
                // 處理純文字模式
                const mailContentText = card.querySelector('#mailContentText');
                if (mailContentText) {
                    mailContentText.style.cssText = mailContentText.style.display === 'none' ? 'display: none;' : 'flex: 1 !important; min-height: 0 !important; overflow-y: auto !important;';
                }
                
                card.classList.add('card-fullscreen');
                overlay.style.display = 'block';
                currentFullscreenCard = card;
                
                // 重繪圖表和調整郵件內容高度
                setTimeout(() => {
                    if (chart1) chart1.resize();
                    if (chart2) chart2.resize();
                    if (chart3) chart3.resize();
                    if (chart4) chart4.resize();
                    // 調整郵件內容高度
                    if (cardId === 'cardMailContent') {
                        adjustMailContentHeight();
                    }
                }, 100);
            }
        }
        
        function exitFullscreen() {
            if (currentFullscreenCard) {
                currentFullscreenCard.classList.remove('card-fullscreen');
                document.getElementById('fullscreenOverlay').style.display = 'none';
                currentFullscreenCard = null;
                
                // 重繪圖表
                if (chart1) chart1.resize();
                if (chart2) chart2.resize();
                if (chart3) chart3.resize();
                if (chart4) chart4.resize();
            }
        }
        
        // ESC 鍵退出全螢幕
        document.addEventListener('keydown', (e) => {
            if (e.key === 'Escape') exitFullscreen();
        });
    </script>
</body>
</html>
'''


def generate_export_html(data, report_date, mail_contents=None, mails_list=None):
    """生成匯出用的 HTML - 包含統計分析和 Review 頁籤"""
    import json
    data_json = json.dumps(data, ensure_ascii=False)
    mail_contents_json = json.dumps(mail_contents or {}, ensure_ascii=False)
    mails_list_json = json.dumps(mails_list or [], ensure_ascii=False)
    
    return f'''<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>Task Report - {report_date}</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
    <link href="https://cdn.jsdelivr.net/npm/bootstrap-icons@1.10.0/font/bootstrap-icons.css" rel="stylesheet">
    <script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
    <style>
        :root {{ --primary: #2E75B6; --primary-dark: #1a4f7a; }}
        body {{ background: #f5f7fa; font-size: 14px; }}
        .navbar {{ background: #2E75B6; }}
        .card {{ border: none; border-radius: 10px; box-shadow: 0 2px 12px rgba(0,0,0,0.08); margin-bottom: 12px; }}
        .card-header {{ background: #2E75B6; color: white; border-radius: 10px 10px 0 0 !important; padding: 8px 12px; display: flex; justify-content: space-between; align-items: center; }}
        .card-header-title {{ font-weight: 500; }}
        .stat-card {{ text-align: center; padding: 10px; cursor: pointer; transition: all 0.2s; height: 85px; display: flex; flex-direction: column; justify-content: center; }}
        .stat-card:hover {{ transform: translateY(-2px); box-shadow: 0 4px 15px rgba(0,0,0,0.15); }}
        .stat-number {{ font-size: 1.5rem; font-weight: bold; color: var(--primary); }}
        .stat-number.danger {{ color: #dc3545; }}
        .stat-number.warning {{ color: #FFA500; }}
        .stat-number.success {{ color: #28a745; }}
        .stat-number.info {{ color: #17a2b8; }}
        .stat-label {{ color: #666; font-size: 0.7rem; }}
        .badge-high {{ background: #FF6B6B !important; }}
        .badge-medium {{ background: #FFE066 !important; color: #333 !important; }}
        .badge-normal {{ background: #74C0FC !important; }}
        .badge-completed {{ background: #28a745 !important; }}
        .badge-pending {{ background: #FFA500 !important; }}
        .badge-in_progress {{ background: #17a2b8 !important; }}
        
        .data-table {{ width: 100%; font-size: 0.8rem; border-collapse: collapse; }}
        .data-table thead th {{ background: #4a4a4a !important; color: white !important; font-weight: 600; cursor: pointer; padding: 8px 5px; border: 1px solid #666; }}
        .data-table tbody td {{ padding: 6px 5px; vertical-align: middle; border: 1px solid #ddd; }}
        .data-table tbody tr:nth-child(even) {{ background: #f9f9f9; }}
        .data-table tbody tr:hover {{ background: #e8f4fc !important; }}
        .data-table tbody tr.row-pending {{ background: #fff8e1; }}
        .data-table tbody tr.row-in_progress {{ background: #e3f2fd; }}
        .data-table tbody tr.row-overdue {{ background: #ffebee; }}
        .table-toolbar {{ display: flex; gap: 8px; padding: 8px 10px; background: #f8f9fa; border-bottom: 1px solid #dee2e6; flex-wrap: wrap; }}
        .table-container {{ overflow-x: auto; height: 400px; overflow-y: auto; }}
        .text-overdue {{ color: #dc3545 !important; font-weight: bold; }}
        
        .footer {{ text-align: center; padding: 12px; color: #999; font-size: 0.7rem; }}
        .rank-badge {{ display: inline-block; width: 22px; height: 22px; line-height: 22px; border-radius: 50%; text-align: center; font-weight: bold; color: white; font-size: 0.7rem; }}
        .rank-1 {{ background: linear-gradient(135deg, #FFD700, #FFA500); }}
        .rank-2 {{ background: linear-gradient(135deg, #C0C0C0, #A0A0A0); }}
        .rank-3 {{ background: linear-gradient(135deg, #CD7F32, #8B4513); }}
        .rank-other {{ background: #6c757d; }}
        .chart-container {{ height: 280px; }}
        .chart-select {{ font-size: 0.75rem; padding: 3px 8px; width: 80px; }}
        
        /* 分頁控制 */
        .pagination-controls {{ display: flex; justify-content: space-between; align-items: center; padding: 8px 10px; background: #f8f9fa; border-top: 1px solid #dee2e6; font-size: 0.75rem; }}
        .pagination-controls button {{ padding: 3px 10px; font-size: 0.75rem; }}
        .pagination-controls select {{ font-size: 0.75rem; padding: 2px 5px; width: 70px; }}
        
        /* 最大化功能 */
        .card-maximize-btn {{ cursor: pointer; opacity: 0.7; font-size: 0.8rem; }}
        .card-maximize-btn:hover {{ opacity: 1; }}
        .card-fullscreen {{ position: fixed !important; top: 0 !important; left: 0 !important; width: 100vw !important; height: 100vh !important; z-index: 9999; border-radius: 0 !important; margin: 0 !important; display: flex !important; flex-direction: column !important; }}
        .card-fullscreen > .card-header {{ flex-shrink: 0 !important; }}
        .card-fullscreen > .card-body, .card-fullscreen .card-body {{ flex: 1 !important; height: 0 !important; min-height: 0 !important; max-height: none !important; overflow: hidden !important; display: flex !important; flex-direction: column !important; }}
        .card-fullscreen .table-container {{ flex: 1 !important; height: 0 !important; min-height: 0 !important; max-height: none !important; }}
        .card-fullscreen .chart-container {{ flex: 1 !important; min-height: 0 !important; }}
        .card-fullscreen #mailContentHtml {{ flex: 1 !important; min-height: 0 !important; max-height: none !important; overflow: hidden !important; position: relative !important; }}
        .card-fullscreen #mailIframe {{ position: absolute !important; top: 0 !important; left: 0 !important; width: 100% !important; height: 100% !important; border: none !important; }}
        .card-fullscreen .mail-preview {{ max-height: none !important; }}
        .fullscreen-overlay {{ position: fixed; top: 0; left: 0; width: 100vw; height: 100vh; background: rgba(0,0,0,0.5); z-index: 9998; display: none; }}
        
        /* 頁籤樣式 */
        #resultTabs .nav-link {{ color: #666; background: #f8f9fa; border: 1px solid #dee2e6; border-bottom: none; margin-right: 2px; font-weight: 600; }}
        #resultTabs .nav-link:hover {{ color: #2E75B6; background: #e8f4fc; }}
        #resultTabs .nav-link.active {{ color: #fff; background: #2E75B6; border-color: #2E75B6; }}
        
        /* Review 樣式 */
        .mail-item {{ padding: 10px; border-bottom: 1px solid #eee; cursor: pointer; }}
        .mail-item:hover {{ background: #f8f9fa; }}
        .mail-item.selected {{ background: #e3f2fd; }}
        .mail-subject {{ font-weight: 500; }}
        .mail-meta {{ font-size: 0.75rem; color: #666; }}
        
        @media print {{ .no-print {{ display: none !important; }} body {{ background: white; }} }}
    </style>
</head>
<body>
    <!-- 全螢幕遮罩 -->
    <div id="fullscreenOverlay" class="fullscreen-overlay" onclick="exitFullscreen()"></div>
    
    <nav class="navbar navbar-dark mb-2 py-1 no-print">
        <div class="container-fluid">
            <span class="navbar-brand mb-0 h6"><i class="bi bi-clipboard-data me-2"></i>Task Report - {report_date}</span>
            <div class="d-flex gap-2">
                <button class="btn btn-outline-light btn-sm" onclick="window.print()"><i class="bi bi-printer me-1"></i>列印</button>
            </div>
        </div>
    </nav>

    <div class="container-fluid">
        <!-- 統計卡片 -->
        <div class="row g-2 mb-2">
            <div class="col"><div class="card stat-card" onclick="showAllTasks()"><div class="stat-number" id="totalTasks">0</div><div class="stat-label">總任務</div></div></div>
            <div class="col"><div class="card stat-card" onclick="showByStatus('pending')"><div class="stat-number warning" id="pendingCount">0</div><div class="stat-label">Pending</div></div></div>
            <div class="col"><div class="card stat-card" onclick="showByStatus('in_progress')"><div class="stat-number info" id="inProgressCount">0</div><div class="stat-label">進行中</div></div></div>
            <div class="col"><div class="card stat-card" onclick="showByStatus('completed')"><div class="stat-number success" id="completedCount">0</div><div class="stat-label">已完成</div></div></div>
            <div class="col"><div class="card stat-card" onclick="showOverdue()"><div class="stat-number danger" id="overdueCount">0</div><div class="stat-label">超期</div></div></div>
        </div>

        <!-- 進度條 -->
        <div class="card mb-2">
            <div class="card-body py-2">
                <div class="progress">
                    <div class="progress-bar bg-success" id="progressCompleted" style="width:0%"></div>
                    <div class="progress-bar bg-info" id="progressInProgress" style="width:0%"></div>
                    <div class="progress-bar bg-warning" id="progressPending" style="width:0%"></div>
                </div>
                <div class="d-flex justify-content-between mt-1 small text-muted">
                    <span>最後郵件: <span id="lastMailDate">-</span></span>
                    <span>成員: <span id="memberCount">0</span></span>
                </div>
            </div>
        </div>

        <!-- 頁籤結構 -->
        <ul class="nav nav-tabs mb-2" id="resultTabs" role="tablist" style="border-bottom: 2px solid #2E75B6;">
            <li class="nav-item">
                <button class="nav-link active" id="tab-stats" data-bs-toggle="tab" data-bs-target="#pane-stats" type="button">
                    <i class="bi bi-bar-chart me-1"></i>統計分析
                </button>
            </li>
            <li class="nav-item">
                <button class="nav-link" id="tab-review" data-bs-toggle="tab" data-bs-target="#pane-review" type="button">
                    <i class="bi bi-eye me-1"></i>Review <span id="reviewMailCount" class="badge bg-warning text-dark ms-1">0</span>
                </button>
            </li>
        </ul>
        
        <div class="tab-content">
            <!-- 統計分析頁籤 -->
            <div class="tab-pane fade show active" id="pane-stats" role="tabpanel">
                <!-- 圖表區 -->
                <div class="row g-2 mb-2">
                    <div class="col-md-6">
                        <div class="card" id="cardChart1">
                            <div class="card-header">
                                <span class="card-header-title"><i class="bi bi-pie-chart me-1"></i>狀態分佈</span>
                                <div class="d-flex align-items-center">
                                    <select class="form-select chart-select me-2" id="chart1Type" onchange="updateChart1()">
                                        <option value="doughnut">環形</option><option value="pie">圓餅</option><option value="bar">長條</option>
                                    </select>
                                    <i class="bi bi-arrows-fullscreen card-maximize-btn text-white" onclick="toggleFullscreen('cardChart1')" title="最大化"></i>
                                </div>
                            </div>
                            <div class="card-body py-2"><div class="chart-container"><canvas id="chart1"></canvas></div></div>
                        </div>
                    </div>
                    <div class="col-md-6">
                        <div class="card" id="cardChart4">
                            <div class="card-header">
                                <span class="card-header-title"><i class="bi bi-person-exclamation me-1"></i>成員超期天數</span>
                                <div class="d-flex align-items-center">
                                    <select class="form-select chart-select me-2" style="width:120px" id="chart4Type" onchange="updateChart4()">
                                        <option value="stacked" selected>水平堆疊</option><option value="vstacked">垂直堆疊</option>
                                    </select>
                                    <i class="bi bi-arrows-fullscreen card-maximize-btn text-white" onclick="toggleFullscreen('cardChart4')" title="最大化"></i>
                                </div>
                            </div>
                            <div class="card-body py-2"><div class="chart-container"><canvas id="chart4"></canvas></div></div>
                        </div>
                    </div>
                </div>
                <div class="row g-2 mb-2">
                    <div class="col-md-6">
                        <div class="card" id="cardChart2">
                            <div class="card-header">
                                <span class="card-header-title"><i class="bi bi-bar-chart me-1"></i>優先級分佈</span>
                                <div class="d-flex align-items-center">
                                    <select class="form-select chart-select me-2" id="chart2Type" onchange="updateChart2()">
                                        <option value="doughnut">環形</option><option value="pie">圓餅</option><option value="bar">長條</option>
                                    </select>
                                    <i class="bi bi-arrows-fullscreen card-maximize-btn text-white" onclick="toggleFullscreen('cardChart2')" title="最大化"></i>
                                </div>
                            </div>
                            <div class="card-body py-2"><div class="chart-container"><canvas id="chart2"></canvas></div></div>
                        </div>
                    </div>
                    <div class="col-md-6">
                        <div class="card" id="cardChart3">
                            <div class="card-header">
                                <span class="card-header-title"><i class="bi bi-exclamation-triangle me-1"></i>超期狀況</span>
                                <div class="d-flex align-items-center">
                                    <select class="form-select chart-select me-2" id="chart3Type" onchange="updateChart3()">
                                        <option value="doughnut">環形</option><option value="pie">圓餅</option><option value="bar">長條</option>
                                    </select>
                                    <i class="bi bi-arrows-fullscreen card-maximize-btn text-white" onclick="toggleFullscreen('cardChart3')" title="最大化"></i>
                                </div>
                            </div>
                            <div class="card-body py-2"><div class="chart-container"><canvas id="chart3"></canvas></div></div>
                        </div>
                    </div>
                </div>

                <!-- 任務列表 -->
                <div class="card mb-2" id="cardTaskList">
                    <div class="card-header">
                        <span class="card-header-title"><i class="bi bi-list-task me-1"></i>任務列表</span>
                        <div class="d-flex align-items-center">
                            <button class="btn btn-outline-light btn-sm me-1" onclick="toggleTaskFilter()"><i class="bi bi-funnel me-1"></i>篩選</button>
                            <button class="btn btn-outline-light btn-sm me-2" onclick="exportTableCSV('task')"><i class="bi bi-download me-1"></i>CSV</button>
                            <i class="bi bi-arrows-fullscreen card-maximize-btn text-white" onclick="toggleFullscreen('cardTaskList')" title="最大化"></i>
                        </div>
                    </div>
                    <div class="table-toolbar" id="taskFilterBar" style="display:none;">
                        <input type="text" class="form-control form-control-sm" style="width:150px" placeholder="🔍 搜尋..." id="taskSearch" onkeyup="filterAndRenderTaskTable()">
                        <select class="form-select form-select-sm" style="width:130px" id="filterModule" onchange="filterAndRenderTaskTable()"><option value="">全部模組</option></select>
                        <select class="form-select form-select-sm" style="width:130px" id="filterOwner" onchange="filterAndRenderTaskTable()"><option value="">全部負責人</option></select>
                        <select class="form-select form-select-sm" style="width:110px" id="filterPriority" onchange="filterAndRenderTaskTable()">
                            <option value="">全部優先</option><option value="high">High</option><option value="medium">Medium</option><option value="normal">Normal</option>
                        </select>
                        <select class="form-select form-select-sm" style="width:110px" id="filterStatus" onchange="filterAndRenderTaskTable()">
                            <option value="">全部狀態</option><option value="in_progress">進行中</option><option value="pending">Pending</option><option value="completed">已完成</option>
                        </select>
                        <select class="form-select form-select-sm" style="width:110px" id="filterOverdue" onchange="filterAndRenderTaskTable()">
                            <option value="">全部超期</option><option value="yes">超期</option><option value="no">未超期</option>
                        </select>
                        <button class="btn btn-outline-secondary btn-sm" onclick="clearTaskFilters()"><i class="bi bi-x-circle"></i> 清除</button>
                    </div>
                    <div class="table-container">
                        <table class="table table-sm data-table mb-0">
                            <thead>
                                <tr>
                                    <th onclick="sortTable('task','last_seen')">Mail日期 ↕</th>
                                    <th onclick="sortTable('task','module')">模組 ↕</th>
                                    <th onclick="sortTable('task','title')">任務 ↕</th>
                                    <th onclick="sortTable('task','owners_str')">負責人 ↕</th>
                                    <th onclick="sortTable('task','priority')">優先級 ↕</th>
                                    <th onclick="sortTable('task','due')">Due ↕</th>
                                    <th onclick="sortTable('task','overdue_days')">超期 ↕</th>
                                    <th onclick="sortTable('task','task_status')">狀態 ↕</th>
                                </tr>
                            </thead>
                            <tbody id="taskTableBody"></tbody>
                        </table>
                    </div>
                    <div class="pagination-controls">
                        <div>
                            <button class="btn btn-outline-secondary btn-sm" onclick="prevPage('task')">上一頁</button>
                            <select class="form-select form-select-sm d-inline-block ms-1" id="taskPageSize" onchange="renderTaskTable()">
                                <option value="30">30</option><option value="50" selected>50</option><option value="100">100</option><option value="200">200</option>
                            </select>
                        </div>
                        <span id="taskPageInfo">-</span>
                        <button class="btn btn-outline-secondary btn-sm" onclick="nextPage('task')">下一頁</button>
                    </div>
                </div>

                <!-- 成員統計 & 貢獻度 -->
                <div class="row g-2">
                    <div class="col-md-7">
                        <div class="card" id="cardMemberStats">
                            <div class="card-header">
                                <span class="card-header-title"><i class="bi bi-people me-1"></i>成員統計</span>
                                <div class="d-flex align-items-center">
                                    <button class="btn btn-outline-light btn-sm me-1" onclick="toggleMemberFilter()"><i class="bi bi-funnel me-1"></i>篩選</button>
                                    <button class="btn btn-outline-light btn-sm me-2" onclick="exportTableCSV('member')"><i class="bi bi-download me-1"></i>CSV</button>
                                    <i class="bi bi-arrows-fullscreen card-maximize-btn text-white" onclick="toggleFullscreen('cardMemberStats')" title="最大化"></i>
                                </div>
                            </div>
                            <div class="table-toolbar" id="memberFilterBar" style="display:none;">
                                <input type="text" class="form-control form-control-sm" style="width:150px" placeholder="🔍 搜尋..." id="memberSearch" onkeyup="filterAndRenderMemberTable()">
                                <select class="form-select form-select-sm" style="width:130px" id="filterMemberModule" onchange="filterAndRenderMemberTable()"><option value="">全部模組</option></select>
                                <select class="form-select form-select-sm" style="width:110px" id="filterMemberPriority" onchange="filterAndRenderMemberTable()">
                                    <option value="">全部優先</option><option value="high">High</option><option value="medium">Medium</option><option value="normal">Normal</option>
                                </select>
                                <select class="form-select form-select-sm" style="width:110px" id="filterMemberTaskStatus" onchange="filterAndRenderMemberTable()">
                                    <option value="">全部狀態</option><option value="in_progress">進行中</option><option value="pending">Pending</option><option value="completed">已完成</option>
                                </select>
                                <select class="form-select form-select-sm" style="width:110px" id="filterMemberOverdue" onchange="filterAndRenderMemberTable()">
                                    <option value="">全部超期</option><option value="hasOverdue">有超期</option><option value="noOverdue">無超期</option>
                                </select>
                                <button class="btn btn-outline-secondary btn-sm" onclick="clearMemberFilters()"><i class="bi bi-x-circle"></i></button>
                            </div>
                            <div class="table-container" style="height:400px;">
                                <table class="table table-sm data-table mb-0">
                                    <thead>
                                        <tr>
                                            <th onclick="sortTable('member','name')">成員 ↕</th>
                                            <th onclick="sortTable('member','total')">總數 ↕</th>
                                            <th onclick="sortTable('member','completed')">完成 ↕</th>
                                            <th onclick="sortTable('member','in_progress')">進行 ↕</th>
                                            <th onclick="sortTable('member','pending')">Pend ↕</th>
                                            <th onclick="sortTable('member','high')">H ↕</th>
                                            <th onclick="sortTable('member','medium')">M ↕</th>
                                            <th onclick="sortTable('member','normal')">N ↕</th>
                                        </tr>
                                    </thead>
                                    <tbody id="memberTableBody"></tbody>
                                </table>
                            </div>
                        </div>
                    </div>
                    <div class="col-md-5">
                        <div class="card" id="cardContrib">
                            <div class="card-header">
                                <span class="card-header-title"><i class="bi bi-trophy me-1"></i>貢獻度 <small class="text-warning">(含超期減分)</small></span>
                                <div class="d-flex align-items-center">
                                    <button class="btn btn-outline-light btn-sm me-1" onclick="toggleContribFilter()"><i class="bi bi-funnel me-1"></i>篩選</button>
                                    <button class="btn btn-outline-light btn-sm me-2" onclick="exportTableCSV('contrib')"><i class="bi bi-download me-1"></i>CSV</button>
                                    <i class="bi bi-arrows-fullscreen card-maximize-btn text-white" onclick="toggleFullscreen('cardContrib')" title="最大化"></i>
                                </div>
                            </div>
                            <div class="table-toolbar" id="contribFilterBar" style="display:none;">
                                <input type="text" class="form-control form-control-sm" style="width:150px" placeholder="🔍 搜尋..." id="contribSearch" onkeyup="filterAndRenderContribTable()">
                                <select class="form-select form-select-sm" style="width:130px" id="filterContribModule" onchange="filterAndRenderContribTable()"><option value="">全部模組</option></select>
                                <select class="form-select form-select-sm" style="width:110px" id="filterContribPriority" onchange="filterAndRenderContribTable()">
                                    <option value="">全部優先</option><option value="high">High</option><option value="medium">Medium</option><option value="normal">Normal</option>
                                </select>
                                <select class="form-select form-select-sm" style="width:110px" id="filterContribTaskStatus" onchange="filterAndRenderContribTable()">
                                    <option value="">全部狀態</option><option value="in_progress">進行中</option><option value="pending">Pending</option><option value="completed">已完成</option>
                                </select>
                                <select class="form-select form-select-sm" style="width:110px" id="filterContribOverdue" onchange="filterAndRenderContribTable()">
                                    <option value="">全部超期</option><option value="hasOverdue">有超期</option><option value="noOverdue">無超期</option>
                                </select>
                                <button class="btn btn-outline-secondary btn-sm" onclick="clearContribFilters()"><i class="bi bi-x-circle"></i></button>
                            </div>
                            <div class="table-container" style="height:400px;">
                                <table class="table table-sm data-table mb-0">
                                    <thead>
                                        <tr>
                                            <th onclick="sortTable('contrib','rank')"># ↕</th>
                                            <th onclick="sortTable('contrib','name')">成員 ↕</th>
                                            <th onclick="sortTable('contrib','task_count')">任務 ↕</th>
                                            <th onclick="sortTable('contrib','base_score')">基礎分 ↕</th>
                                            <th onclick="sortTable('contrib','overdue_count')">超期數 ↕</th>
                                            <th onclick="sortTable('contrib','overdue_penalty')">扣分 ↕</th>
                                            <th onclick="sortTable('contrib','score')">總分 ↕</th>
                                        </tr>
                                    </thead>
                                    <tbody id="contribTableBody"></tbody>
                                </table>
                            </div>
                        </div>
                    </div>
                </div>
            </div>

            <!-- Review 頁籤 -->
            <div class="tab-pane fade" id="pane-review" role="tabpanel">
                <div class="row g-2">
                    <div class="col-md-4">
                        <div class="card" style="height:600px;">
                            <div class="card-header">
                                <span class="card-header-title"><i class="bi bi-envelope me-1"></i>郵件列表</span>
                                <small id="reviewMailCountDetail" class="text-white-50"></small>
                            </div>
                            <div class="table-toolbar">
                                <input type="text" class="form-control form-control-sm" placeholder="🔍 搜尋主旨/寄件者..." id="mailSearch" onkeyup="filterMailList()">
                            </div>
                            <div id="mailList" style="flex:1;overflow-y:auto;"></div>
                        </div>
                    </div>
                    <div class="col-md-8">
                        <div class="card" style="height:600px;">
                            <div class="card-header">
                                <span class="card-header-title"><i class="bi bi-file-text me-1"></i>郵件內容</span>
                                <div class="btn-group btn-group-sm">
                                    <button class="btn btn-outline-light btn-sm active" id="btnMailHtml" onclick="setMailViewMode('html')">HTML</button>
                                    <button class="btn btn-outline-light btn-sm" id="btnMailText" onclick="setMailViewMode('text')">純文字</button>
                                </div>
                            </div>
                            <div id="mailHeader" class="p-2 bg-light border-bottom small" style="display:none;">
                                <div><strong>主旨:</strong> <span id="mailSubjectView">-</span></div>
                                <div><strong>日期:</strong> <span id="mailDateView">-</span></div>
                                <div id="mailAttachmentsRow" style="display:none;"><strong>附件:</strong> <span id="mailAttachmentsList"></span></div>
                            </div>
                            <div id="mailBodyHtml" style="flex:1;overflow:hidden;">
                                <iframe id="mailIframe" style="width:100%;height:100%;border:none;"></iframe>
                            </div>
                            <div id="mailContentText" style="flex:1;overflow-y:auto;padding:15px;font-family:monospace;font-size:13px;white-space:pre-wrap;background:#fafafa;display:none;"></div>
                        </div>
                    </div>
                </div>
            </div>
        </div>
    </div>
    
    <div class="footer">© 2025 Task Dashboard v23 | Powered by Vince</div>

    <!-- Modal -->
    <div class="modal fade" id="detailModal" tabindex="-1">
        <div class="modal-dialog modal-xl">
            <div class="modal-content">
                <div class="modal-header py-2">
                    <h6 class="modal-title" id="modalTitle"></h6>
                    <button type="button" class="btn-close" data-bs-dismiss="modal"></button>
                </div>
                <div class="modal-body" id="modalContent"></div>
                <div class="modal-footer py-1">
                    <button class="btn btn-outline-primary btn-sm" onclick="exportModalCSV()"><i class="bi bi-download me-1"></i>CSV</button>
                    <button type="button" class="btn btn-secondary btn-sm" data-bs-dismiss="modal">關閉</button>
                </div>
            </div>
        </div>
    </div>
    
    <!-- Mail Preview Modal -->
    <div class="modal fade" id="mailModal" tabindex="-1">
        <div class="modal-dialog modal-xl">
            <div class="modal-content">
                <div class="modal-header py-2 bg-primary text-white">
                    <h6 class="modal-title"><i class="bi bi-envelope me-1"></i>Mail 預覽</h6>
                    <button type="button" class="btn-close btn-close-white" data-bs-dismiss="modal"></button>
                </div>
                <div class="modal-body p-0">
                    <div class="p-2 bg-light border-bottom d-flex justify-content-between align-items-center">
                        <div>
                            <div><strong>主旨：</strong><span id="mailSubject">-</span></div>
                            <div><strong>日期：</strong><span id="mailDate">-</span> <span id="mailTime" class="text-muted"></span></div>
                            <div id="mailPreviewAttachments" style="display:none;" class="mt-1"></div>
                        </div>
                        <div class="btn-group btn-group-sm">
                            <button class="btn btn-outline-secondary active" onclick="setMailView('html')" id="btnHtml">HTML</button>
                            <button class="btn btn-outline-secondary" onclick="setMailView('text')" id="btnText">純文字</button>
                        </div>
                    </div>
                    <div id="mailBodyHtml" style="height:60vh;overflow:hidden;">
                        <iframe id="mailPreviewIframe" style="width:100%;height:100%;border:none;"></iframe>
                    </div>
                    <div id="mailBodyText" style="max-height:60vh;overflow-y:auto;padding:15px;font-family:monospace;font-size:13px;white-space:pre-wrap;background:#fafafa;display:none;"></div>
                </div>
                <div class="modal-footer py-1">
                    <button type="button" class="btn btn-secondary btn-sm" data-bs-dismiss="modal">關閉</button>
                </div>
            </div>
        </div>
    </div>

    <script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/js/bootstrap.bundle.min.js"></script>
    <script>
        // 數據
        const resultData = {data_json};
        const mailContents = {mail_contents_json};
        const allMails = {mails_list_json};
        
        const statusLabels = {{ completed: '已完成', pending: 'Pending', in_progress: '進行中' }};
        let chart1 = null, chart2 = null, chart3 = null, chart4 = null, currentModal = null, modalTasks = [];
        let currentFullscreenCard = null;
        let mailViewMode = 'html';
        
        // 表格狀態
        let tableState = {{
            task: {{ data: [], filtered: [], sortKey: 'last_seen', sortDir: -1, page: 0, pageSize: 50 }},
            member: {{ data: [], filtered: [], sortKey: 'total', sortDir: -1 }},
            contrib: {{ data: [], filtered: [], sortKey: 'rank', sortDir: 1 }}
        }};
        
        // 填充篩選選項
        function fillFilterOptions() {{
            const modules = [...new Set(resultData.all_tasks.map(t => t.module).filter(m => m))].sort();
            const owners = [...new Set(resultData.all_tasks.flatMap(t => t.owners || []).filter(o => o))].sort();
            
            const filterModule = document.getElementById('filterModule');
            const filterOwner = document.getElementById('filterOwner');
            const filterMemberModule = document.getElementById('filterMemberModule');
            const filterContribModule = document.getElementById('filterContribModule');
            
            const moduleOptions = '<option value="">全部模組</option>' + modules.map(m => `<option value="${{m}}">${{m}}</option>`).join('');
            
            if (filterModule) filterModule.innerHTML = moduleOptions;
            if (filterOwner) filterOwner.innerHTML = '<option value="">全部負責人</option>' + owners.map(o => `<option value="${{o}}">${{o}}</option>`).join('');
            if (filterMemberModule) filterMemberModule.innerHTML = moduleOptions;
            if (filterContribModule) filterContribModule.innerHTML = moduleOptions;
        }}
        
        // 初始化
        function updateUI() {{
            if (!resultData) return;
            
            document.getElementById('totalTasks').textContent = resultData.total_tasks || 0;
            document.getElementById('pendingCount').textContent = resultData.pending_count || 0;
            document.getElementById('inProgressCount').textContent = resultData.in_progress_count || 0;
            document.getElementById('completedCount').textContent = resultData.completed_count || 0;
            document.getElementById('overdueCount').textContent = resultData.overdue_count || 0;
            document.getElementById('lastMailDate').textContent = resultData.last_mail_date || '-';
            document.getElementById('memberCount').textContent = resultData.unique_members || 0;
            document.getElementById('reviewMailCount').textContent = allMails.length;
            document.getElementById('reviewMailCountDetail').textContent = `共 ${{allMails.length}} 封`;
            
            const total = resultData.total_tasks || 1;
            document.getElementById('progressCompleted').style.width = (resultData.completed_count / total * 100) + '%';
            document.getElementById('progressInProgress').style.width = (resultData.in_progress_count / total * 100) + '%';
            document.getElementById('progressPending').style.width = (resultData.pending_count / total * 100) + '%';
            
            // 初始化表格
            tableState.task.data = resultData.all_tasks || [];
            tableState.task.filtered = [...tableState.task.data];
            tableState.member.data = resultData.members || [];
            tableState.member.filtered = [...tableState.member.data];
            tableState.contrib.data = resultData.contribution || [];
            tableState.contrib.filtered = [...tableState.contrib.data];
            
            fillFilterOptions();
            renderTaskTable();
            renderMemberTable();
            renderContribTable();
            renderMailList();
            
            updateChart1();
            updateChart2();
            updateChart3();
            updateChart4();
        }}
        
        // 表格渲染 - 附件圖示（可點擊開啟 Mail 預覽）
        function getAttachmentIcons(attachments, hasAttachments, mailId = null) {{
            const clickAttr = mailId ? `style="cursor:pointer;font-size:0.75rem" onclick="showMailPreview('${{mailId}}', event)"` : `style="font-size:0.75rem"`;
            
            if (attachments && attachments.length > 0) {{
                const icons = [];
                let hasExcel = false, hasWord = false, hasPpt = false, hasPdf = false, hasImage = false, hasOther = false;
                attachments.forEach(att => {{
                    const name = (att.name || '').toLowerCase();
                    if (name.endsWith('.xlsx') || name.endsWith('.xls') || name.endsWith('.csv')) hasExcel = true;
                    else if (name.endsWith('.docx') || name.endsWith('.doc')) hasWord = true;
                    else if (name.endsWith('.pptx') || name.endsWith('.ppt')) hasPpt = true;
                    else if (name.endsWith('.pdf')) hasPdf = true;
                    else if (name.endsWith('.png') || name.endsWith('.jpg') || name.endsWith('.jpeg') || name.endsWith('.gif')) hasImage = true;
                    else hasOther = true;
                }});
                if (hasExcel) icons.push(`<i class="bi bi-file-earmark-excel text-success" ${{clickAttr}} title="Excel 附件"></i>`);
                if (hasWord) icons.push(`<i class="bi bi-file-earmark-word text-primary" ${{clickAttr}} title="Word 附件"></i>`);
                if (hasPpt) icons.push(`<i class="bi bi-file-earmark-ppt text-danger" ${{clickAttr}} title="PPT 附件"></i>`);
                if (hasPdf) icons.push(`<i class="bi bi-file-earmark-pdf text-danger" ${{clickAttr}} title="PDF 附件"></i>`);
                if (hasImage) icons.push(`<i class="bi bi-file-earmark-image text-info" ${{clickAttr}} title="圖片附件"></i>`);
                if (hasOther) icons.push(`<i class="bi bi-paperclip text-secondary" ${{clickAttr}} title="其他附件"></i>`);
                return icons.length > 0 ? `<span class="ms-1">${{icons.join('')}}</span>` : '';
            }}
            if (hasAttachments) return `<i class="bi bi-paperclip ms-1 text-secondary" ${{clickAttr}} title="有附件"></i>`;
            return '';
        }}
        
        function renderTaskTable() {{
            const state = tableState.task;
            state.pageSize = parseInt(document.getElementById('taskPageSize')?.value || 50);
            const start = state.page * state.pageSize;
            const pageData = state.filtered.slice(start, start + state.pageSize);
            
            document.getElementById('taskTableBody').innerHTML = pageData.map(t => `
                <tr class="row-${{t.task_status}} ${{t.overdue_days > 0 ? 'row-overdue' : ''}}">
                    <td>${{t.last_seen || '-'}}</td>
                    <td><span class="badge bg-secondary" style="font-size:0.65rem">${{t.module || '-'}}</span></td>
                    <td>
                        <span style="cursor:pointer" onclick="showTaskDetail('${{esc(t.title)}}')">${{t.title}}</span>
                        ${{t.mail_id ? `<i class="bi bi-envelope ms-1 text-primary" style="cursor:pointer;font-size:0.8rem" onclick="showMailPreview('${{t.mail_id}}', event)" title="預覽"></i>` : ''}}
                        ${{getAttachmentIcons(t.attachments, t.has_attachments, t.mail_id)}}
                    </td>
                    <td>${{t.owners_str}}</td>
                    <td><span class="badge badge-${{t.priority}}">${{t.priority}}</span></td>
                    <td class="${{t.overdue_days > 0 ? 'text-overdue' : ''}}">${{t.due || '-'}}</td>
                    <td class="${{t.overdue_days > 0 ? 'text-overdue' : ''}}">${{t.overdue_days > 0 ? '+' + t.overdue_days + '天' : '-'}}</td>
                    <td><span class="badge badge-${{t.task_status}}">${{statusLabels[t.task_status]}}</span></td>
                </tr>
            `).join('');
            
            const totalPages = Math.ceil(state.filtered.length / state.pageSize) || 1;
            const pageInfoEl = document.getElementById('taskPageInfo');
            if (pageInfoEl) pageInfoEl.textContent = `第 ${{state.page + 1}}/${{totalPages}} 頁 (共 ${{state.filtered.length}} 筆)`;
        }}
        
        function renderMemberTable() {{
            document.getElementById('memberTableBody').innerHTML = tableState.member.filtered.map(m => `
                <tr>
                    <td><strong style="cursor:pointer" onclick="showMemberTasks('${{esc(m.name)}}')">${{m.name}}</strong></td>
                    <td style="cursor:pointer" onclick="showMemberTasks('${{esc(m.name)}}')">${{m.total}}</td>
                    <td style="cursor:pointer" onclick="showMemberTasksByStatus('${{esc(m.name)}}', 'completed')"><span class="badge badge-completed">${{m.completed}}</span></td>
                    <td style="cursor:pointer" onclick="showMemberTasksByStatus('${{esc(m.name)}}', 'in_progress')"><span class="badge badge-in_progress">${{m.in_progress}}</span></td>
                    <td style="cursor:pointer" onclick="showMemberTasksByStatus('${{esc(m.name)}}', 'pending')"><span class="badge badge-pending">${{m.pending}}</span></td>
                    <td style="cursor:pointer" onclick="showMemberTasksByPriority('${{esc(m.name)}}', 'high')"><span class="badge badge-high">${{m.high}}</span></td>
                    <td style="cursor:pointer" onclick="showMemberTasksByPriority('${{esc(m.name)}}', 'medium')"><span class="badge badge-medium">${{m.medium}}</span></td>
                    <td style="cursor:pointer" onclick="showMemberTasksByPriority('${{esc(m.name)}}', 'normal')"><span class="badge badge-normal">${{m.normal}}</span></td>
                </tr>
            `).join('');
        }}
        
        function renderContribTable() {{
            document.getElementById('contribTableBody').innerHTML = tableState.contrib.filtered.map(c => `
                <tr>
                    <td><span class="rank-badge ${{c.rank <= 3 ? 'rank-' + c.rank : 'rank-other'}}">${{c.rank}}</span></td>
                    <td style="cursor:pointer" onclick="showMemberTasks('${{esc(c.name)}}')">${{c.name}}</td>
                    <td style="cursor:pointer" onclick="showMemberTasks('${{esc(c.name)}}')">${{c.task_count}}</td>
                    <td style="cursor:pointer" onclick="showContribDetail('${{esc(c.name)}}')">${{c.base_score}}</td>
                    <td class="${{c.overdue_count > 0 ? 'text-overdue' : ''}}" style="cursor:pointer" onclick="showMemberOverdueTasks('${{esc(c.name)}}')">${{c.overdue_count || 0}}</td>
                    <td class="${{c.overdue_penalty > 0 ? 'text-overdue' : ''}}" style="cursor:pointer" onclick="showContribDetail('${{esc(c.name)}}')">-${{c.overdue_penalty}}</td>
                    <td style="cursor:pointer" onclick="showContribDetail('${{esc(c.name)}}')"><strong>${{c.score}}</strong></td>
                </tr>
            `).join('');
        }}
        
        // 郵件列表
        function renderMailList() {{
            const search = document.getElementById('mailSearch')?.value?.toLowerCase() || '';
            const filtered = allMails.filter(m => !search || (m.subject || '').toLowerCase().includes(search) || (m.sender || '').toLowerCase().includes(search));
            
            document.getElementById('mailList').innerHTML = filtered.map((m, i) => {{
                const hasAtt = (m.attachments && m.attachments.length > 0) || m.has_attachments;
                const mailId = m.mail_id || '';
                // 使用統一的附件圖示函數，點擊可開啟 Mail 預覽
                const attIcons = hasAtt ? getAttachmentIcons(m.attachments, m.has_attachments, mailId) : '';
                return `
                <div class="mail-item" onclick="selectMail(${{i}})" data-index="${{i}}">
                    <div class="mail-subject d-flex align-items-center justify-content-between">
                        <span>${{m.subject || '(無主旨)'}} ${{attIcons}}</span>
                        ${{mailId ? `<i class="bi bi-box-arrow-up-right text-primary" style="cursor:pointer;font-size:0.8rem" onclick="showMailPreview('${{mailId}}', event)" title="預覽"></i>` : ''}}
                    </div>
                    <div class="mail-meta">${{m.date}} ${{m.time || ''}} | ${{m.sender || ''}}</div>
                </div>
            `}}).join('') || '<div class="p-3 text-muted">無郵件</div>';
        }}
        
        function selectMail(index) {{
            document.querySelectorAll('.mail-item').forEach(el => el.classList.remove('selected'));
            document.querySelectorAll('.mail-item')[index]?.classList.add('selected');
            
            const mail = allMails[index];
            if (!mail) return;
            
            // 從 mailContents 取得完整內容
            const fullMail = mailContents[mail.mail_id] || mail;
            
            document.getElementById('mailHeader').style.display = 'block';
            document.getElementById('mailSubjectView').textContent = fullMail.subject || mail.subject || '-';
            document.getElementById('mailDateView').textContent = `${{fullMail.date || mail.date}} ${{fullMail.time || mail.time || ''}}`;
            
            // 附件 - 優先使用 fullMail
            const attRow = document.getElementById('mailAttachmentsRow');
            const attList = document.getElementById('mailAttachmentsList');
            const attachments = fullMail.attachments || mail.attachments || [];
            if (attachments.length > 0) {{
                attRow.style.display = 'block';
                attList.innerHTML = attachments.map(att => {{
                    if (att.data) {{
                        return `<span class="badge bg-primary me-1" style="cursor:pointer" onclick="downloadAttachment('${{att.name.replace(/'/g, "\\\\'")}}', '${{att.data}}', '${{att.mime}}')" title="點擊下載"><i class="bi bi-download me-1"></i>${{att.name}}</span>`;
                    }} else {{
                        return `<span class="badge bg-secondary me-1" title="離線模式">${{att.name}}</span>`;
                    }}
                }}).join('');
            }} else {{
                attRow.style.display = 'none';
            }}
            
            // 內容
            if (fullMail.html_body && fullMail.html_body.length > 0) {{
                document.getElementById('mailIframe').srcdoc = fullMail.html_body;
            }} else {{
                const textHtml = `<!DOCTYPE html><html><head><meta charset="UTF-8"><style>body{{font-family:sans-serif;font-size:14px;padding:15px;}}</style></head><body><pre style="white-space:pre-wrap;">${{escapeHtml(fullMail.body || mail.body || '')}}</pre></body></html>`;
                document.getElementById('mailIframe').srcdoc = textHtml;
            }}
            document.getElementById('mailContentText').textContent = fullMail.body || mail.body || '';
        }}
        
        function filterMailList() {{ renderMailList(); }}
        
        function setMailViewMode(mode) {{
            mailViewMode = mode;
            document.getElementById('mailBodyHtml').style.display = mode === 'html' ? 'block' : 'none';
            document.getElementById('mailContentText').style.display = mode === 'text' ? 'block' : 'none';
            document.getElementById('btnMailHtml').classList.toggle('active', mode === 'html');
            document.getElementById('btnMailText').classList.toggle('active', mode === 'text');
        }}
        
        // 排序
        function sortTable(table, key) {{
            const state = tableState[table];
            if (state.sortKey === key) state.sortDir *= -1;
            else {{ state.sortKey = key; state.sortDir = 1; }}
            
            state.filtered.sort((a, b) => {{
                let va = a[key], vb = b[key];
                if (va == null) va = '';
                if (vb == null) vb = '';
                if (typeof va === 'number') return (va - vb) * state.sortDir;
                return String(va).localeCompare(String(vb)) * state.sortDir;
            }});
            
            if (table === 'task') renderTaskTable();
            else if (table === 'member') renderMemberTable();
            else renderContribTable();
        }}
        
        function filterTaskTable() {{
            const search = document.getElementById('taskSearch').value.toLowerCase();
            tableState.task.filtered = tableState.task.data.filter(t => !search || JSON.stringify(t).toLowerCase().includes(search));
            renderTaskTable();
        }}
        
        function esc(s) {{ return String(s || '').replace(/'/g, "\\\\'").replace(/"/g, '&quot;'); }}
        function escapeHtml(text) {{ const div = document.createElement('div'); div.textContent = text; return div.innerHTML; }}
        
        // Modal 功能
        function showModal(title, content) {{
            document.getElementById('modalTitle').textContent = title;
            document.getElementById('modalContent').innerHTML = content;
            currentModal = new bootstrap.Modal(document.getElementById('detailModal'));
            currentModal.show();
        }}
        
        function modalTableWithFilters(tasks) {{
            modalTasks = tasks;
            // 取得唯一值
            const modules = [...new Set(tasks.map(t => t.module || '未分類'))].sort();
            const owners = [...new Set(tasks.flatMap(t => t.owners || []))].sort();
            const priorities = ['high', 'medium', 'normal'];
            const statuses = ['in_progress', 'pending', 'completed'];
            
            return `
                <div class="d-flex flex-wrap gap-2 mb-2 align-items-center">
                    <input type="text" class="form-control form-control-sm" style="width:150px" placeholder="🔍 搜尋..." id="modal_search" onkeyup="filterModalTasks()">
                    <select class="form-select form-select-sm" style="width:130px" id="modal_module" onchange="filterModalTasks()">
                        <option value="">全部模組</option>
                        ${{modules.map(m => `<option value="${{m}}">${{m}}</option>`).join('')}}
                    </select>
                    <select class="form-select form-select-sm" style="width:130px" id="modal_owner" onchange="filterModalTasks()">
                        <option value="">全部負責人</option>
                        ${{owners.map(o => `<option value="${{o}}">${{o}}</option>`).join('')}}
                    </select>
                    <select class="form-select form-select-sm" style="width:110px" id="modal_priority" onchange="filterModalTasks()">
                        <option value="">全部優先</option>
                        ${{priorities.map(p => `<option value="${{p}}">${{p}}</option>`).join('')}}
                    </select>
                    <select class="form-select form-select-sm" style="width:110px" id="modal_status" onchange="filterModalTasks()">
                        <option value="">全部狀態</option>
                        ${{statuses.map(s => `<option value="${{s}}">${{statusLabels[s]}}</option>`).join('')}}
                    </select>
                    <select class="form-select form-select-sm" style="width:110px" id="modal_overdue" onchange="filterModalTasks()">
                        <option value="">全部超期</option>
                        <option value="yes">超期</option>
                        <option value="no">未超期</option>
                    </select>
                    <span id="modal_count" class="small text-muted">共 ${{tasks.length}} 筆</span>
                </div>
                <div style="max-height:50vh;overflow-y:auto;">
                    <table class="table table-sm data-table">
                        <thead><tr><th>Mail日期</th><th>模組</th><th>任務</th><th>負責人</th><th>優先級</th><th>Due</th><th>超期</th><th>狀態</th></tr></thead>
                        <tbody id="modalTableBody">${{tasks.map(t => `
                            <tr class="row-${{t.task_status}} ${{t.overdue_days > 0 ? 'row-overdue' : ''}}">
                                <td>${{t.last_seen || '-'}}</td>
                                <td><span class="badge bg-secondary" style="font-size:0.6rem">${{t.module || '-'}}</span></td>
                                <td>${{t.title}} ${{t.mail_id ? `<i class="bi bi-envelope ms-1 text-primary" style="cursor:pointer;font-size:0.8rem" onclick="showMailPreview('${{t.mail_id}}', event)"></i>` : ''}}${{getAttachmentIcons(t.attachments, t.has_attachments, t.mail_id)}}</td>
                                <td>${{t.owners_str || '-'}}</td>
                                <td><span class="badge badge-${{t.priority}}">${{t.priority}}</span></td>
                                <td class="${{t.overdue_days > 0 ? 'text-overdue' : ''}}">${{t.due || '-'}}</td>
                                <td class="${{t.overdue_days > 0 ? 'text-overdue' : ''}}">${{t.overdue_days > 0 ? '+' + t.overdue_days + '天' : '-'}}</td>
                                <td><span class="badge badge-${{t.task_status}}">${{statusLabels[t.task_status]}}</span></td>
                            </tr>
                        `).join('')}}</tbody>
                    </table>
                </div>`;
        }}
        
        function filterModalTasks() {{
            const search = (document.getElementById('modal_search')?.value || '').toLowerCase();
            const module = document.getElementById('modal_module')?.value || '';
            const owner = document.getElementById('modal_owner')?.value || '';
            const priority = document.getElementById('modal_priority')?.value || '';
            const status = document.getElementById('modal_status')?.value || '';
            const overdue = document.getElementById('modal_overdue')?.value || '';
            
            const filtered = modalTasks.filter(t => {{
                if (search && !JSON.stringify(t).toLowerCase().includes(search)) return false;
                if (module && (t.module || '未分類') !== module) return false;
                if (owner && !(t.owners || []).includes(owner) && !t.owners_str?.includes(owner)) return false;
                if (priority && t.priority !== priority) return false;
                if (status && t.task_status !== status) return false;
                if (overdue === 'yes' && t.overdue_days <= 0) return false;
                if (overdue === 'no' && t.overdue_days > 0) return false;
                return true;
            }});
            
            document.getElementById('modal_count').textContent = `共 ${{filtered.length}} 筆`;
            document.getElementById('modalTableBody').innerHTML = filtered.map(t => `
                <tr class="row-${{t.task_status}} ${{t.overdue_days > 0 ? 'row-overdue' : ''}}">
                    <td>${{t.last_seen || '-'}}</td>
                    <td><span class="badge bg-secondary" style="font-size:0.6rem">${{t.module || '-'}}</span></td>
                    <td>${{t.title}} ${{t.mail_id ? `<i class="bi bi-envelope ms-1 text-primary" style="cursor:pointer;font-size:0.8rem" onclick="showMailPreview('${{t.mail_id}}', event)"></i>` : ''}}${{getAttachmentIcons(t.attachments, t.has_attachments, t.mail_id)}}</td>
                    <td>${{t.owners_str || '-'}}</td>
                    <td><span class="badge badge-${{t.priority}}">${{t.priority}}</span></td>
                    <td class="${{t.overdue_days > 0 ? 'text-overdue' : ''}}">${{t.due || '-'}}</td>
                    <td class="${{t.overdue_days > 0 ? 'text-overdue' : ''}}">${{t.overdue_days > 0 ? '+' + t.overdue_days + '天' : '-'}}</td>
                    <td><span class="badge badge-${{t.task_status}}">${{statusLabels[t.task_status]}}</span></td>
                </tr>
            `).join('');
        }}
        
        function showAllTasks() {{ showModal(`全部任務 (${{resultData.total_tasks}})`, modalTableWithFilters(resultData.all_tasks)); }}
        function showByStatus(status) {{ const tasks = resultData.all_tasks.filter(t => t.task_status === status); showModal(`${{statusLabels[status]}} (${{tasks.length}})`, modalTableWithFilters(tasks)); }}
        function showOverdue() {{ const tasks = resultData.all_tasks.filter(t => t.overdue_days > 0 && t.task_status !== 'completed'); showModal(`超期任務 (${{tasks.length}})`, modalTableWithFilters(tasks)); }}
        function showTaskDetail(title) {{ const tasks = resultData.all_tasks.filter(t => t.title === title); showModal(`任務: ${{title}}`, modalTableWithFilters(tasks)); }}
        function showMemberTasks(name) {{ const tasks = resultData.all_tasks.filter(t => t.owners_str.includes(name)); showModal(`${{name}} 的任務 (${{tasks.length}})`, modalTableWithFilters(tasks)); }}
        function showMemberTasksByStatus(name, status) {{ const tasks = resultData.all_tasks.filter(t => t.owners_str.includes(name) && t.task_status === status); showModal(`${{name}} - ${{statusLabels[status]}} (${{tasks.length}})`, modalTableWithFilters(tasks)); }}
        function showMemberTasksByPriority(name, priority) {{ const tasks = resultData.all_tasks.filter(t => t.owners_str.includes(name) && t.priority === priority); showModal(`${{name}} - ${{priority.toUpperCase()}} (${{tasks.length}})`, modalTableWithFilters(tasks)); }}
        function showContribDetail(name) {{ 
            const c = tableState.contrib.filtered.find(x => x.name === name) || resultData.contribution.find(x => x.name === name); 
            if (!c) return; 
            const detail = `
                <div class="p-3">
                    <div class="d-flex align-items-center mb-3">
                        <i class="bi bi-person-circle fs-2 text-primary me-2"></i>
                        <h5 class="mb-0">${{name}} 貢獻度計算明細</h5>
                    </div>
                    <table class="table table-sm data-table mb-0">
                        <tbody>
                            <tr><td class="fw-bold" style="width:50%">任務數</td><td>${{c.task_count}}</td></tr>
                            <tr><td>High 任務 × 3</td><td><span class="badge badge-high me-1">${{c.high}}</span>× 3 = ${{c.high * 3}}</td></tr>
                            <tr><td>Medium 任務 × 2</td><td><span class="badge badge-medium me-1">${{c.medium}}</span>× 2 = ${{c.medium * 2}}</td></tr>
                            <tr><td>Normal 任務 × 1</td><td><span class="badge badge-normal me-1">${{c.normal}}</span>× 1 = ${{c.normal}}</td></tr>
                            <tr class="table-active"><td class="fw-bold">基礎分</td><td class="fw-bold">${{c.base_score}}</td></tr>
                        </tbody>
                    </table>
                    <table class="table table-sm data-table mt-2 mb-0">
                        <tbody>
                            <tr class="row-overdue"><td style="width:50%">超期任務數</td><td>${{c.overdue_count}}</td></tr>
                            <tr class="row-overdue"><td>總超期天數</td><td>${{c.overdue_days}}</td></tr>
                            <tr class="row-overdue"><td class="fw-bold">扣分 (天數 × 0.1 × -1)</td><td class="fw-bold text-danger">-${{c.overdue_penalty}}</td></tr>
                        </tbody>
                    </table>
                    <table class="table table-sm data-table mt-2 mb-0">
                        <tbody>
                            <tr style="background:#d4edda"><td style="width:50%" class="fw-bold fs-5">總分</td><td class="fw-bold fs-5 text-success">${{c.score}}</td></tr>
                        </tbody>
                    </table>
                    <div class="text-muted small mt-3">
                        <i class="bi bi-info-circle me-1"></i>計算公式: 總分 = 基礎分 - 扣分 = ${{c.base_score}} - ${{c.overdue_penalty}} = ${{c.score}}
                    </div>
                </div>
            `;
            showModal(`${{name}} 貢獻度明細`, detail);
        }}
        function showByPriority(priority) {{ const tasks = resultData.all_tasks.filter(t => t.priority === priority); showModal(`${{priority.toUpperCase()}} 優先級 (${{tasks.length}})`, modalTableWithFilters(tasks)); }}
        function showNotOverdue() {{ const tasks = resultData.all_tasks.filter(t => t.overdue_days <= 0 && t.task_status !== 'completed'); showModal(`未超期任務 (${{tasks.length}})`, modalTableWithFilters(tasks)); }}
        function showMemberOverdueTasks(name) {{ const tasks = resultData.all_tasks.filter(t => t.owners_str.includes(name) && t.overdue_days > 0); showModal(`${{name}} 超期任務 (${{tasks.length}})`, modalTableWithFilters(tasks)); }}
        
        // Mail Preview
        function showMailPreview(mailId, event) {{
            if (event) event.stopPropagation();
            
            // 優先從 mailContents 取得，否則從 allMails 尋找
            let mail = mailContents[mailId];
            if (!mail) {{
                mail = allMails.find(m => m.mail_id === mailId);
            }}
            if (!mail) {{ 
                alert('無法取得 Mail 內容'); 
                return; 
            }}
            
            document.getElementById('mailSubject').textContent = mail.subject || '-';
            document.getElementById('mailDate').textContent = mail.date || '-';
            document.getElementById('mailTime').textContent = mail.time ? `(${{mail.time}})` : '';
            
            // 附件（如果有 Base64 資料則可下載）
            const attachContainer = document.getElementById('mailPreviewAttachments');
            if (mail.attachments && mail.attachments.length > 0) {{
                attachContainer.innerHTML = '<strong class="me-2">附件:</strong>' + mail.attachments.map(att => {{
                    if (att.data) {{
                        // 有 Base64 資料，可下載
                        return `<span class="badge bg-primary me-1" style="cursor:pointer" onclick="downloadAttachment('${{att.name.replace(/'/g, "\\\\'")}}', '${{att.data}}', '${{att.mime}}')" title="點擊下載"><i class="bi bi-download me-1"></i>${{att.name}}</span>`;
                    }} else {{
                        // 無資料，僅顯示名稱
                        return `<span class="badge bg-secondary me-1" title="離線模式無法下載">${{att.name}}</span>`;
                    }}
                }}).join('');
                attachContainer.style.display = 'block';
            }} else {{
                attachContainer.style.display = 'none';
            }}
            
            if (mail.html_body && mail.html_body.trim().length > 0) {{
                setMailView('html');
                document.getElementById('mailPreviewIframe').srcdoc = mail.html_body;
            }} else {{
                setMailView('html');
                const textAsHtml = `<!DOCTYPE html><html><head><meta charset="UTF-8"><style>body{{font-family:Segoe UI,Arial,sans-serif;font-size:14px;padding:20px;}}</style></head><body><pre style="white-space:pre-wrap;">${{escapeHtml(mail.body || '')}}</pre></body></html>`;
                document.getElementById('mailPreviewIframe').srcdoc = textAsHtml;
            }}
            document.getElementById('mailBodyText').textContent = mail.body || '';
            
            new bootstrap.Modal(document.getElementById('mailModal')).show();
        }}
        
        // 下載附件 (Base64)
        function downloadAttachment(name, data, mime) {{
            const link = document.createElement('a');
            link.href = `data:${{mime}};base64,${{data}}`;
            link.download = name;
            link.click();
        }}
        
        function setMailView(mode) {{
            document.getElementById('mailBodyHtml').style.display = mode === 'html' ? 'block' : 'none';
            document.getElementById('mailBodyText').style.display = mode === 'text' ? 'block' : 'none';
            document.getElementById('btnHtml').classList.toggle('active', mode === 'html');
            document.getElementById('btnText').classList.toggle('active', mode === 'text');
        }}
        
        // CSV 匯出
        function exportTableCSV(table) {{
            let csv = [], headers = [], data = [];
            if (table === 'task') {{
                headers = ['Mail日期', '模組', '任務', '負責人', '優先級', 'Due', '超期天數', '狀態'];
                data = tableState.task.filtered.map(t => [t.last_seen || '', t.module || '', t.title, t.owners_str || '', t.priority, t.due || '', t.overdue_days, statusLabels[t.task_status]]);
            }} else if (table === 'member') {{
                headers = ['成員', '總數', '完成', '進行中', 'Pending', 'High', 'Medium', 'Normal'];
                data = tableState.member.filtered.map(m => [m.name, m.total, m.completed, m.in_progress, m.pending, m.high, m.medium, m.normal]);
            }} else if (table === 'contrib') {{
                headers = ['排名', '成員', '任務數', '基礎分', '扣分', '總分'];
                data = tableState.contrib.filtered.map(c => [c.rank, c.name, c.task_count, c.base_score, c.overdue_penalty, c.score]);
            }}
            csv.push(headers.join(','));
            data.forEach(row => csv.push(row.map(v => '"' + String(v).replace(/"/g, '""') + '"').join(',')));
            downloadCSV(csv.join('\\n'), table + '_export.csv');
        }}
        
        function downloadCSV(content, filename) {{
            const blob = new Blob(['\\ufeff' + content], {{ type: 'text/csv;charset=utf-8' }});
            const a = document.createElement('a'); a.href = URL.createObjectURL(blob); a.download = filename; a.click();
        }}
        
        function exportModalCSV() {{
            if (modalTasks.length === 0) return;
            let csv = ['Mail日期,模組,任務,負責人,優先級,Due,超期天數,狀態'];
            modalTasks.forEach(t => csv.push([t.last_seen || '', t.module || '', '"' + t.title.replace(/"/g, '""') + '"', t.owners_str || '', t.priority, t.due || '', t.overdue_days, statusLabels[t.task_status]].join(',')));
            downloadCSV(csv.join('\\n'), 'modal_export.csv');
        }}
        
        // 圖表
        function updateChart1() {{
            const type = document.getElementById('chart1Type').value;
            if (chart1) chart1.destroy();
            chart1 = new Chart(document.getElementById('chart1'), {{
                type: type,
                data: {{ labels: ['進行中', 'Pending', '已完成'], datasets: [{{ data: [resultData.in_progress_count, resultData.pending_count, resultData.completed_count], backgroundColor: ['#17a2b8', '#FFA500', '#28a745'] }}] }},
                options: {{ maintainAspectRatio: false, plugins: {{ legend: {{ display: type !== 'bar', position: 'right' }} }}, onClick: (e, el) => {{ if (el.length) showByStatus(['in_progress', 'pending', 'completed'][el[0].index]); }} }}
            }});
        }}
        
        function updateChart2() {{
            const type = document.getElementById('chart2Type').value;
            if (chart2) chart2.destroy();
            chart2 = new Chart(document.getElementById('chart2'), {{
                type: type,
                data: {{ labels: ['High', 'Medium', 'Normal'], datasets: [{{ data: [resultData.priority_counts.high, resultData.priority_counts.medium, resultData.priority_counts.normal], backgroundColor: ['#FF6B6B', '#FFE066', '#74C0FC'] }}] }},
                options: {{ maintainAspectRatio: false, plugins: {{ legend: {{ display: type !== 'bar', position: 'right' }} }}, onClick: (e, el) => {{ if (el.length) showByPriority(['high', 'medium', 'normal'][el[0].index]); }} }}
            }});
        }}
        
        function updateChart3() {{
            const type = document.getElementById('chart3Type').value;
            if (chart3) chart3.destroy();
            const notOverdueCount = (resultData.not_overdue_count !== undefined) ? resultData.not_overdue_count : (resultData.total_tasks - resultData.overdue_count - resultData.completed_count);
            chart3 = new Chart(document.getElementById('chart3'), {{
                type: type,
                data: {{ labels: ['超期', '未超期'], datasets: [{{ data: [resultData.overdue_count, notOverdueCount], backgroundColor: ['#dc3545', '#28a745'] }}] }},
                options: {{ maintainAspectRatio: false, plugins: {{ legend: {{ display: type !== 'bar', position: 'right' }} }}, onClick: (e, el) => {{ if (el.length && el[0].index === 0) showOverdue(); else if (el.length && el[0].index === 1) showNotOverdue(); }} }}
            }});
        }}
        
        function updateChart4() {{
            const type = document.getElementById('chart4Type').value;
            const ctx = document.getElementById('chart4');
            if (chart4) chart4.destroy();
            
            const overdueData = resultData.contribution.filter(c => c.overdue_days > 0).sort((a, b) => b.overdue_days - a.overdue_days).slice(0, 10);
            const labels = overdueData.map(c => c.name);
            
            if (overdueData.length === 0) {{
                chart4 = new Chart(ctx, {{ type: 'bar', data: {{ labels: ['無超期'], datasets: [{{ data: [0], backgroundColor: '#28a745' }}] }}, options: {{ maintainAspectRatio: false, plugins: {{ legend: {{ display: false }} }} }} }});
                return;
            }}
            
            if (type === 'vstacked') {{
                chart4 = new Chart(ctx, {{
                    type: 'bar',
                    data: {{ labels, datasets: [
                        {{ label: '已完成超期', data: overdueData.map(c => c.completed_overdue_days || 0), backgroundColor: '#6c757d', stack: 's' }},
                        {{ label: '未完成超期', data: overdueData.map(c => c.active_overdue_days || 0), backgroundColor: '#dc3545', stack: 's' }}
                    ]}},
                    options: {{ maintainAspectRatio: false, plugins: {{ legend: {{ display: true }} }}, scales: {{ x: {{ stacked: true }}, y: {{ stacked: true, beginAtZero: true }} }}, onClick: (e, el) => {{ if (el.length) showMemberOverdueTasks(labels[el[0].index]); }} }}
                }});
            }} else {{
                chart4 = new Chart(ctx, {{
                    type: 'bar',
                    data: {{ labels, datasets: [
                        {{ label: '已完成超期', data: overdueData.map(c => c.completed_overdue_days || 0), backgroundColor: '#6c757d', stack: 's' }},
                        {{ label: '未完成超期', data: overdueData.map(c => c.active_overdue_days || 0), backgroundColor: '#dc3545', stack: 's' }}
                    ]}},
                    options: {{ maintainAspectRatio: false, indexAxis: 'y', plugins: {{ legend: {{ display: true }} }}, scales: {{ x: {{ stacked: true, beginAtZero: true }}, y: {{ stacked: true }} }}, onClick: (e, el) => {{ if (el.length) showMemberOverdueTasks(labels[el[0].index]); }} }}
                }});
            }}
        }}
        
        // 分頁控制
        function prevPage(table) {{ if (tableState[table].page > 0) {{ tableState[table].page--; if (table === 'task') renderTaskTable(); }} }}
        function nextPage(table) {{ 
            const state = tableState[table];
            const totalPages = Math.ceil(state.filtered.length / state.pageSize);
            if (state.page < totalPages - 1) {{ state.page++; if (table === 'task') renderTaskTable(); }}
        }}
        
        // 篩選切換
        function toggleTaskFilter() {{ const bar = document.getElementById('taskFilterBar'); bar.style.display = bar.style.display === 'none' ? 'flex' : 'none'; }}
        function toggleMemberFilter() {{ const bar = document.getElementById('memberFilterBar'); bar.style.display = bar.style.display === 'none' ? 'flex' : 'none'; }}
        function toggleContribFilter() {{ const bar = document.getElementById('contribFilterBar'); bar.style.display = bar.style.display === 'none' ? 'flex' : 'none'; }}
        
        // 清除篩選
        function clearTaskFilters() {{
            document.getElementById('taskSearch').value = '';
            document.getElementById('filterModule').value = '';
            document.getElementById('filterOwner').value = '';
            document.getElementById('filterPriority').value = '';
            document.getElementById('filterStatus').value = '';
            document.getElementById('filterOverdue').value = '';
            filterAndRenderTaskTable();
        }}
        
        // 任務表格篩選
        function filterAndRenderTaskTable() {{
            const search = (document.getElementById('taskSearch')?.value || '').toLowerCase();
            const module = document.getElementById('filterModule')?.value || '';
            const owner = document.getElementById('filterOwner')?.value || '';
            const priority = document.getElementById('filterPriority')?.value || '';
            const status = document.getElementById('filterStatus')?.value || '';
            const overdue = document.getElementById('filterOverdue')?.value || '';
            
            tableState.task.filtered = tableState.task.data.filter(t => {{
                if (search && !JSON.stringify(t).toLowerCase().includes(search)) return false;
                if (module && t.module !== module) return false;
                if (owner && !(t.owners || []).includes(owner)) return false;
                if (priority && t.priority !== priority) return false;
                if (status && t.task_status !== status) return false;
                if (overdue === 'yes' && t.overdue_days <= 0) return false;
                if (overdue === 'no' && t.overdue_days > 0) return false;
                return true;
            }});
            tableState.task.page = 0;
            renderTaskTable();
        }}
        
        // 成員表格篩選
        function filterAndRenderMemberTable() {{
            const search = (document.getElementById('memberSearch')?.value || '').toLowerCase();
            const module = document.getElementById('filterMemberModule')?.value || '';
            const priority = document.getElementById('filterMemberPriority')?.value || '';
            const status = document.getElementById('filterMemberTaskStatus')?.value || '';
            const overdue = document.getElementById('filterMemberOverdue')?.value || '';
            
            // 根據篩選條件重新計算成員統計
            const filteredTasks = resultData.all_tasks.filter(t => {{
                if (module && t.module !== module) return false;
                if (priority && t.priority !== priority) return false;
                if (status && t.task_status !== status) return false;
                if (overdue === 'hasOverdue' && t.overdue_days <= 0) return false;
                if (overdue === 'noOverdue' && t.overdue_days > 0) return false;
                return true;
            }});
            
            // 重新計算成員統計
            const memberMap = {{}};
            filteredTasks.forEach(t => {{
                (t.owners || []).forEach(name => {{
                    if (!memberMap[name]) memberMap[name] = {{ name, total: 0, completed: 0, in_progress: 0, pending: 0, high: 0, medium: 0, normal: 0 }};
                    memberMap[name].total++;
                    memberMap[name][t.task_status]++;
                    memberMap[name][t.priority]++;
                }});
            }});
            
            tableState.member.data = Object.values(memberMap);
            tableState.member.filtered = tableState.member.data.filter(m => !search || m.name.toLowerCase().includes(search));
            renderMemberTable();
        }}
        
        // 貢獻度表格篩選
        function filterAndRenderContribTable() {{
            const search = (document.getElementById('contribSearch')?.value || '').toLowerCase();
            const module = document.getElementById('filterContribModule')?.value || '';
            const priority = document.getElementById('filterContribPriority')?.value || '';
            const status = document.getElementById('filterContribTaskStatus')?.value || '';
            const overdue = document.getElementById('filterContribOverdue')?.value || '';
            
            // 先根據模組和優先級篩選所有任務（用於計算任務數，包含 pending）
            let allFilteredTasks = resultData.all_tasks.filter(t => {{
                if (module && t.module !== module) return false;
                if (priority && t.priority !== priority) return false;
                if (status && t.task_status !== status) return false;
                return true;
            }});
            
            // 用於計算分數的任務（排除 pending）
            let scoringTasks = allFilteredTasks.filter(t => t.task_status !== 'pending');
            
            // 重新計算貢獻度
            const contribMap = {{}};
            
            // 先計算任務數（包含 pending）
            allFilteredTasks.forEach(t => {{
                (t.owners || []).forEach(name => {{
                    if (!contribMap[name]) contribMap[name] = {{ name, task_count: 0, high: 0, medium: 0, normal: 0, overdue_count: 0, overdue_days: 0 }};
                    contribMap[name].task_count++;
                }});
            }});
            
            // 再計算分數（排除 pending）
            scoringTasks.forEach(t => {{
                (t.owners || []).forEach(name => {{
                    if (!contribMap[name]) contribMap[name] = {{ name, task_count: 0, high: 0, medium: 0, normal: 0, overdue_count: 0, overdue_days: 0 }};
                    contribMap[name][t.priority]++;
                    if (t.overdue_days > 0 && t.task_status !== 'completed') {{
                        contribMap[name].overdue_count++;
                        contribMap[name].overdue_days += t.overdue_days;
                    }}
                }});
            }});
            
            // 計算分數
            let contribList = Object.values(contribMap).map(c => {{
                c.base_score = c.high * 3 + c.medium * 2 + c.normal * 1;
                c.overdue_penalty = Math.round(c.overdue_days * 0.1 * 10) / 10;
                c.score = Math.round((c.base_score - c.overdue_penalty) * 10) / 10;
                return c;
            }});
            
            // 篩選超期
            if (overdue === 'hasOverdue') contribList = contribList.filter(c => c.overdue_count > 0);
            if (overdue === 'noOverdue') contribList = contribList.filter(c => c.overdue_count === 0);
            
            // 排序和排名
            contribList.sort((a, b) => b.score - a.score);
            contribList.forEach((c, i) => c.rank = i + 1);
            
            tableState.contrib.data = contribList;
            tableState.contrib.filtered = contribList.filter(c => !search || c.name.toLowerCase().includes(search));
            renderContribTable();
        }}
        
        function clearMemberFilters() {{ 
            document.getElementById('memberSearch').value = ''; 
            const filterMemberModule = document.getElementById('filterMemberModule');
            if (filterMemberModule) filterMemberModule.value = '';
            const filterMemberPriority = document.getElementById('filterMemberPriority');
            if (filterMemberPriority) filterMemberPriority.value = '';
            const filterMemberTaskStatus = document.getElementById('filterMemberTaskStatus');
            if (filterMemberTaskStatus) filterMemberTaskStatus.value = '';
            const filterMemberOverdue = document.getElementById('filterMemberOverdue');
            if (filterMemberOverdue) filterMemberOverdue.value = '';
            // 還原原始資料
            tableState.member.data = resultData.members || [];
            tableState.member.filtered = [...tableState.member.data];
            renderMemberTable(); 
        }}
        function clearContribFilters() {{ 
            document.getElementById('contribSearch').value = ''; 
            const filterContribModule = document.getElementById('filterContribModule');
            if (filterContribModule) filterContribModule.value = '';
            const filterContribPriority = document.getElementById('filterContribPriority');
            if (filterContribPriority) filterContribPriority.value = '';
            const filterContribTaskStatus = document.getElementById('filterContribTaskStatus');
            if (filterContribTaskStatus) filterContribTaskStatus.value = '';
            const filterContribOverdue = document.getElementById('filterContribOverdue');
            if (filterContribOverdue) filterContribOverdue.value = '';
            filterAndRenderContribTable(); 
        }}
        
        // 最大化功能
        function toggleFullscreen(cardId) {{
            const card = document.getElementById(cardId);
            const overlay = document.getElementById('fullscreenOverlay');
            
            if (card.classList.contains('card-fullscreen')) {{
                // 還原
                card.classList.remove('card-fullscreen');
                overlay.style.display = 'none';
                currentFullscreenCard = null;
                
                // 還原卡片高度
                if (card._originalHeight) {{
                    card.style.height = card._originalHeight;
                }} else {{
                    card.style.height = '';
                }}
                
                // 還原 card-body 樣式
                const cardBody = card.querySelector('.card-body');
                if (cardBody) {{
                    if (cardBody._originalStyle) {{
                        cardBody.style.cssText = cardBody._originalStyle;
                    }} else {{
                        cardBody.style.cssText = '';
                    }}
                }}
                
                // 還原 chart-container
                const chartContainer = card.querySelector('.chart-container');
                if (chartContainer) {{
                    chartContainer.style.cssText = '';
                }}
                
                // 延遲重繪圖表
                setTimeout(() => {{
                    if (chart1) chart1.resize();
                    if (chart2) chart2.resize();
                    if (chart3) chart3.resize();
                    if (chart4) chart4.resize();
                }}, 50);
            }} else {{
                // 最大化
                if (currentFullscreenCard) {{
                    // 先還原之前最大化的卡片
                    currentFullscreenCard.classList.remove('card-fullscreen');
                    if (currentFullscreenCard._originalHeight) {{
                        currentFullscreenCard.style.height = currentFullscreenCard._originalHeight;
                    }} else {{
                        currentFullscreenCard.style.height = '';
                    }}
                    const prevCardBody = currentFullscreenCard.querySelector('.card-body');
                    if (prevCardBody && prevCardBody._originalStyle) {{
                        prevCardBody.style.cssText = prevCardBody._originalStyle;
                    }} else if (prevCardBody) {{
                        prevCardBody.style.cssText = '';
                    }}
                }}
                
                // 儲存原始狀態
                card._originalHeight = card.style.height || '';
                card.style.height = '100vh';
                
                const cardBody = card.querySelector('.card-body');
                if (cardBody) {{
                    cardBody._originalStyle = cardBody.style.cssText || '';
                    cardBody.style.cssText = 'flex: 1 !important; height: 0 !important; min-height: 0 !important; display: flex !important; flex-direction: column !important; overflow: hidden !important;';
                }}
                
                // 圖表容器
                const chartContainer = card.querySelector('.chart-container');
                if (chartContainer) {{
                    chartContainer.style.cssText = 'flex: 1 !important; min-height: 0 !important;';
                }}
                
                card.classList.add('card-fullscreen');
                overlay.style.display = 'block';
                currentFullscreenCard = card;
                
                setTimeout(() => {{
                    if (chart1) chart1.resize();
                    if (chart2) chart2.resize();
                    if (chart3) chart3.resize();
                    if (chart4) chart4.resize();
                }}, 100);
            }}
        }}
        
        function exitFullscreen() {{
            if (currentFullscreenCard) {{
                currentFullscreenCard.classList.remove('card-fullscreen');
                document.getElementById('fullscreenOverlay').style.display = 'none';
                if (currentFullscreenCard._originalHeight) currentFullscreenCard.style.height = currentFullscreenCard._originalHeight;
                currentFullscreenCard = null;
                if (chart1) chart1.resize();
                if (chart2) chart2.resize();
                if (chart3) chart3.resize();
                if (chart4) chart4.resize();
            }}
        }}
        
        // 初始化
        updateUI();
    </script>
</body>
</html>'''

@app.route('/')
def index():
    return render_template_string(HTML, tree=FOLDER_TREE, fc=len(FOLDERS))

@app.route('/api/outlook', methods=['POST'])
def api_outlook():
    global LAST_RESULT, LAST_DATA, MAIL_CONTENTS, LAST_MAILS_LIST, MAIL_ENTRIES
    MAIL_CONTENTS.clear()
    MAIL_ENTRIES.clear()
    LAST_MAILS_LIST = []
    try:
        j = request.json
        exclude_middle_priority = j.get('exclude_middle_priority', True)
        exclude_after_5pm = j.get('exclude_after_5pm', True)
        include_mails = j.get('include_mails', False)
        
        msgs = get_messages(j['entry_id'], j['store_id'], j['start'], j['end'], exclude_after_5pm)
        parser = TaskParser(exclude_middle_priority=exclude_middle_priority)
        for m in msgs:
            parser.parse(m['subject'], m['body'], m['date'], m.get('time', ''), m.get('html_body', ''), 
                        m.get('has_attachments', False), m.get('attachments', []), m.get('mail_id'))
        stats = Stats()
        for t in parser.tasks:
            stats.add(t)
        LAST_RESULT = stats
        LAST_DATA = stats.summary()
        LAST_MAILS_LIST = msgs  # 儲存郵件列表供匯出用
        
        # 加入郵件列表（用於 Review 模式）
        result = dict(LAST_DATA)
        result['mails'] = msgs
        
        return jsonify(result)
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/upload', methods=['POST'])
def api_upload():
    global LAST_RESULT, LAST_DATA, MAIL_CONTENTS, LAST_MAILS_LIST, MAIL_ENTRIES
    MAIL_CONTENTS.clear()
    MAIL_ENTRIES.clear()  # 清除舊的 entry（上傳時不需要）
    LAST_MAILS_LIST = []  # 清除舊的郵件列表
    
    exclude_middle_priority = request.form.get('exclude_middle_priority', 'true').lower() == 'true'
    exclude_after_5pm = request.form.get('exclude_after_5pm', 'true').lower() == 'true'
    
    parser = TaskParser(exclude_middle_priority=exclude_middle_priority)
    mails = []
    
    import hashlib
    for f in request.files.getlist('f'):
        if not f.filename.endswith('.msg'): continue
        try:
            # 建立暫存檔案並關閉，讓 Outlook 可以開啟
            tmp_path = tempfile.mktemp(suffix='.msg')
            f.save(tmp_path)
            
            # 優先使用 Outlook COM 讀取 .msg（可以正確處理 RTF 轉 HTML）
            html_body = ""
            body = ""
            subject = ""
            mail_time = None
            sender = ""
            attachments_info = []
            outlook_success = False
            
            if HAS_OUTLOOK:
                try:
                    outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
                    msg = outlook.OpenSharedItem(tmp_path)
                    
                    subject = msg.Subject or ""
                    body = msg.Body or ""
                    html_body = msg.HTMLBody or ""
                    mail_time = msg.ReceivedTime if hasattr(msg, 'ReceivedTime') else msg.SentOn
                    sender = str(msg.SenderName) if hasattr(msg, 'SenderName') else ""
                    
                    # 取得附件資訊並處理 CID 圖片，同時保存 Base64 供匯出使用
                    cid_images = {}
                    if hasattr(msg, 'Attachments') and msg.Attachments.Count > 0:
                        import base64
                        import re
                        import mimetypes
                        
                        for i in range(1, msg.Attachments.Count + 1):
                            att = msg.Attachments.Item(i)
                            att_name = att.FileName if hasattr(att, 'FileName') else f"attachment_{i}"
                            att_size = att.Size if hasattr(att, 'Size') else 0
                            
                            # 檢查 Content-ID
                            content_id = ""
                            try:
                                content_id = att.PropertyAccessor.GetProperty("http://schemas.microsoft.com/mapi/proptag/0x3712001F")
                            except:
                                pass
                            
                            # 儲存附件並讀取 Base64（供匯出 HTML 使用）
                            att_b64_data = ""
                            att_mime_type = ""
                            try:
                                att_tmp = tempfile.mktemp(suffix=os.path.splitext(att_name)[1])
                                att.SaveAsFile(att_tmp)
                                with open(att_tmp, 'rb') as attf:
                                    att_data = attf.read()
                                os.unlink(att_tmp)
                                
                                att_b64_data = base64.b64encode(att_data).decode('utf-8')
                                att_mime_type, _ = mimetypes.guess_type(att_name)
                                if not att_mime_type:
                                    att_mime_type = 'application/octet-stream'
                            except Exception as att_err:
                                print(f"[Upload] Error reading attachment {att_name}: {att_err}")
                            
                            # 如果是圖片且有 Content-ID，處理 CID
                            is_image = att_name.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp', '.webp'))
                            if is_image and content_id and att_b64_data:
                                cid_key = content_id.strip('<>') if content_id else att_name
                                cid_images[cid_key] = f"data:{att_mime_type};base64,{att_b64_data}"
                            
                            attachments_info.append({
                                "index": i,
                                "name": att_name,
                                "size": att_size,
                                "data": att_b64_data,  # Base64 資料供匯出使用
                                "mime": att_mime_type
                            })
                        
                        # 替換 HTML 中的 cid: 連結
                        if cid_images and html_body:
                            for cid, data_url in cid_images.items():
                                html_body = re.sub(f'src=["\']cid:{re.escape(cid)}["\']', f'src="{data_url}"', html_body, flags=re.IGNORECASE)
                                html_body = re.sub(f'src=["\']cid:{re.escape(cid.split("@")[0] if "@" in cid else cid)}["\']', f'src="{data_url}"', html_body, flags=re.IGNORECASE)
                            print(f"[Upload] Replaced {len(cid_images)} CID images")
                    
                    outlook_success = True
                    print(f"[Upload] Via Outlook COM: {subject[:50]}, HTML len={len(html_body)}")
                except Exception as outlook_err:
                    print(f"[Upload] Outlook COM failed: {outlook_err}")
                    html_body = ""
                
                # 如果 Outlook COM 失敗，使用 extract_msg
                if not outlook_success and HAS_EXTRACT_MSG:
                    try:
                        msg = extract_msg.Message(tmp_path)
                        subject = msg.subject or ""
                        body = msg.body or ""
                        mail_time = msg.date
                        sender = str(msg.sender) if hasattr(msg, 'sender') else ""
                        
                        # 取得附件資訊
                        if hasattr(msg, 'attachments') and msg.attachments:
                            for i, att in enumerate(msg.attachments, 1):
                                att_name = att.longFilename or att.shortFilename or f"attachment_{i}"
                                attachments_info.append({
                                    "index": i,
                                    "name": att_name,
                                    "size": len(att.data) if hasattr(att, 'data') and att.data else 0
                                })
                        
                        print(f"[Upload] Via extract_msg: {subject[:50]}")
                        
                        # 嘗試取得 HTML（可能會失敗）
                        try:
                            raw_html = msg.htmlBody
                            if raw_html:
                                if isinstance(raw_html, bytes):
                                    html_body = raw_html.decode('utf-8', errors='ignore')
                                else:
                                    html_body = str(raw_html)
                        except:
                            pass
                    except Exception as msg_err:
                        print(f"[Upload] extract_msg failed: {msg_err}")
                
                # 如果還是沒有 HTML，將純文字轉為 HTML
                if not html_body and body:
                    import html as html_module
                    escaped_body = html_module.escape(body)
                    html_body = f'''<!DOCTYPE html>
<html><head><meta charset="UTF-8"></head>
<body style="font-family: Calibri, Arial, sans-serif; font-size: 14px; padding: 20px;">
<pre style="white-space: pre-wrap; font-family: inherit;">{escaped_body}</pre>
</body></html>'''
                    print(f"[Upload] Converted text to HTML, len={len(html_body)}")
                
                if exclude_after_5pm and mail_time and hasattr(mail_time, 'hour'):
                    if mail_time.hour >= 17:
                        try:
                            os.unlink(tmp_path)
                        except:
                            pass
                        continue
                
                mail_date_str = mail_time.strftime("%Y-%m-%d") if mail_time else ""
                mail_time_str = mail_time.strftime("%H:%M") if mail_time else ""
                
                # 生成 mail_id
                mail_id = hashlib.md5(f"{mail_date_str}_{mail_time_str}_{subject}".encode()).hexdigest()[:12]
                
                # 檢查是否有附件
                has_attachments = len(attachments_info) > 0
                if has_attachments:
                    print(f"[Upload] Attachments: {len(attachments_info)}")
                
                # 存入 MAIL_CONTENTS（供 API 和匯出使用）
                MAIL_CONTENTS[mail_id] = {
                    "subject": subject,
                    "body": body,
                    "html_body": html_body,
                    "date": mail_date_str,
                    "time": mail_time_str,
                    "attachments": attachments_info,
                    "cid_processed": True  # 已處理 CID 圖片
                }
                
                parser.parse(subject, body, mail_date_str, mail_time_str, html_body, has_attachments, attachments_info, mail_id)
                
                mails.append({
                    "mail_id": mail_id,
                    "subject": subject,
                    "body": body,
                    "html_body": html_body,
                    "date": mail_date_str,
                    "time": mail_time_str,
                    "sender": sender,
                    "has_attachments": has_attachments,
                    "attachments": attachments_info,
                    "cid_processed": True  # 已處理 CID 圖片
                })
                
                # 清理暫存檔
                try:
                    os.unlink(tmp_path)
                except:
                    pass
        except Exception as file_err:
            print(f"[Upload] Error processing file: {file_err}")
            # 嘗試清理暫存檔
            try:
                if 'tmp_path' in locals():
                    os.unlink(tmp_path)
            except:
                pass
    
    stats = Stats()
    for t in parser.tasks:
        stats.add(t)
    LAST_RESULT = stats
    LAST_DATA = stats.summary()
    LAST_MAILS_LIST = mails  # 儲存郵件列表供匯出用
    
    # 調試輸出
    print(f"[Upload] MAIL_CONTENTS has {len(MAIL_CONTENTS)} mails")
    for mid, mc in MAIL_CONTENTS.items():
        has_html = bool(mc.get('html_body'))
        html_len = len(mc.get('html_body', '')) if mc.get('html_body') else 0
        has_att_data = any(a.get('data') for a in mc.get('attachments', []))
        print(f"  - {mid}: has_html={has_html}, html_len={html_len}, has_att_data={has_att_data}")
    
    print(f"[Upload] LAST_MAILS_LIST has {len(LAST_MAILS_LIST)} mails")
    
    # 調試：檢查任務的 has_attachments
    print(f"[Upload] Tasks with attachments:")
    for t in LAST_DATA.get('all_tasks', []):
        print(f"  - {t.get('title', '')[:30]}: has_attachments={t.get('has_attachments')}")
    
    result = dict(LAST_DATA)
    result['mails'] = mails
    
    return jsonify(result)

@app.route('/api/mail/<mail_id>')
def api_mail(mail_id):
    # 如果已經有完整內容且已處理過 CID 和附件，直接返回
    cached = MAIL_CONTENTS.get(mail_id)
    if cached and cached.get('cid_processed') and cached.get('attachments') is not None:
        print(f"[api_mail] Returning cached data for {mail_id}, attachments: {len(cached.get('attachments', []))}")
        return jsonify(cached)
    
    # 如果有 entry_id，從 Outlook 讀取完整內容
    if mail_id in MAIL_ENTRIES and HAS_OUTLOOK:
        try:
            entry_info = MAIL_ENTRIES[mail_id]
            outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
            msg = outlook.GetItemFromID(entry_info['entry_id'], entry_info.get('store_id'))
            
            body = ""
            html_body = ""
            try:
                body = msg.Body or ""
            except:
                pass
            try:
                html_body = msg.HTMLBody or ""
            except:
                pass
            
            # 取得附件資訊並處理 CID 圖片
            attachments = []
            cid_images = {}  # cid -> base64 data
            try:
                if hasattr(msg, 'Attachments') and msg.Attachments.Count > 0:
                    import tempfile
                    import base64
                    import re
                    
                    print(f"[api_mail] Processing {msg.Attachments.Count} attachments")
                    
                    for j in range(1, msg.Attachments.Count + 1):
                        att = msg.Attachments.Item(j)
                        att_name = att.FileName if hasattr(att, 'FileName') else f"attachment_{j}"
                        att_size = att.Size if hasattr(att, 'Size') else 0
                        
                        # 檢查是否為嵌入圖片 (有 Content-ID)
                        content_id = ""
                        try:
                            # PropertyAccessor 取得 Content-ID
                            content_id = att.PropertyAccessor.GetProperty("http://schemas.microsoft.com/mapi/proptag/0x3712001F")
                        except:
                            pass
                        
                        # 如果是圖片且有 Content-ID，轉為 base64
                        is_image = att_name.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp', '.webp'))
                        if is_image and (content_id or 'image' in str(getattr(att, 'Type', '')).lower()):
                            try:
                                # 儲存到暫存檔再讀取
                                tmp_path = tempfile.mktemp(suffix=os.path.splitext(att_name)[1])
                                att.SaveAsFile(tmp_path)
                                with open(tmp_path, 'rb') as f:
                                    img_data = f.read()
                                os.unlink(tmp_path)
                                
                                # 轉為 base64
                                b64_data = base64.b64encode(img_data).decode('utf-8')
                                
                                # 判斷 MIME 類型
                                ext = os.path.splitext(att_name)[1].lower()
                                mime_types = {'.png': 'image/png', '.jpg': 'image/jpeg', '.jpeg': 'image/jpeg', '.gif': 'image/gif', '.bmp': 'image/bmp', '.webp': 'image/webp'}
                                mime_type = mime_types.get(ext, 'image/png')
                                
                                # 儲存 CID 映射
                                cid_key = content_id.strip('<>') if content_id else att_name
                                cid_images[cid_key] = f"data:{mime_type};base64,{b64_data}"
                                
                            except Exception as img_err:
                                print(f"[api_mail] Error processing image {att_name}: {img_err}")
                        
                        attachments.append({
                            "index": j,
                            "name": att_name,
                            "size": att_size,
                            "content_id": content_id
                        })
                    
                    # 替換 HTML 中的 cid: 連結
                    if cid_images and html_body:
                        for cid, data_url in cid_images.items():
                            # 替換 cid:xxx 格式
                            html_body = re.sub(f'src=["\']cid:{re.escape(cid)}["\']', f'src="{data_url}"', html_body, flags=re.IGNORECASE)
                            # 也嘗試替換檔名
                            html_body = re.sub(f'src=["\']cid:{re.escape(cid.split("@")[0] if "@" in cid else cid)}["\']', f'src="{data_url}"', html_body, flags=re.IGNORECASE)
                        
                        print(f"[api_mail] Replaced {len(cid_images)} CID images")
            except Exception as att_err:
                print(f"[api_mail] Error processing attachments: {att_err}")
            
            mail_time = None
            try:
                mail_time = msg.ReceivedTime
            except:
                try:
                    mail_time = msg.SentOn
                except:
                    pass
            
            mail_data = {
                "subject": msg.Subject or "",
                "body": body,
                "html_body": html_body,
                "date": mail_time.strftime("%Y-%m-%d") if mail_time else "",
                "time": mail_time.strftime("%H:%M") if mail_time else "",
                "attachments": attachments,
                "cid_processed": True
            }
            
            print(f"[api_mail] Returning data for {mail_id}, attachments: {len(attachments)}")
            for att in attachments:
                print(f"  - {att['name']} ({att['size']} bytes)")
            
            # 快取供下次使用
            MAIL_CONTENTS[mail_id] = mail_data
            
            return jsonify(mail_data)
        except Exception as e:
            print(f"[api_mail] Error reading from Outlook: {e}")
    
    # 返回已有的部分資料或錯誤
    if mail_id in MAIL_CONTENTS:
        return jsonify(MAIL_CONTENTS[mail_id])
    
    return jsonify({'error': 'Mail not found'}), 404

@app.route('/api/review-mails', methods=['POST'])
def api_review_mails():
    """Review 模式專用 - 只載入郵件列表，不做分析，支援分頁"""
    global MAIL_CONTENTS, MAIL_ENTRIES
    
    if not HAS_OUTLOOK:
        return jsonify({'error': 'Outlook not available'}), 500
    
    data = request.json or {}
    entry_id = data.get('entry_id')
    store_id = data.get('store_id')
    start_date = data.get('start')
    end_date = data.get('end')
    offset = data.get('offset', 0)  # 分頁偏移
    limit = data.get('limit', 100)  # 每頁筆數
    
    print(f"[Review] Request: start={start_date}, end={end_date}, offset={offset}, limit={limit}")
    
    if not entry_id:
        return jsonify({'error': 'No folder selected'}), 400
    
    try:
        outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
        folder = outlook.GetFolderFromID(entry_id, store_id) if store_id else outlook.GetFolderFromID(entry_id)
        
        items = folder.Items
        items.Sort("[ReceivedTime]", True)  # 降序排列
        
        print(f"[Review] Folder has {items.Count} items before filter")
        
        # 日期篩選 - 使用正確的 Outlook 日期格式
        if start_date and end_date:
            try:
                # Outlook Restrict 需要 MM/DD/YYYY 格式
                from datetime import datetime
                start_dt = datetime.strptime(start_date, "%Y-%m-%d")
                end_dt = datetime.strptime(end_date, "%Y-%m-%d")
                start_fmt = start_dt.strftime("%m/%d/%Y")
                end_fmt = end_dt.strftime("%m/%d/%Y")
                filter_str = f"[ReceivedTime] >= '{start_fmt}' AND [ReceivedTime] <= '{end_fmt} 11:59 PM'"
                print(f"[Review] Date filter: {filter_str}")
                items = items.Restrict(filter_str)
                print(f"[Review] After filter: {items.Count} items")
            except Exception as e:
                print(f"[Review] Restrict failed: {e}, using all items")
        
        # 取得總數
        try:
            total_count = items.Count
        except:
            total_count = 0
            
        mails = []
        
        if total_count == 0:
            return jsonify({
                'mails': [],
                'total': 0,
                'offset': offset,
                'limit': limit,
                'has_more': False
            })
        
        # 分頁載入
        start_idx = offset + 1  # Outlook 是 1-based
        end_idx = min(offset + limit, total_count)
        
        import hashlib
        error_count = 0
        for i in range(start_idx, end_idx + 1):
            try:
                msg = items.Item(i)
                if not hasattr(msg, 'Subject'):
                    continue
                
                # 取得時間，處理可能的錯誤
                mail_date_str = ""
                mail_time_str = ""
                try:
                    mail_time = msg.ReceivedTime
                    if mail_time:
                        mail_date_str = mail_time.strftime("%Y-%m-%d")
                        mail_time_str = mail_time.strftime("%H:%M")
                except Exception as time_err:
                    # 嘗試其他時間欄位
                    try:
                        mail_time = msg.SentOn
                        if mail_time:
                            mail_date_str = mail_time.strftime("%Y-%m-%d")
                            mail_time_str = mail_time.strftime("%H:%M")
                    except:
                        pass
                
                # 生成 mail_id
                mail_id = hashlib.md5(f"{mail_date_str}_{mail_time_str}_{msg.Subject or ''}".encode()).hexdigest()[:12]
                
                # 延遲讀取 body 和 html_body - 只在需要時才讀取
                # 這裡只記錄基本資訊，實際內容在 /api/mail/<id> 時讀取
                
                # 儲存 entry_id 用於後續讀取完整內容和下載附件
                try:
                    MAIL_ENTRIES[mail_id] = {
                        "entry_id": msg.EntryID,
                        "store_id": store_id
                    }
                except:
                    pass
                
                # 快速檢查是否有附件（不讀取附件詳細資訊）
                attachment_count = 0
                try:
                    if hasattr(msg, 'Attachments'):
                        attachment_count = msg.Attachments.Count
                except:
                    pass
                
                # 取得收件者
                recipient = ""
                try:
                    if hasattr(msg, 'To'):
                        recipient = str(msg.To)
                except:
                    pass
                
                mails.append({
                    "mail_id": mail_id,
                    "subject": msg.Subject or "(無主旨)",
                    "date": mail_date_str,
                    "time": mail_time_str,
                    "sender": str(msg.SenderName) if hasattr(msg, 'SenderName') else "",
                    "recipient": recipient,
                    "attachment_count": attachment_count
                })
            except Exception as item_err:
                error_count += 1
                if error_count <= 3:  # 只顯示前3個錯誤
                    print(f"Error reading item {i}: {item_err}")
                elif error_count == 4:
                    print(f"... (更多錯誤省略)")
                continue
        
        return jsonify({
            'mails': mails,
            'total': total_count,
            'offset': offset,
            'limit': limit,
            'has_more': end_idx < total_count
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/folder-mails', methods=['POST'])
def api_folder_mails():
    """直接載入資料夾郵件（不套用日期篩選）- 用於點擊資料夾"""
    global MAIL_CONTENTS, MAIL_ENTRIES
    
    if not HAS_OUTLOOK:
        return jsonify({'error': 'Outlook not available'}), 500
    
    data = request.json or {}
    entry_id = data.get('entry_id')
    store_id = data.get('store_id')
    offset = data.get('offset', 0)
    limit = data.get('limit', 100)
    
    print(f"[Folder] Direct load: offset={offset}, limit={limit}")
    
    if not entry_id:
        return jsonify({'error': 'No folder selected'}), 400
    
    try:
        outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
        folder = outlook.GetFolderFromID(entry_id, store_id) if store_id else outlook.GetFolderFromID(entry_id)
        
        items = folder.Items
        items.Sort("[ReceivedTime]", True)  # 降序排列（最新的在前）
        
        # 不套用日期篩選，直接取得所有郵件
        try:
            total_count = items.Count
        except:
            total_count = 0
            
        print(f"[Folder] Total items: {total_count}")
        
        mails = []
        
        if total_count == 0:
            return jsonify({
                'mails': [],
                'total': 0,
                'offset': offset,
                'limit': limit,
                'has_more': False
            })
        
        # 分頁載入
        start_idx = offset + 1  # Outlook 是 1-based
        end_idx = min(offset + limit, total_count)
        
        import hashlib
        error_count = 0
        for i in range(start_idx, end_idx + 1):
            try:
                msg = items.Item(i)
                if not hasattr(msg, 'Subject'):
                    continue
                
                mail_date_str = ""
                mail_time_str = ""
                try:
                    mail_time = msg.ReceivedTime
                    if mail_time:
                        mail_date_str = mail_time.strftime("%Y-%m-%d")
                        mail_time_str = mail_time.strftime("%H:%M")
                except:
                    try:
                        mail_time = msg.SentOn
                        if mail_time:
                            mail_date_str = mail_time.strftime("%Y-%m-%d")
                            mail_time_str = mail_time.strftime("%H:%M")
                    except:
                        pass
                
                mail_id = hashlib.md5(f"{mail_date_str}_{mail_time_str}_{msg.Subject or ''}".encode()).hexdigest()[:12]
                
                try:
                    MAIL_ENTRIES[mail_id] = {
                        "entry_id": msg.EntryID,
                        "store_id": store_id
                    }
                except:
                    pass
                
                # 取得附件資訊（包含檔名，用於顯示圖示）
                attachments = []
                try:
                    if hasattr(msg, 'Attachments') and msg.Attachments.Count > 0:
                        for j in range(1, msg.Attachments.Count + 1):
                            try:
                                att = msg.Attachments.Item(j)
                                att_name = str(att.FileName) if hasattr(att, 'FileName') else f"attachment_{j}"
                                attachments.append({
                                    "index": j,
                                    "name": att_name
                                })
                            except:
                                pass
                except:
                    pass
                
                recipient = ""
                try:
                    if hasattr(msg, 'To'):
                        recipient = str(msg.To)
                except:
                    pass
                
                mails.append({
                    "mail_id": mail_id,
                    "subject": msg.Subject or "(無主旨)",
                    "date": mail_date_str,
                    "time": mail_time_str,
                    "sender": str(msg.SenderName) if hasattr(msg, 'SenderName') else "",
                    "recipient": recipient,
                    "attachment_count": len(attachments),
                    "attachments": attachments
                })
            except Exception as item_err:
                error_count += 1
                if error_count <= 3:
                    print(f"Error reading item {i}: {item_err}")
                continue
        
        print(f"[Folder] Loaded {len(mails)} mails")
        
        return jsonify({
            'mails': mails,
            'total': total_count,
            'offset': offset,
            'limit': limit,
            'has_more': end_idx < total_count
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


# 附件資訊 API（需要另外處理實際下載）
@app.route('/api/mail/<mail_id>/attachments')
def api_mail_attachments(mail_id):
    if mail_id in MAIL_CONTENTS:
        return jsonify(MAIL_CONTENTS[mail_id].get('attachments', []))
    return jsonify([])

# 附件下載 API
@app.route('/api/mail/<mail_id>/attachment/<int:att_index>')
def api_download_attachment(mail_id, att_index):
    """下載郵件附件"""
    if not HAS_OUTLOOK:
        return jsonify({'error': 'Outlook not available'}), 500
    
    if mail_id not in MAIL_ENTRIES:
        return jsonify({'error': 'Mail entry not found'}), 404
    
    entry_info = MAIL_ENTRIES[mail_id]
    
    try:
        outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
        msg = outlook.GetItemFromID(entry_info['entry_id'], entry_info.get('store_id'))
        
        if not hasattr(msg, 'Attachments') or msg.Attachments.Count < att_index:
            return jsonify({'error': 'Attachment not found'}), 404
        
        att = msg.Attachments.Item(att_index)
        filename = att.FileName if hasattr(att, 'FileName') else f"attachment_{att_index}"
        
        # 儲存到暫存檔案
        temp_path = os.path.join(tempfile.gettempdir(), f"att_{mail_id}_{att_index}_{filename}")
        att.SaveAsFile(temp_path)
        
        # 讀取並返回
        with open(temp_path, 'rb') as f:
            content = f.read()
        
        # 清理暫存檔
        try:
            os.unlink(temp_path)
        except:
            pass
        
        # 判斷 MIME 類型
        import mimetypes
        mime_type, _ = mimetypes.guess_type(filename)
        if not mime_type:
            mime_type = 'application/octet-stream'
        
        response = Response(content, mimetype=mime_type)
        response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/excel')
def api_excel():
    if not LAST_RESULT:
        return jsonify({'error': 'No data'}), 400
    return send_file(LAST_RESULT.excel(), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=f'task_report_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx')

@app.route('/api/export-html')
def api_export_html():
    global LAST_DATA, MAIL_CONTENTS, LAST_MAILS_LIST, MAIL_ENTRIES
    
    if not LAST_DATA:
        return "請先執行分析", 400
    
    # 調試輸出
    print(f"[Export HTML] MAIL_CONTENTS has {len(MAIL_CONTENTS)} mails")
    print(f"[Export HTML] LAST_MAILS_LIST has {len(LAST_MAILS_LIST)} mails")
    print(f"[Export HTML] MAIL_ENTRIES has {len(MAIL_ENTRIES)} entries")
    
    # 收集所有郵件的完整內容（包含 CID 處理後的 html_body 和附件 Base64）
    mail_contents_with_attachments = {}
    
    # 遍歷所有郵件，確保每封都有經過 CID 處理的 html_body
    for mail in LAST_MAILS_LIST:
        mail_id = mail.get('mail_id')
        if not mail_id:
            continue
        
        # 檢查是否已經有處理過的資料（來自 MAIL_CONTENTS 快取）
        cached = MAIL_CONTENTS.get(mail_id)
        if cached and cached.get('cid_processed'):
            # 使用快取的資料（已處理 CID）
            mail_contents_with_attachments[mail_id] = dict(cached)
            print(f"[Export HTML] Mail {mail_id}: using cached data (cid_processed=True)")
            continue
        
        # 檢查 mail 本身是否已處理過 CID（來自上傳）
        if mail.get('cid_processed'):
            mail_contents_with_attachments[mail_id] = {
                "subject": mail.get('subject', ''),
                "body": mail.get('body', ''),
                "html_body": mail.get('html_body', ''),
                "date": mail.get('date', ''),
                "time": mail.get('time', ''),
                "attachments": mail.get('attachments', []),
                "cid_processed": True
            }
            print(f"[Export HTML] Mail {mail_id}: from LAST_MAILS_LIST (cid_processed=True)")
            continue
        
        # 需要從 Outlook 重新讀取並處理 CID
        if mail_id in MAIL_ENTRIES and HAS_OUTLOOK:
            try:
                entry_info = MAIL_ENTRIES[mail_id]
                outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
                msg = outlook.GetItemFromID(entry_info['entry_id'], entry_info.get('store_id'))
                
                body = ""
                html_body = ""
                try:
                    body = msg.Body or ""
                except:
                    pass
                try:
                    html_body = msg.HTMLBody or ""
                except:
                    pass
                
                # 處理 CID 圖片和附件
                cid_images = {}
                attachments_with_data = []
                
                if hasattr(msg, 'Attachments') and msg.Attachments.Count > 0:
                    import base64
                    import re
                    import mimetypes
                    
                    for j in range(1, msg.Attachments.Count + 1):
                        att = msg.Attachments.Item(j)
                        att_name = att.FileName if hasattr(att, 'FileName') else f"attachment_{j}"
                        att_size = att.Size if hasattr(att, 'Size') else 0
                        
                        # 檢查 Content-ID
                        content_id = ""
                        try:
                            content_id = att.PropertyAccessor.GetProperty("http://schemas.microsoft.com/mapi/proptag/0x3712001F")
                        except:
                            pass
                        
                        # 儲存並讀取 Base64
                        att_b64 = ""
                        att_mime = "application/octet-stream"
                        try:
                            temp_path = os.path.join(tempfile.gettempdir(), f"export_att_{mail_id}_{j}")
                            att.SaveAsFile(temp_path)
                            with open(temp_path, 'rb') as f:
                                att_b64 = base64.b64encode(f.read()).decode('utf-8')
                            os.unlink(temp_path)
                            att_mime, _ = mimetypes.guess_type(att_name)
                            if not att_mime:
                                att_mime = 'application/octet-stream'
                        except Exception as att_err:
                            print(f"[Export HTML] Error reading attachment {att_name}: {att_err}")
                        
                        # 處理 CID 圖片
                        is_image = att_name.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp', '.webp'))
                        if is_image and content_id and att_b64:
                            cid_key = content_id.strip('<>') if content_id else att_name
                            cid_images[cid_key] = f"data:{att_mime};base64,{att_b64}"
                        
                        attachments_with_data.append({
                            'name': att_name,
                            'size': att_size,
                            'data': att_b64,
                            'mime': att_mime
                        })
                    
                    # 替換 HTML 中的 cid: 連結
                    if cid_images and html_body:
                        for cid, data_url in cid_images.items():
                            html_body = re.sub(f'src=["\']cid:{re.escape(cid)}["\']', f'src="{data_url}"', html_body, flags=re.IGNORECASE)
                            html_body = re.sub(f'src=["\']cid:{re.escape(cid.split("@")[0] if "@" in cid else cid)}["\']', f'src="{data_url}"', html_body, flags=re.IGNORECASE)
                        print(f"[Export HTML] Mail {mail_id}: replaced {len(cid_images)} CID images")
                
                mail_time = None
                try:
                    mail_time = msg.ReceivedTime
                except:
                    try:
                        mail_time = msg.SentOn
                    except:
                        pass
                
                mail_contents_with_attachments[mail_id] = {
                    "subject": msg.Subject or mail.get('subject', ''),
                    "body": body,
                    "html_body": html_body,
                    "date": mail_time.strftime("%Y-%m-%d") if mail_time else mail.get('date', ''),
                    "time": mail_time.strftime("%H:%M") if mail_time else mail.get('time', ''),
                    "attachments": attachments_with_data,
                    "cid_processed": True
                }
                print(f"[Export HTML] Mail {mail_id}: loaded from Outlook with CID processing")
                
            except Exception as e:
                print(f"[Export HTML] Error loading mail {mail_id}: {e}")
                # 使用原始資料（沒有 CID 處理）
                mail_contents_with_attachments[mail_id] = {
                    "subject": mail.get('subject', ''),
                    "body": mail.get('body', ''),
                    "html_body": mail.get('html_body', ''),
                    "date": mail.get('date', ''),
                    "time": mail.get('time', ''),
                    "attachments": mail.get('attachments', []),
                    "cid_processed": False
                }
        else:
            # 沒有 MAIL_ENTRIES，使用原始資料
            mail_contents_with_attachments[mail_id] = {
                "subject": mail.get('subject', ''),
                "body": mail.get('body', ''),
                "html_body": mail.get('html_body', ''),
                "date": mail.get('date', ''),
                "time": mail.get('time', ''),
                "attachments": mail.get('attachments', []),
                "cid_processed": False
            }
            print(f"[Export HTML] Mail {mail_id}: using original data (no MAIL_ENTRIES)")
    
    print(f"[Export HTML] Total mails in export: {len(mail_contents_with_attachments)}")
    for mid, mc in mail_contents_with_attachments.items():
        has_html = bool(mc.get('html_body'))
        html_len = len(mc.get('html_body', '')) if mc.get('html_body') else 0
        cid_proc = mc.get('cid_processed', False)
        atts = mc.get('attachments', [])
        has_att_data = any(a.get('data') for a in atts) if atts else False
        print(f"  - {mid}: html_len={html_len}, cid_processed={cid_proc}, atts={len(atts)}, has_att_data={has_att_data}")
    
    # 使用主頁面模板，但注入預載數據
    import json
    report_date = datetime.now().strftime("%Y-%m-%d %H:%M")
    
    print(f"[Export HTML] LAST_MAILS_LIST has {len(LAST_MAILS_LIST)} mails for Review tab")
    
    # 生成完整 HTML（包含統計分析和 Review 頁籤）
    html = generate_export_html(LAST_DATA, report_date, mail_contents_with_attachments, LAST_MAILS_LIST)
    
    return Response(html, mimetype='text/html', headers={'Content-Disposition': f'attachment; filename=task_report_{datetime.now().strftime("%Y%m%d_%H%M")}.html'})

if __name__ == '__main__':
    print("=" * 50)
    print("Task Dashboard v23")
    print("=" * 50)
    load_folders()
    print("開啟: http://127.0.0.1:5000")
    print("=" * 50)
    from werkzeug.serving import run_simple
    run_simple('127.0.0.1', 5000, app, use_reloader=False, threaded=False)
